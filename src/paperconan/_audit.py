#!/usr/bin/env python3
"""
paper_audit.py — scan a paper's published source data (xlsx) for statistical signals.

Usage:
    python3 paper_audit.py <dir-with-xlsx-files> [--out OUT_DIR]

Outputs to <OUT_DIR or <dir>/audit>:
  - scan.json   structured findings (every block, every detector)
  - REPORT.md   ranked top-5 + supporting evidence in markdown

What it detects (numeric patterns requiring contextual review):
  1. Identical / constant-offset / constant-ratio / exact-linear column relations
  2. Arithmetic-progression columns (constant first difference)
  3. Repeated last-two-decimal endings beyond chance
  4. Last-digit chi-square (true measurements have ~uniform last digits)
  5. Suspicious row pairs that sum to integer / equal-value column pairs
  6. Reverse-engineered generation rules (col_b = col_a + k, col_b = K - col_a, etc.)

Dependencies: openpyxl, numpy, scipy
"""
from __future__ import annotations
import argparse
import csv as _csv
import datetime
import glob
import json
import math
import os
import re
import sys
import time
from collections import Counter, defaultdict
from fractions import Fraction
from pathlib import Path

import openpyxl
import numpy as np
from scipy import stats

from ._profiles import apply_profile_to_findings, normalize_profile
from ._sheet import Sheet
from .schema import PaperconanInputError

# Canonical list of the per-block finding-group keys emitted into every
# `relations_blocks[]` entry (see scan_dir's report_blocks.append). This is the
# SINGLE SOURCE OF TRUTH: the markdown report, the packet distiller, and the
# paperconan-watch severity counters / triage gate all iterate this set, so a
# HIGH finding in ANY group (notably row_pairs) is counted and can reach review.
BLOCK_FINDING_GROUPS = (
    "relations", "equal_pairs", "progressions", "row_pairs", "row_relations",
    "within_col", "identical_after_rounding", "grim",
)


def _version():
    """paperconan version, resolved lazily to avoid an import cycle with __init__."""
    try:
        from . import __version__
        return __version__
    except Exception:
        return "unknown"


# ---------- value helpers ----------

def is_num(x):
    if x is None or isinstance(x, bool):
        return False
    if isinstance(x, (int, float)):
        return not (isinstance(x, float) and (math.isnan(x) or math.isinf(x)))
    return False


def to_float(x):
    return float(x) if is_num(x) else None


def last_significant_digit(x):
    if x is None or x == 0:
        return None
    s = f"{x:.10g}"
    digits = [c for c in s if c.isdigit()]
    return digits[-1] if digits else None


def trailing_decimal_digits(x, k=2):
    if x is None:
        return None
    try:
        s = repr(float(x))
    except (TypeError, ValueError):
        return None
    if "e" in s or "E" in s or "." not in s:
        return None
    frac = s.split(".", 1)[1]
    return frac[-k:] if len(frac) >= k else None


def _decimals_of(x, cap=6):
    """Number of significant decimal places in x's shortest float repr, capped.

    Cells are coerced to float on load, so displayed trailing zeros are lost.
    Recovering decimals from the float repr therefore UNDER-counts precision for
    values like 2.50 -> 2.5. That is conservatively safe for GRIM: fewer decimals
    means a coarser grid and fewer flags, never a false flag."""
    s = repr(float(x))
    if "e" in s or "E" in s:
        return cap  # scientific notation: assume high precision (conservative)
    if "." not in s:
        return 0
    frac = s.split(".", 1)[1].rstrip("0")
    return min(len(frac), cap)


def grim_consistent(mean, n, decimals):
    """True if `mean`, reported to `decimals` places, is achievable as an integer
    total divided by `n`. Conservative: any bracketing integer total that rounds
    back to the reported mean counts as consistent (tolerant of the rounding
    convention used by the authors)."""
    if n <= 0:
        return True
    scale = 10 ** decimals
    target = round(mean * scale)
    base = mean * n
    for t in (math.floor(base), math.ceil(base), round(base)):
        if round((t / n) * scale) == target:
            return True
    return False


def grimmer_consistent(mean, sd, n, mean_decimals, sd_decimals):
    """True if a sample of `n` integers can have both the reported `mean` and the
    reported `sd` (to their stated decimals). Implements the GRIMMER test: for the
    integer total T fixed by the mean, search the integer sum-of-squares values
    whose implied sd rounds to the reported sd, and require one with the correct
    parity (since sum(x^2) == sum(x) mod 2 for integers). Accepts either sample
    (n-1) or population (n) SD convention so an unknown convention never
    false-positives."""
    if n <= 1 or sd < 0:
        return True
    T = round(mean * n)
    half = 0.5 / (10 ** sd_decimals)
    lo_sd = max(0.0, sd - half)
    hi_sd = sd + half
    for ddof in (1, 0):
        denom = n - ddof
        if denom <= 0:
            continue
        corr = (T * T) / n
        ss_lo = lo_sd * lo_sd * denom + corr
        ss_hi = hi_sd * hi_sd * denom + corr
        for ss in range(math.ceil(ss_lo - 1e-9), math.floor(ss_hi + 1e-9) + 1):
            if ss < 0:
                continue
            if (ss % 2) != (T % 2):       # integer parity test
                continue
            if ss + 1e-9 >= corr:          # variance >= 0
                return True
    return False


# ---------- sheet I/O ----------

def _fill_sheet_from_rows(rows_iter, mr, mc, loaded):
    """Stream rows of openpyxl-shaped cell values (int/float/str/datetime/bool/None)
    into a Sheet, honouring the cumulative `_MAX_CELLS` budget that `loaded` cells
    have already consumed across this file.

    Returns (sheet_or_None, cells): None means the per-file cumulative budget was
    exceeded mid-stream (oversized). Both readers (openpyxl, calamine) funnel through
    this so they produce a byte-identical Sheet; the calamine path normalizes its
    typed values to openpyxl's shape BEFORE calling here.

    The produced Sheet matches Sheet.from_rows of the same rows: nrows == rows
    consumed, ncols == max row width seen (trailing all-empty rows/cols are kept as
    NaN padding, not trimmed). `mr`/`mc` are only the pre-allocation hint; the array
    grows on demand if a reader under-declares dimensions."""
    numeric = np.full((mr, mc), np.nan, dtype=float) if (mr and mc) else np.empty((0, 0))
    text = {}
    ints = set()
    r = 0                                        # rows consumed (== final nrows)
    cells = 0
    max_w = 0                                    # max row width seen (== final ncols)
    oversized = False
    for row in rows_iter:
        if r >= numeric.shape[0]:                # reader under-reported rows: grow by one
            grow = np.full((1, numeric.shape[1]), np.nan)
            numeric = np.vstack([numeric, grow]) if numeric.size or numeric.shape[1] else grow
        width = len(row)
        if width > numeric.shape[1]:             # row wider than declared: grow columns
            pad = np.full((numeric.shape[0], width - numeric.shape[1]), np.nan)
            numeric = np.hstack([numeric, pad]) if numeric.shape[1] else pad
        for c, v in enumerate(row):
            if is_num(v):
                numeric[r, c] = float(v)
                if isinstance(v, int) and not isinstance(v, bool):
                    ints.add((r, c))
            elif v is not None:
                text[(r, c)] = v
        if width > max_w:
            max_w = width
        cells += width
        if loaded + cells > _MAX_CELLS:          # per-file cumulative budget — bail mid-stream
            oversized = True
            break
        r += 1
    if oversized:
        return None, cells
    # Trim to the geometry Sheet.from_rows would produce: nrows == rows consumed,
    # ncols == max(len(row)). (numeric may be larger if the reader over-declared.)
    n_rows, n_cols = r, max_w
    if n_rows and n_cols:
        numeric = numeric[:n_rows, :n_cols]
    else:
        numeric = np.full((n_rows, n_cols), np.nan, dtype=float)
    text = {(rr, cc): val for (rr, cc), val in text.items()
            if rr < n_rows and cc < n_cols}
    ints = {(rr, cc) for (rr, cc) in ints if rr < n_rows and cc < n_cols}
    return Sheet(numeric.shape[0], numeric.shape[1], numeric, text, ints), cells


def _load_workbook_openpyxl(path):
    """Return dict of sheet_name -> Sheet via openpyxl (the reference reader). A sheet
    over _MAX_CELLS (on its own, or once this file's cumulative cell budget is spent)
    is returned as None (oversized), preserving the legacy memory guard. Rows stream
    directly into the Sheet's columnar arrays — the full list-of-lists is never
    materialized."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {}
    loaded = 0                                       # cumulative cells across this file's sheets
    for s in wb.sheetnames:
        ws = wb[s]
        mr, mc = ws.max_row or 0, ws.max_column or 0
        # Skip a sheet that is too big on its own, OR once this file's cumulative cell budget is
        # spent (a many-sheet workbook materialized at once OOMs even if each sheet is under cap).
        if loaded >= _MAX_CELLS or (mr and mc and mr * mc > _MAX_CELLS):
            out[s] = None
            continue
        sheet, cells = _fill_sheet_from_rows(ws.iter_rows(values_only=True), mr, mc, loaded)
        out[s] = sheet
        if sheet is not None:
            loaded += cells
    wb.close()
    return out


def _calamine_cell(v):
    """Normalize one python_calamine typed value to the shape openpyxl's read_only
    reader produces, so a Sheet built from calamine rows is byte-identical:
      - "" (calamine's empty cell) -> None (openpyxl yields None for empty cells)
      - whole-number float -> int (openpyxl coerces every integral value to int)
      - datetime.date -> datetime.datetime at midnight (openpyxl never yields bare date)
    bool / str / datetime / non-integral float pass through unchanged."""
    if isinstance(v, bool):
        return v
    if v == "":
        return None
    if isinstance(v, float):
        if math.isfinite(v) and v == int(v):
            return int(v)
        return v
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    return v


def _load_workbook_calamine(path):
    """Return dict of sheet_name -> Sheet via python-calamine (a fast Rust reader),
    producing a Sheet byte-identical to _load_workbook_openpyxl. Same _MAX_CELLS
    per-sheet + cumulative guard, same oversized->None, same trim-to-max-width."""
    import python_calamine
    wb = python_calamine.CalamineWorkbook.from_path(path)
    out = {}
    loaded = 0                                       # cumulative cells across this file's sheets
    for name in wb.sheet_names:
        sh = wb.get_sheet_by_name(name)
        # Reject from the cheap DECLARED dimensions BEFORE to_python materializes the
        # full bounding box: a sheet that declares e.g. C1000000 would otherwise
        # allocate millions of cells just to be discarded — that materialization is
        # what OOMs in prod. `height` × `width` matches to_python(skip_empty_area=False).
        h, w = sh.height, sh.width
        if loaded >= _MAX_CELLS or (h and w and h * w > _MAX_CELLS):
            out[name] = None
            continue
        rows = sh.to_python(skip_empty_area=False)
        mr = len(rows)
        mc = max((len(row) for row in rows), default=0)
        if loaded >= _MAX_CELLS or (mr and mc and mr * mc > _MAX_CELLS):
            out[name] = None
            continue
        norm = ([_calamine_cell(v) for v in row] for row in rows)
        sheet, cells = _fill_sheet_from_rows(norm, mr, mc, loaded)
        out[name] = sheet
        if sheet is not None:
            loaded += cells
    return out


def load_workbook_rows(path):
    """Return dict of sheet_name -> Sheet. Uses python-calamine (a fast Rust xlsx
    reader) when installed, falling back to the openpyxl reference path otherwise or
    on any reader quirk. Both paths produce a byte-identical Sheet."""
    try:
        import python_calamine  # noqa: F401
    except Exception:
        return _load_workbook_openpyxl(path)
    try:
        return _load_workbook_calamine(path)
    except Exception:
        return _load_workbook_openpyxl(path)  # any reader quirk → reference path


def _coerce_cell(s):
    """Parse a CSV string cell into int / float / text. Empty -> None.
    Deliberately conservative: no thousands separators, no percent, no currency."""
    if s is None:
        return None
    s = s.strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def load_csv_rows(path, delimiter):
    """Load a delimited text file as {sheet_name: Sheet|None}, mirroring load_workbook_rows.
    A flat file has no sheets, so it becomes a single sheet named after the file stem.
    Oversized (> _MAX_CELLS) -> {stem: None}; otherwise the rows are wrapped in a Sheet."""
    stem = os.path.splitext(os.path.basename(path))[0]
    rows = []
    oversized = False
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            rows = []
            cells = 0
            with open(path, newline="", encoding=enc) as fh:
                for r in _csv.reader(fh, delimiter=delimiter):
                    rows.append([_coerce_cell(c) for c in r])
                    cells += len(r)
                    if cells > _MAX_CELLS:           # oversized: stop before exhausting memory
                        oversized = True
                        break
            break
        except UnicodeDecodeError:
            continue
    if oversized:
        return {stem: None}
    maxc = max((len(r) for r in rows), default=0)
    for r in rows:
        if len(r) < maxc:
            r.extend([None] * (maxc - len(r)))
    return {stem: Sheet.from_rows(rows)}


def load_table(path):
    """Dispatch by extension to a {sheet_name: Sheet|None} loader."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".tsv":
        return load_csv_rows(path, delimiter="\t")
    if ext == ".csv":
        return load_csv_rows(path, delimiter=",")
    if ext == ".pdf":
        from ._extract import load_pdf_tables
        return {k: (None if v is None else Sheet.from_rows(v)) for k, v in load_pdf_tables(path).items()}
    if ext == ".docx":
        from ._extract import load_docx_tables
        return {k: (None if v is None else Sheet.from_rows(v)) for k, v in load_docx_tables(path).items()}
    return load_workbook_rows(path)


def find_numeric_blocks(sheet, min_rows=3, min_cols=1):
    R, C = sheet.nrows, sheet.ncols
    if R == 0 or C == 0:
        return []
    num = ~np.isnan(sheet.numeric)
    blocks = []
    visited = np.zeros_like(num)
    for j in range(C):
        i = 0
        while i < R:
            if num[i, j] and not visited[i, j]:
                i0 = i
                while i < R and num[i, j]:
                    i += 1
                i1 = i
                j1 = j + 1
                while j1 < C:
                    col_density = num[i0:i1, j1].mean() if i1 > i0 else 0
                    if col_density >= 0.7:
                        j1 += 1
                    else:
                        break
                visited[i0:i1, j:j1] = True
                if (i1 - i0) >= min_rows and (j1 - j) >= min_cols:
                    blocks.append((i0, i1, j, j1))
            else:
                i += 1
    return blocks


def header_for(sheet, r0, c0, c1):
    for r in range(r0 - 1, max(-1, r0 - 5), -1):
        if r < 0:
            continue
        line = [sheet.cell(r, c) for c in range(c0, c1)]
        texty = [x for x in line if x is not None and not is_num(x)]
        if texty:
            return [str(sheet.cell(r, c)).strip() if sheet.cell(r, c) is not None else ""
                    for c in range(c0, c1)]
    return [""] * (c1 - c0)


def col_array(sheet, r0, r1, c):
    return sheet.numeric[r0:r1, c].copy()


def _sample(arr, k=8):
    """A tiny value peek for downstream LLM triage: the first <=k finite numeric
    values of `arr` as built-in floats rounded to 6 significant figures. Bounded to
    <=k elements so it CANNOT reintroduce the evidence-bloat OOM (~64 bytes here)."""
    out = []
    for v in arr[:k]:
        fv = float(v)
        if math.isnan(fv) or math.isinf(fv):
            continue
        out.append(round(fv, 6))
    return out


# ---------- evidence helpers ----------

def _cell_value(v):
    """JSON-serializable cell value: keep numbers as-is, stringify dates/objects, None stays None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    return str(v)


def _block_evidence(sheet, r0, r1, c0, c1, header, highlight_cols, highlight_rows=None):
    """Slice a numeric block (with 1 row of context above/below if available) into a
    JSON-friendly evidence dict that the HTML renderer can show as a table.

    The emitted snippet is bounded to a contiguous _MAX_EV_ROWS × _MAX_EV_COLS
    sub-rectangle inside the block, always covering the highlighted columns (and rows
    when given). This stops a dense block from being copied whole into every finding
    (which balloons the scan dict / scan.json to GBs). Small blocks are emitted whole
    and stay byte-identical (no `truncated` key)."""
    truncated = False

    # --- column window -------------------------------------------------------
    ec0, ec1 = c0, c1
    if (c1 - c0) > _MAX_EV_COLS:
        truncated = True
        if highlight_cols:
            lo = min(highlight_cols)
            hi = max(highlight_cols)
        else:
            lo = hi = c0
        if hi - lo + 1 > _MAX_EV_COLS:
            ec0 = lo
            ec1 = lo + _MAX_EV_COLS
        else:
            # Center a _MAX_EV_COLS-wide window on [lo, hi], then clamp into [c0, c1).
            pad = (_MAX_EV_COLS - (hi - lo + 1)) // 2
            ec0 = lo - pad
            ec1 = ec0 + _MAX_EV_COLS
            if ec0 < c0:
                ec0, ec1 = c0, c0 + _MAX_EV_COLS
            if ec1 > c1:
                ec1 = c1
                ec0 = ec1 - _MAX_EV_COLS
            if ec0 < c0:
                ec0 = c0

    # --- row window ----------------------------------------------------------
    r_start = max(0, r0 - 1)
    r_end = min(sheet.nrows, r1 + 1)
    if (r_end - r_start) > _MAX_EV_ROWS:
        truncated = True
        if highlight_rows:
            # highlight_rows are 1-based row numbers; center the window on them.
            rlo = min(highlight_rows) - 1
            rhi = max(highlight_rows) - 1
            if rhi - rlo + 1 >= _MAX_EV_ROWS:
                wr0 = rlo
            else:
                pad = (_MAX_EV_ROWS - (rhi - rlo + 1)) // 2
                wr0 = rlo - pad
        else:
            wr0 = r_start
        if wr0 < r_start:
            wr0 = r_start
        wr1 = wr0 + _MAX_EV_ROWS
        if wr1 > r_end:
            wr1 = r_end
            wr0 = max(r_start, wr1 - _MAX_EV_ROWS)
        r_start, r_end = wr0, wr1

    data_rows = []
    for r in range(r_start, r_end):
        vals = [_cell_value(sheet.cell(r, c)) for c in range(ec0, ec1)]
        data_rows.append({
            "row_idx": r + 1,
            "is_context": r < r0 or r >= r1,
            "values": vals,
        })
    out = {
        "headers": list(header[ec0 - c0:ec1 - c0]),
        "col_offset": ec0,
        "highlight_cols": list(highlight_cols),
        "highlight_rows": list(highlight_rows) if highlight_rows else [],
        "rows": data_rows,
    }
    if truncated:
        out["truncated"] = True
    return out


def _norm_label(s):
    """Normalize a row label for equality: lowercased, whitespace-collapsed; '' for
    None or a synthetic 'row N' placeholder (two unlabeled rows are not 'same-named')."""
    if not s:
        return ""
    t = " ".join(str(s).split()).strip().lower()
    return "" if re.fullmatch(r"row \d+", t) else t


def _is_round_power_of_ten(ratio):
    """True if `ratio` is (approximately) 10**k for some non-zero integer k — the
    fingerprint of a unit conversion or a percent/fraction restatement (x10, x100,
    x0.01, ...). Arbitrary ratios (1.14, 1.042) are NOT, and stay unexplained."""
    r = abs(float(ratio))
    if r <= 0 or r == 1.0:
        return False
    exp = math.log10(r)
    return abs(exp - round(exp)) < 1e-9 and round(exp) != 0


def benign_reason(f):
    """Return a common innocent explanation for a finding kind, or None.

    Attached to findings as `likely_benign` so the agent always has the
    false-positive context in hand and the HTML report can show it inline.
    """
    kind = f.get("kind")
    if kind == "arithmetic_progression":
        if f.get("reused_progression"):
            return ("this exact progression is re-plotted across >=2 panels — an "
                    "independent-variable axis (field / angle / time / dose / wavelength "
                    "sweep), not measured data")
        step = f.get("step")
        if step is not None and abs(step - round(step)) < 1e-9:
            return ("an integer-step progression is usually an axis (day / dose / "
                    "timepoint), not measured data")
        return None
    if kind == "rounded_to_half_or_int":
        return ("values ending in .0/.5 are common for derived or instrument-rounded "
                "quantities (cell counts, scores, calibrated readouts)")
    if kind == "identical_after_rounding":
        return ("cells share a rounded value but differ at full precision — usually "
                "display rounding, not duplication")
    if kind in ("constant_ratio_row", "scaled_row_reuse", "identical_row_reuse"):
        # A same-named row reused across two PANELS of one figure (different sheet, same
        # figure number) is the classic shared control/baseline replot — benign. A
        # same-sheet cross-block pair (e.g. a DMSO vs MMS arm) or a DIFFERENT-named row
        # is NOT a shared control and stays unexplained.
        if (f.get("same_figure") and not f.get("same_sheet")
                and _norm_label(f.get("row_a")) == _norm_label(f.get("row_b"))
                and _norm_label(f.get("row_a"))):
            return ("the same-named row reused across two panels of one figure is usually "
                    "a shared control/baseline replot — confirm the legend discloses the reuse")
        if kind != "identical_row_reuse":
            ratio = f.get("ratio")
            if ratio is not None and _is_round_power_of_ten(float(ratio)):
                return ("a whole power-of-ten ratio between two rows is usually a unit "
                        "conversion or percentage-vs-fraction restatement of the same row, "
                        "not two independent measurements")
        return None
    if kind in ("cross_sheet_value_overlap", "cross_sheet_position_identical"):
        if f.get("same_figure"):
            return f.get("context")
        if f.get("same_file") is False:
            return ("a control/baseline cohort is often reused across a main figure and "
                    "its extended-data figure — confirm the legend discloses the reuse")
    if kind in ("grim_inconsistent", "grimmer_inconsistent"):
        return ("GRIM/GRIMMER assume the statistic is a mean of integer-valued "
                "items (counts/scores); verify the measure is integer-granular "
                "before acting")
    return None


def _attach_benign(findings):
    """Mutate findings in-place to add a `likely_benign` note where one applies."""
    for f in findings:
        reason = benign_reason(f)
        if reason:
            f["likely_benign"] = reason
    return findings


def _attach_evidence(findings, sheet, r0, r1, c0, c1, header):
    """Mutate each finding in-place to add an `evidence` field, derived from the same
    block coordinates the detector was scanning. Highlight columns come from the
    finding's own col_*_idx / col_idx fields."""
    for f in findings:
        hi_cols = []
        for k in ("col_a_idx", "col_b_idx", "col_idx"):
            if k in f and isinstance(f[k], int):
                hi_cols.append(f[k])
        hi_rows = []
        for k in ("row_a_idx", "row_b_idx", "row_idx"):
            if k in f and isinstance(f[k], int):
                hi_rows.append(f[k] + 1)
        # identical_after_rounding / within_col_dispersed_repeats list specific
        # (row, col) example cells (1-based).
        for ex in f.get("example_cells", []) or []:
            try:
                hi_rows.append(int(ex[0]))
                hi_cols.append(int(ex[1]) - 1)
            except (TypeError, ValueError, IndexError):
                pass
        # De-duplicate (order-preserving): a column/row referenced by both an *_idx
        # field and one or more example_cells should highlight once, not N times.
        hi_cols = list(dict.fromkeys(hi_cols))
        hi_rows = list(dict.fromkeys(hi_rows))
        f["evidence"] = _block_evidence(sheet, r0, r1, c0, c1, header,
                                        highlight_cols=hi_cols,
                                        highlight_rows=hi_rows)
    return findings


# ---------- detectors ----------

def _isclose_rowwise(actual, expected, rtol=1e-9):
    """Return row-wise closeness at each row's own numeric scale.

    A single coordinate/metadata row can be orders of magnitude larger than
    the measurement rows. A block-wide absolute tolerance lets those small
    measurement rows drift substantially while still passing a relation check.
    """
    actual = np.asarray(actual, dtype=float)
    expected = np.asarray(expected, dtype=float)
    row_scale = np.maximum.reduce([
        np.abs(actual),
        np.abs(expected),
        np.full_like(actual, 1e-300, dtype=float),
    ])
    typical_scale = max(float(np.median(row_scale)), 1e-300)
    tol = rtol * row_scale + (np.finfo(float).eps * typical_scale * 64)
    return np.abs(actual - expected) <= tol


def _allclose_rowwise(actual, expected, rtol=1e-9):
    return bool(np.all(_isclose_rowwise(actual, expected, rtol=rtol)))


_GRIM_MEAN_RE = re.compile(r"\b(mean|average|avg)\b|均值|平均", re.I)
_GRIM_SD_RE = re.compile(r"\b(s\.?d\.?|std)\b|标准差", re.I)
_GRIM_N_RE = re.compile(r"\bn\b|sample.?size|样本量|例数", re.I)
_GRIM_INT_RE = re.compile(
    r"count|number|cells|foci|colon|nuclei|score|rating|likert"
    r"|个数|数目|计数|数量|评分|#", re.I)
_GRIM_RATIO_RE = re.compile(
    r"%|percent|percentage|\bratio\b|\brate\b|\bindex\b|proportion|fraction"
    r"|百分|比例|比率|占比|指数", re.I)


def detect_relations(sheet, r0, r1, c0, c1, header):
    findings = []
    cols = [(c, col_array(sheet, r0, r1, c)) for c in range(c0, c1)]
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            ci, ai = cols[i]
            cj, aj = cols[j]
            mask = ~np.isnan(ai) & ~np.isnan(aj)
            n = int(mask.sum())
            if n < 4:
                continue
            x, y = ai[mask], aj[mask]
            # Compact value peek for downstream LLM triage (bounded <=8 each, ~tiny).
            sa, sb = _sample(x), _sample(y)
            # Scale-relative tolerance. A fixed absolute atol (1e-9) misfires on tiny-magnitude
            # data: e.g. MEG fields ~1e-14 T are all within 1e-9 of each other, so every column
            # pair falsely reads as identical/linear. Tie the tolerance to the data magnitude so
            # these are tests of RELATIVE precision at any scale (and large-magnitude columns
            # aren't held to an unreasonably tight absolute bound either).
            tol = 1e-9 * max(float(np.max(np.abs(x))), float(np.max(np.abs(y))), 1e-300)
            # identical
            if _allclose_rowwise(x, y):
                findings.append(dict(kind="identical_column", col_a=header[ci - c0], col_b=header[cj - c0],
                                     col_a_idx=ci, col_b_idx=cj, n=n, severity="high",
                                     col_a_sample=sa, col_b_sample=sb,
                                     rule=f"col[{cj}] == col[{ci}]"))
                continue
            # constant offset
            diff = y - x
            mean_diff = float(np.mean(diff))
            if abs(mean_diff) > tol and _allclose_rowwise(y, x + mean_diff):
                findings.append(dict(kind="constant_offset", col_a=header[ci - c0], col_b=header[cj - c0],
                                     col_a_idx=ci, col_b_idx=cj, n=n, offset=mean_diff,
                                     severity="high",
                                     col_a_sample=sa, col_b_sample=sb,
                                     rule=f"col[{cj}] = col[{ci}] + {mean_diff:.6g}"))
                continue
            # constant ratio
            ratio_emitted = False
            if np.all(np.abs(x) > 1e-12):
                ratio = y / x
                mean_ratio = float(np.mean(ratio))
                ratio_tol = 1e-9 * max(abs(mean_ratio), 1e-300)
                if (
                    np.std(ratio) < ratio_tol
                    and abs(mean_ratio - 1) > 1e-9
                    and abs(mean_ratio) > 1e-9
                    and _allclose_rowwise(y, mean_ratio * x)
                ):
                    findings.append(dict(kind="constant_ratio", col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n, ratio=mean_ratio,
                                         severity="high",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=f"col[{cj}] = col[{ci}] * {mean_ratio:.6g}"))
                    ratio_emitted = True
            # mirror: x + y == constant
            csum = x + y
            if n >= 5 and np.std(csum) < tol:
                K = float(np.mean(csum))
                if abs(K) > tol:
                    findings.append(dict(kind="sum_constant", col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n, sum=K,
                                         severity="high",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=f"col[{ci}] + col[{cj}] = {K:.6g}"))
            # exact linear (non-identical)
            if n >= 5 and np.ptp(x) > 1e-12:
                try:
                    slope, intercept, r, _p, _se = stats.linregress(x, y)
                except ValueError:
                    continue
                fitted = slope * x + intercept
                if np.std(y) > 0 and _allclose_rowwise(y, fitted, rtol=1e-7) and abs(r) > 0.99:
                    # A scale-relatively zero intercept means the fit is y = slope*x: the
                    # identity (slope~=1, caught by identical_column) or a pure scaling. When a
                    # constant_ratio already captured that scaling, a second exact_linear finding
                    # is redundant (same relationship, b==0 to round-off) and only inflates the
                    # count — suppress it. exact_linear is reserved for a genuine non-zero
                    # intercept (an affine offset constant_ratio cannot express), and still fires
                    # when no constant_ratio covered the pair (e.g. a zero in x skips its guard).
                    intercept_is_zero = abs(intercept) < tol
                    is_identity = abs(slope - 1) < 1e-9 and intercept_is_zero
                    redundant_scaling = intercept_is_zero and ratio_emitted
                    if not (is_identity or redundant_scaling):
                        findings.append(dict(kind="exact_linear", col_a=header[ci - c0], col_b=header[cj - c0],
                                             col_a_idx=ci, col_b_idx=cj, n=n,
                                             slope=float(slope), intercept=float(intercept),
                                             severity="high",
                                             col_a_sample=sa, col_b_sample=sb,
                                             rule=f"col[{cj}] = {slope:.4g} * col[{ci}] + {intercept:.4g}"))
            # B4: partial constant offset — a long CONSECUTIVE run where y = x + k for a fixed
            # non-zero k, while the rest of the column diverges (the whole-column case is
            # constant_offset above). A contiguous block shifted by a fixed amount is a
            # copy-then-shift fingerprint; two independent columns do not hold a fixed offset
            # over a long contiguous run. Guarded to non-trivial offsets and long runs only.
            if n >= 24:
                # Scale-relative run detection on the raw diff (a fixed decimal round would be
                # inert on small-magnitude data — the exact regime `tol` above was written for).
                best_len = cur_len = 1
                best_val = float(diff[0])
                for t in range(1, len(diff)):
                    if abs(diff[t] - diff[t - 1]) < tol:
                        cur_len += 1
                    else:
                        if cur_len > best_len:
                            best_len, best_val = cur_len, float(diff[t - 1])
                        cur_len = 1
                if cur_len > best_len:
                    best_len, best_val = cur_len, float(diff[-1])
                run_floor = max(20, int(round(0.5 * n)))
                col_hp = sum(1 for v in x if _sig_frac_digits(v) >= 2) >= 0.6 * len(x)
                # The benign case to exclude is a run shifted by a small WHOLE number on
                # low-precision data (e.g. B = A + 5). Test that scale-relatively (tol), so a
                # genuine small-magnitude offset like 3e-14 is not mistaken for "integer 0".
                off_is_small_integer = abs(best_val - round(best_val)) < tol and abs(round(best_val)) >= 1
                non_trivial_offset = (not off_is_small_integer) or col_hp
                if (best_len >= run_floor and best_len < n
                        and abs(best_val) > tol and non_trivial_offset):
                    findings.append(dict(kind="partial_constant_offset",
                                         col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n,
                                         run_length=int(best_len), offset=float(best_val),
                                         severity="high",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=(f"col[{cj}] = col[{ci}] + {best_val:.6g} over a run of "
                                               f"{int(best_len)}/{n} consecutive rows")))
                    continue
            # integer difference with shared decimal fractions (B5), else small discrete diff set
            # B5: y and x reproduce each other's HIGH-PRECISION decimal fractions row-wise while
            # differing only by whole numbers that VARY across rows (a constant integer offset is
            # already caught above as constant_offset). Independent measurements do not reproduce
            # another column's 4+-decimal fractions on several rows — a copy-then-shift fingerprint
            # (e.g. 178.7615 vs 112.7615, 169.8687 vs 115.8687). The precision requirement lets this
            # fire from n>=5 without the false positives a bare small-diff-set floor would admit.
            if n >= 5:
                # Per-row tolerance for the integer-difference test: representation noise at each
                # row's OWN magnitude, not the column-wide max. A single extreme value (an inf /
                # placeholder like a 1e99 fold-change for a zero-denominator row) must not inflate
                # the tolerance so that every row's diff reads as a whole number — that produced
                # spurious whole-sheet integer_diff_shared_fraction findings (M2-1).
                diff_tol = 1e-9 * np.maximum(np.maximum(np.abs(x), np.abs(y)), 1e-300)
                diff_is_int = np.abs(diff - np.round(diff)) < diff_tol
                frac_x = x - np.round(x)                       # signed distance to nearest integer
                hp_rows = diff_is_int & (np.abs(frac_x) > 1e-6)
                hp_fracs = [float(v) for v in frac_x[hp_rows] if _sig_frac_digits(v) >= 4]
                # Exclude shared fractions that are a small-denominator residue (k/13, k/19,
                # …): those recur across integers as a quantization artifact, not a copy. The
                # distinct-count then reflects only arbitrary high-entropy fractions. (Shared
                # helper with the row-oriented shared-fraction detectors.)
                distinct_hp = len({round(v, 6) for v in hp_fracs
                                   if not _shared_frac_is_small_denominator(v)})
                int_diffs = np.unique(np.round(diff[diff_is_int]))
                n_real_frac = int(hp_rows.sum())        # rows sharing a genuine (non-.0) fraction
                if (int(diff_is_int.sum()) >= max(5, int(round(0.8 * n)))
                        and distinct_hp >= 3
                        and len(int_diffs) >= 2):
                    findings.append(dict(kind="integer_diff_shared_fraction",
                                         col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n,
                                         n_shared_fraction=n_real_frac,
                                         n_high_precision=distinct_hp,
                                         severity="high",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=(f"col[{cj}] and col[{ci}] share the same decimal fraction on "
                                               f"{n_real_frac}/{n} rows ({distinct_hp} distinct high-precision "
                                               f"fractions) but differ by whole numbers")))
                    continue
                # B5b: the same shared-fraction + integer-difference signal that B5 above
                # requires >=4 fraction digits for, admitted at lower precision ONLY when
                # every non-zero integer difference is a multiple of 10. Shifting a value
                # by a round number while keeping its decimals has no benign additive
                # transform, so that constraint compensates for a less-distinctive 2-digit
                # fraction. The evidence floor counts rows that are BOTH a round-shift AND
                # carry a genuine (non-.0) fraction — so a mostly-integer column with a few
                # stray decimals cannot satisfy it on thin fractional evidence.
                nz = diff_is_int & (np.abs(np.round(diff)) >= 1)
                round10 = nz & (np.abs(np.round(diff) - np.round(diff / 10.0) * 10.0) < 0.5)
                shared_frac = diff_is_int & (np.abs(frac_x) > 1e-6)
                rs_frac = round10 & shared_frac
                distinct_frac = len({round(float(v), 6) for v in frac_x[rs_frac]})
                if (int(rs_frac.sum()) >= max(5, int(round(0.7 * n)))
                        and int(round10.sum()) == int(nz.sum())
                        and distinct_frac >= 3):
                    findings.append(dict(kind="round_shift_shared_fraction",
                                         col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n,
                                         n_shared_fraction=int(rs_frac.sum()),
                                         severity="high",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=(f"col[{cj}] and col[{ci}] share the same decimal fraction on "
                                               f"{int(rs_frac.sum())}/{n} rows and differ only by non-zero "
                                               f"integer multiples of 10")))
                    continue
            # small discrete diff set
            if n >= 8:
                diff_rounded = np.round(diff, 4)
                uniq = np.unique(diff_rounded)
                if 2 <= len(uniq) <= min(6, n // 3):
                    findings.append(dict(kind="small_diff_set", col_a=header[ci - c0], col_b=header[cj - c0],
                                         col_a_idx=ci, col_b_idx=cj, n=n,
                                         unique_diffs=[float(x) for x in uniq],
                                         severity="medium",
                                         col_a_sample=sa, col_b_sample=sb,
                                         rule=f"col[{cj}] - col[{ci}] only takes {len(uniq)} discrete values"))
    return findings


_ROW_PAIR_MAX_ROWS = 80
_ROW_PAIR_MAX_COLS = 200
_ROW_PAIR_MAX_FINDINGS_PER_BLOCK = 25

# `detect_row_relations` gates: it compares every ROW PAIR across all columns, so it
# is bounded to few-rows/many-columns (condition-in-rows, measurement-in-columns)
# layouts — the ones the column-oriented `detect_relations` is blind to. Capping the
# row count keeps the O(rows^2) pair loop cheap on tall entity-in-rows tables (which
# are not this orientation anyway); the column floor keeps a proportional pair from
# firing on too few cells to be distinctive.
_ROW_REL_MAX_ROWS = int(os.environ.get("PAPERCONAN_ROW_REL_MAX_ROWS", "60"))
_ROW_REL_MIN_COLS = int(os.environ.get("PAPERCONAN_ROW_REL_MIN_COLS", "12"))
# Short-run row reuse (detect_short_row_reuse): the long-run detectors above miss the
# JCI "Supporting Data Values" layout, where each sub-panel is its own 1-4 row block and
# the copied/scaled segment is only 3-8 columns. A short run is safe from chance only if
# every value carries enough significant figures — so this path requires >=5 sig figs per
# cell and a shorter minimum run. FP math: two independent >=5-sig-fig cells collide at
# ~1e-4, so a 3-cell identical/ratio run is ~1e-8..1e-12 by chance.
_SHORT_ROW_MIN_COLS = int(os.environ.get("PAPERCONAN_SHORT_ROW_MIN_COLS", "3"))
_SHORT_ROW_MIN_SIGFIGS = int(os.environ.get("PAPERCONAN_SHORT_ROW_MIN_SIGFIGS", "5"))
# Tighter than _ROW_REL_RTOL: a short ratio run has fewer cells to corroborate the
# constant, so the constancy must be crisp to stay clear of chance.
_SHORT_ROW_RTOL = float(os.environ.get("PAPERCONAN_SHORT_ROW_RTOL", "1e-4"))
# Per-sheet cap on high-precision candidate rows (bounds the O(rows^2) pair loop).
_SHORT_ROW_MAX_ROWS_PER_SHEET = int(os.environ.get("PAPERCONAN_SHORT_ROW_MAX_ROWS", "400"))
# A run value that recurs often across the sheet is QUANTIZED (a k/19 normalization grid,
# a dose-response plateau), so two rows sharing it is not distinctive — the same trap
# `detect_decimal_tail_clustering` guards. A genuine reuse duplicates values that are
# otherwise unique to the two rows (freq 2-4); anything above this is a common-pool match.
_SHORT_ROW_MAX_VALUE_FREQ = int(os.environ.get("PAPERCONAN_SHORT_ROW_MAX_VALUE_FREQ", "8"))
# Ratio-run membership tolerance (relative). A genuine exact scaling read back from
# stored source data wobbles by ~2x the per-cell rounding: ~1e-6 at 6 sig figs but
# ~1e-4 for the very common 2-3 sig-fig bench readouts (percent, viability, OD). 1e-3
# absorbs all of these while staying 2+ orders below the ~0.1-0.3 spread of ratios
# between genuinely independent rows — at >=12 consecutive columns, random-chance FP
# is still ~(1e-3/0.3)**11 ≈ 1e-27, so it cannot manufacture a run.
_ROW_REL_RTOL = float(os.environ.get("PAPERCONAN_ROW_REL_RTOL", "1e-3"))
# Per-call column-op budget for detect_row_relations. rows are capped, but each row
# pair runs a pure-Python O(cols) scan, so a very wide block (e.g. 60x160000, still
# under _MAX_CELLS) would otherwise cost ~minutes. Bound total pair*cols work; a
# starved run stops early (stderr note) rather than hanging.
_ROW_REL_BUDGET = int(os.environ.get("PAPERCONAN_ROW_REL_BUDGET", "6000000"))


def _row_label(sheet, r, c0):
    labels = []
    for c in range(max(0, c0 - 4), c0):
        v = sheet.cell(r, c)
        if v is not None and not is_num(v):
            s = str(v).strip()
            if s:
                labels.append(s)
    return " | ".join(labels) if labels else f"row {r + 1}"


def _has_fractional_part(v):
    fv = float(v)
    return abs(fv - round(fv)) > 1e-9


def _ones_digit(v):
    return int(math.floor(abs(float(v)) + 1e-9)) % 10


def _decimal_digit(v, place=1):
    scale = 10 ** place
    return int(math.floor(abs(float(v)) * scale + 1e-8)) % 10


def _sig_frac_digits(v):
    """Count significant fractional decimal digits of v's distance to the nearest integer.
    167.93 -> 2 (.07), 178.7615 -> 4 (.2385), 100.5 -> 1, an integer -> 0."""
    fv = abs(float(v) - round(float(v)))
    if fv < 1e-9:
        return 0
    return len(f"{fv:.9f}".split(".")[1].rstrip("0"))


def _is_multiple_of_ten_diff(d):
    if abs(d) < 10 - 1e-8:
        return False
    nearest = round(d / 10.0) * 10.0
    return abs(d - nearest) <= 1e-7


def _row_pair_low_cardinality_integer_like(x, y):
    combined = np.concatenate([x, y])
    finite = combined[np.isfinite(combined)]
    if len(finite) == 0:
        return True
    near_integer = np.mean(np.abs(finite - np.round(finite)) < 1e-9)
    distinct = len(set(np.round(finite, 4).tolist()))
    max_abs = float(np.max(np.abs(finite)))
    return bool(near_integer >= 0.9 and max_abs <= 20 and distinct <= max(5, len(finite) // 4))


def detect_row_pair_digit_coupling(sheet, r0, r1, c0, c1, header, min_n=10):
    """Detect suspicious paired rows that preserve low-order digits across many cells.

    This targets source-data layouts where replicate/condition rows are aligned by
    measurement column. A concerning pattern is: row B differs from row A in value,
    but the first decimal digit and often the ones digit are preserved across many
    paired cells, with differences frequently landing on coarse multiples of 10.
    """
    findings = []
    n_rows = r1 - r0
    n_cols = c1 - c0
    if n_rows < 2 or n_cols < min_n:
        return findings
    if n_rows > _ROW_PAIR_MAX_ROWS or n_cols > _ROW_PAIR_MAX_COLS:
        return findings

    labels = {r: _row_label(sheet, r, c0) for r in range(r0, r1)}
    for i, ra in enumerate(range(r0, r1)):
        label_a = labels[ra]
        if _AXIS_CONTEXT_LABEL_RE.search(label_a):
            continue
        a = sheet.numeric[ra, c0:c1]
        for rb in range(r0 + i + 1, r1):
            label_b = labels[rb]
            if _AXIS_CONTEXT_LABEL_RE.search(label_b):
                continue
            b = sheet.numeric[rb, c0:c1]
            mask = ~np.isnan(a) & ~np.isnan(b)
            n = int(mask.sum())
            if n < min_n:
                continue
            x = a[mask].astype(float)
            y = b[mask].astype(float)
            cols = [c for c, keep in zip(range(c0, c1), mask.tolist()) if keep]

            if _row_pair_low_cardinality_integer_like(x, y):
                continue

            non_integer_pairs = sum(
                1 for xv, yv in zip(x, y)
                if _has_fractional_part(xv) or _has_fractional_part(yv)
            )
            if non_integer_pairs < max(4, math.ceil(0.25 * n)):
                continue

            changed_mask = ~_isclose_rowwise(x, y, rtol=1e-9)
            changed = int(changed_mask.sum())
            if changed / n < 0.5:
                continue

            same_decimal1 = 0
            same_ones = 0
            same_ones_decimal1 = 0
            coarse_10_diff = 0
            examples = []
            diffs = []
            for col, xv, yv, is_changed in zip(cols, x, y, changed_mask.tolist()):
                dec_same = _decimal_digit(xv, 1) == _decimal_digit(yv, 1)
                ones_same = _ones_digit(xv) == _ones_digit(yv)
                if dec_same:
                    same_decimal1 += 1
                if ones_same:
                    same_ones += 1
                if dec_same and ones_same:
                    same_ones_decimal1 += 1
                diff = float(yv - xv)
                diffs.append(round(diff, 6))
                if is_changed and _is_multiple_of_ten_diff(diff):
                    coarse_10_diff += 1
                if len(examples) < 8 and dec_same and is_changed:
                    examples.append({
                        "col": col + 1,
                        "header": header[col - c0] if 0 <= col - c0 < len(header) else "",
                        "a": float(xv),
                        "b": float(yv),
                        "diff": diff,
                    })

            frac_decimal1 = same_decimal1 / n
            frac_ones_decimal1 = same_ones_decimal1 / n
            frac_coarse_10 = coarse_10_diff / n
            severity = None
            if (
                n >= 12
                and frac_decimal1 >= 0.90
                and frac_ones_decimal1 >= 0.50
                and changed / n >= 0.50
                and frac_coarse_10 >= 0.50
            ):
                severity = "high"
            elif (
                n >= 12
                and frac_decimal1 >= 0.85
                and changed / n >= 0.50
                and (frac_ones_decimal1 >= 0.45 or frac_coarse_10 >= 0.45)
            ):
                severity = "medium"
            if not severity:
                continue

            top_diffs = Counter(diffs).most_common(6)
            findings.append(dict(
                kind="row_pair_digit_coupling",
                row_a=label_a,
                row_b=label_b,
                row_a_idx=ra,
                row_b_idx=rb,
                n=n,
                changed=changed,
                same_decimal1=same_decimal1,
                same_decimal1_frac=frac_decimal1,
                same_ones=same_ones,
                same_ones_decimal1=same_ones_decimal1,
                same_ones_decimal1_frac=frac_ones_decimal1,
                coarse_10_diff=coarse_10_diff,
                coarse_10_diff_frac=frac_coarse_10,
                top_diffs=[{"diff": float(d), "count": int(c)} for d, c in top_diffs],
                examples=examples,
                example_cells=[(ra + 1, ex["col"]) for ex in examples[:4]]
                              + [(rb + 1, ex["col"]) for ex in examples[:4]],
                severity=severity,
                rule=(f"rows {ra + 1} and {rb + 1}: first decimal digit matches "
                      f"{same_decimal1}/{n}; ones+decimal matches "
                      f"{same_ones_decimal1}/{n}; coarse 10-step differences "
                      f"{coarse_10_diff}/{n}"),
            ))

    findings.sort(key=lambda f: (
        0 if f["severity"] == "high" else 1,
        -f["same_decimal1_frac"],
        -f["same_ones_decimal1_frac"],
        -f["coarse_10_diff_frac"],
        -f["n"],
    ))
    return findings[:_ROW_PAIR_MAX_FINDINGS_PER_BLOCK]


# Above this many pairwise column relations in ONE block, the sheet is a dense /
# correlated matrix (correlation tables, normalized replicate panels) where identical or
# linear columns are expected by construction — not a duplication red flag. One real
# proteomics sheet produced ~20,000 such 'high' relations, drowning the genuine signal.
RELATION_FLOOD_CAP = 40

# Above this many within-column findings on ONE (file, sheet), the sheet is a large
# data table whose columns are repetitive by construction (categorical codes, dose
# grids, few-value panels). Genuine within-col signals live in low-count sheets
# (offline corpus: genuine-signal sheets held <=2 within_col each), so a sheet-wide
# flood is noise — demote it wholesale instead of flooding the judge.
WITHIN_COL_SHEET_CAP = 25


def _demote_within_col_flood(within_col, cap=WITHIN_COL_SHEET_CAP):
    """Demote a per-sheet flood of within-column findings to low severity, dropping them
    from the packet (prefilter='drop'). Kept in scan.json (reversible via forensic).
    Mutates + returns the same list."""
    if len(within_col) <= cap:
        return within_col
    for f in within_col:
        f["severity"] = "low"
        f["prefilter"] = "drop"
        f["prefilter_reason"] = "within_col_sheet_flood"
        f["within_col_flood_sheet"] = True
    return within_col


def _demote_dense_relations(relations, cap=RELATION_FLOOD_CAP):
    """Demote a flood of pairwise column relations to low severity (tagging them
    ``dense_block``) so a dense matrix stops dominating high-severity output. Findings
    are kept, not dropped — just down-weighted. Returns the same list."""
    if len(relations) <= cap:
        return relations
    for r in relations:
        r["severity"] = "low"
        r["dense_block"] = True
    return relations


def _demote_dense_sheets(report_blocks, cap=RELATION_FLOOD_CAP):
    """Apply the dense-flood demotion per (file, sheet), not per block: a dense matrix
    is split into many numeric blocks, each holding only part of the column relations,
    so the flood must be judged by the SHEET total. Mutates findings in place."""
    by_sheet = {}
    for b in report_blocks:
        key = (b["file"], b["sheet"])
        agg = by_sheet.setdefault(key, {"relations": [], "equal_pairs": [], "within_col": []})
        agg["relations"].extend(b.get("relations", []))
        agg["equal_pairs"].extend(b.get("equal_pairs", []))
        agg["within_col"].extend(b.get("within_col", []))
    for agg in by_sheet.values():
        _demote_dense_relations(agg["relations"], cap)   # same dict objects as in blocks
        _demote_dense_relations(agg["equal_pairs"], cap)
        _demote_within_col_flood(agg["within_col"])      # per-sheet within-col flood gate
    return report_blocks


def _demote_reused_progressions(report_blocks):
    """A perfect arithmetic progression that is REUSED — the identical (step, n, first)
    appears in >=2 numeric blocks/sheets — is an independent-variable axis re-plotted across
    panels (magnetic-field / 2-theta / time / dose / wavelength sweep), not a data inconsistency.
    Real measured data is never a perfect progression; a reused perfect progression is an axis.
    Demote these out of the high/medium review priority (kept in scan.json, reversible via
    forensic). A ONE-OFF perfect progression keeps its severity — that is the genuinely
    suspicious linear-fill case (and matches the golden fixture's single ap_col). Mutates in
    place and returns report_blocks."""
    sig_count = {}
    progs = []
    for b in report_blocks:
        for f in b.get("progressions", []):
            if f.get("kind") != "arithmetic_progression":
                continue
            sig = (round(float(f.get("step", 0.0)), 9), f.get("n"),
                   round(float(f.get("first", 0.0)), 9))
            sig_count[sig] = sig_count.get(sig, 0) + 1
            progs.append((sig, f))
    for sig, f in progs:
        if sig_count.get(sig, 0) >= 2:
            f["severity"] = "low"
            f["reused_progression"] = True
            f["prefilter"] = "drop"
            f["prefilter_reason"] = "reused_progression_axis"
            note = benign_reason(f)               # runs AFTER _attach_benign, so set it here
            if note:
                f["likely_benign"] = note
    return report_blocks


def detect_row_relations(sheet, r0, r1, c0, c1, header):
    """Row-oriented mirror of `detect_relations`: flag two ROWS that hold an exact
    relationship across many columns.

    Source data with experimental CONDITIONS in rows and per-cell MEASUREMENTS in
    columns hides "row B == row A" and "row B == row A * k" relationships from the
    column-pair detector entirely (they never touch a single column pair). Two
    DIFFERENT conditions that are bit-identical, or an exact constant multiple of
    each other cell-for-cell across dozens of columns, is a data-inconsistency
    signal worth an author's explanation — not a verdict.

    Called on every block and SELF-GATES on row/col counts (not the `wide` flag), so
    it covers precisely the wide blocks the column detectors skip. The row-count cap
    plus a per-call column-op budget keep the O(rows^2 * cols) pure-Python scan bounded
    even on genome-scale wide blocks.
    """
    findings = []
    n_rows = r1 - r0
    n_cols = c1 - c0
    if n_rows < 2 or n_cols < _ROW_REL_MIN_COLS or n_rows > _ROW_REL_MAX_ROWS:
        return findings

    budget = _ROW_REL_BUDGET
    labels = {r: _row_label(sheet, r, c0) for r in range(r0, r1)}
    for i, ra in enumerate(range(r0, r1)):
        label_a = labels[ra]
        if _AXIS_CONTEXT_LABEL_RE.search(label_a):
            continue
        a = sheet.numeric[ra, c0:c1]
        for rb in range(ra + 1, r1):
            budget -= n_cols
            if budget <= 0:
                print(f"[paperconan] detect_row_relations: column-op budget exhausted on a "
                      f"{n_rows}x{n_cols} block — coverage bounded", file=sys.stderr)
                return findings
            label_b = labels[rb]
            if _AXIS_CONTEXT_LABEL_RE.search(label_b):
                continue
            b = sheet.numeric[rb, c0:c1]
            mask = ~np.isnan(a) & ~np.isnan(b)
            n = int(mask.sum())
            if n < _ROW_REL_MIN_COLS:
                continue
            xm = a[mask].astype(float)
            # Too few DISTINCT values and a "constant ratio" is unremarkable
            # (e.g. two low-cardinality integer score rows). Require real spread.
            if np.ptp(xm) <= 0 or len(np.unique(xm)) < 6:
                continue
            sa, sb = _sample(xm), _sample(b[mask].astype(float))

            # identical rows (a bit-identical data group under two different labels)
            if _allclose_rowwise(xm, b[mask].astype(float)):
                findings.append(dict(kind="identical_row",
                                     row_a=label_a, row_b=label_b,
                                     row_a_idx=ra, row_b_idx=rb, n=n, severity="high",
                                     row_a_sample=sa, row_b_sample=sb,
                                     rule=f"row[{rb + 1}] == row[{ra + 1}] over {n} columns"))
                continue

            # constant ratio over the longest CONTIGUOUS run of columns where
            # row B == row A * k (k != 1). The real fingerprint scales only PART of a
            # row (a copy-then-scale of a column range), so requiring the whole row
            # would miss it — mirror the column detector's partial_constant_offset run.
            run = _longest_constant_ratio_run(a, b, c0, c1)
            if run is not None:
                k, run_len, x_run = run
                if run_len >= _ROW_REL_MIN_COLS and len(np.unique(x_run)) >= 6:
                    findings.append(dict(kind="constant_ratio_row",
                                         row_a=label_a, row_b=label_b,
                                         row_a_idx=ra, row_b_idx=rb, n=int(run_len),
                                         ratio=k, run_length=int(run_len), severity="high",
                                         row_a_sample=sa, row_b_sample=sb,
                                         rule=f"row[{rb + 1}] = row[{ra + 1}] * {k:.6g} over a run of "
                                              f"{int(run_len)}/{n} columns"))
    return findings


def _longest_constant_ratio_run(a, b, c0, c1):
    """Longest contiguous column run where b[c] == k * a[c] for a fixed k != 1.

    `a`, `b` are the two full row slices (may contain NaN). A column breaks the run
    if either cell is NaN or a[c] is ~0. Membership is anchored to the run's first
    ratio within `_ROW_REL_RTOL` (rounding-tolerant); the returned k is the run MEAN
    for a clean reported value. Only NON-unity runs compete for the longest — a
    near-unity (identical) prefix must not win and mask a shorter genuine scaling
    suffix (that identical part is caught separately). Returns (k, run_length,
    x_values_in_run) for the longest qualifying scaling run, or None."""
    best_len, best_start = 0, 0
    cur_len, cur_k, cur_start = 0, None, 0
    for idx in range(c1 - c0):
        av, bv = a[idx], b[idx]
        if math.isnan(av) or math.isnan(bv) or abs(av) <= 1e-12:
            cur_len, cur_k = 0, None
            continue
        r = bv / av
        if cur_k is None or abs(r - cur_k) > _ROW_REL_RTOL * max(abs(cur_k), 1e-300):
            cur_k, cur_len, cur_start = r, 1, idx
        else:
            cur_len += 1
        # Only a genuinely-scaled run (ratio distinct from 1) may set the best; a long
        # identical run is not a scaling and must not shadow a real one elsewhere.
        if cur_len > best_len and abs(cur_k - 1.0) > _ROW_REL_RTOL:
            best_len, best_start = cur_len, cur_start
    if best_len == 0:
        return None
    x_run = a[best_start:best_start + best_len].astype(float)
    y_run = b[best_start:best_start + best_len].astype(float)
    k = float(np.mean(y_run / x_run))
    if abs(k - 1.0) <= _ROW_REL_RTOL or abs(k) <= 1e-9:
        return None
    return k, best_len, x_run


def detect_arithmetic_progression(sheet, r0, r1, c0, c1, header):
    findings = []
    for c in range(c0, c1):
        a = col_array(sheet, r0, r1, c)
        a = a[~np.isnan(a)]
        if len(a) < 5:
            continue
        diffs = np.diff(a)
        tol = 1e-9 * max(float(np.max(np.abs(a))), 1e-300)   # scale-relative (see detect_relations)
        if np.allclose(diffs, diffs[0], atol=tol, rtol=1e-9) and abs(diffs[0]) > tol:
            sev = "medium" if abs(diffs[0] - round(diffs[0])) < 1e-9 else "high"
            findings.append(dict(kind="arithmetic_progression", col=header[c - c0], col_idx=c,
                                 block_c0=c0,
                                 n=int(len(a)), step=float(diffs[0]), first=float(a[0]),
                                 severity=sev,
                                 rule=f"col[{c}] = arithmetic progression, step={diffs[0]:.6g}"))
    return findings


def detect_within_column_patterns(sheet, r0, r1, c0, c1, header, min_n=6):
    """Detect within-column anomalies:
       - many identical values in one column (Su Jiacao: '13 中 8 个相同')
       - many values sharing same last-2 decimals (Su Jiacao: '13 中 11 个末两位相同')
       - too many .0 / .5 endings (Su Jiacao: '71 个中 51 个末位 0 或 5')
       - missing last digits (Su Jiacao: '70 个数据中末位完全没有 3 或 7')
    """
    findings = []
    for c in range(c0, c1):
        a = col_array(sheet, r0, r1, c)
        a_clean = a[~np.isnan(a)]
        n = len(a_clean)
        if n < min_n:
            continue
        col_name = header[c - c0] if c - c0 < len(header) else f"col{c}"

        # Cheap column descriptors shared by the within-col detectors below, so a
        # downstream prefilter can decide precisely (categorical/integer column,
        # low-cardinality, value peek) instead of guessing from the column name alone.
        vals_rounded = np.round(a_clean, 4)
        counts = Counter(vals_rounded.tolist())
        n_distinct = int(len(counts))
        all_integer = bool(np.all(np.abs(a_clean - np.round(a_clean)) < 1e-9))
        value_sample = [float(v) for v, _ in counts.most_common(8)]
        enrich = dict(n_distinct=n_distinct, all_integer=all_integer, value_sample=value_sample)

        # 1) duplicate values within the column
        top_val, top_count = counts.most_common(1)[0]
        if top_count >= max(4, n // 2) and n - top_count >= 1:
            findings.append(dict(kind="within_col_value_duplication",
                                 col=col_name, col_idx=c, n=n,
                                 dup_value=float(top_val), dup_count=int(top_count),
                                 frac_repeat=top_count / n, **enrich,
                                 severity="high",
                                 rule=f"col[{c}] has value {top_val} repeated {top_count}/{n} times"))

        # 2) last-2-decimal repetition within column
        endings = [trailing_decimal_digits(v, 2) for v in a_clean]
        endings = [e for e in endings if e is not None]
        if len(endings) >= max(min_n, 8):
            ec = Counter(endings)
            top_end, top_end_count = ec.most_common(1)[0]
            if top_end_count >= max(5, 2 * len(endings) // 3):
                findings.append(dict(kind="within_col_decimal_repetition",
                                     col=col_name, col_idx=c, n=len(endings),
                                     ending=top_end, count=int(top_end_count),
                                     frac_repeat=top_end_count / len(endings), **enrich,
                                     severity="high",
                                     rule=f"col[{c}]: {top_end_count}/{len(endings)} values share last-2 decimals '.{top_end}'"))

        # 3) too many .0 / .5 last decimal (rounded to half/int)
        last1 = [last_significant_digit(v) for v in a_clean]
        last1 = [d for d in last1 if d is not None]
        if len(last1) >= max(min_n, 10):
            zeros_fives = sum(1 for d in last1 if d in ("0", "5"))
            if zeros_fives >= max(7, 0.7 * len(last1)):
                findings.append(dict(kind="rounded_to_half_or_int",
                                     col=col_name, col_idx=c, n=len(last1),
                                     count_05=int(zeros_fives),
                                     severity="medium",
                                     rule=f"col[{c}]: {zeros_fives}/{len(last1)} values end in 0 or 5"))

        # 4) missing last-digit (3 or 7 completely absent in a large column)
        if len(last1) >= 20:
            present = set(last1)
            missing = [d for d in "123456789" if d not in present]
            if missing and len(present) <= 6:
                findings.append(dict(kind="missing_last_digits",
                                     col=col_name, col_idx=c, n=len(last1),
                                     missing=missing,
                                     severity="medium",
                                     rule=f"col[{c}]: last digits {missing} never appear in {len(last1)} values"))
    return findings


def detect_dispersed_repeats(sheet, r0, r1, c0, c1, header, min_n=30):
    """Many DISTINCT high-precision values each repeated across DISPERSED rows.

    Complements within_col_value_duplication (single dominant value). Targets a
    continuous, high-precision column whose exact-duplicate mass far exceeds the
    near-zero birthday expectation, where repeats are scattered across the table
    (not adjacent fill-down / technical replicates). Thresholds are conservative
    defaults pinned by tests; not env-tunable.
    """
    findings = []

    def _dec_places(v):
        s = f"{v:.10f}".rstrip("0")
        return len(s.split(".")[1]) if "." in s else 0

    for c in range(c0, c1):
        rows_vals = []
        for r in range(r0, r1):
            v = sheet.cell(r, c)
            if is_num(v) and not (isinstance(v, float) and np.isnan(v)):
                rows_vals.append((r, float(v)))
        n = len(rows_vals)
        if n < min_n:
            continue
        vals = [v for _, v in rows_vals]

        # quick reject: pure-integer columns (counts / codes)
        if all(abs(v - round(v)) < 1e-9 for v in vals):
            continue

        # Strip a dominant boundary/censor value FIRST (e.g. 600s ceiling), so it
        # neither drags down the precision fraction nor counts as a "repeat".
        cnt_all = Counter(round(v, 6) for v in vals)
        top_v, top_c = cnt_all.most_common(1)[0]
        boundary = top_v if top_c > 0.25 * n else None
        core = [(r, v) for (r, v) in rows_vals
                if boundary is None or round(v, 6) != boundary]
        m = len(core)
        if m < min_n:
            continue
        core_vals = [v for _, v in core]

        # Gate 1 — continuity / high precision (computed on core)
        frac_hi_prec = sum(1 for v in core_vals if _dec_places(v) >= 2) / m
        if frac_hi_prec < 0.6:
            continue
        distinct = len({round(v, 6) for v in core_vals})
        if distinct < 50 or distinct / m < 0.3:
            continue

        # Gate 1b — birthday / effective-support gate: the recording precision must
        # be fine ENOUGH relative to the value range that exact collisions are
        # near-zero-expected. A coarse column (e.g. 2 decimals over [0,1] -> only
        # ~100 possible values) collides naturally and must NOT fire.
        dps = sorted(_dec_places(v) for v in core_vals)
        med_dp = dps[len(dps) // 2]
        support = (max(core_vals) - min(core_vals)) * (10 ** med_dp)
        if support < 20 * m:
            continue

        # Gate 2 + 3 — dispersed exact-duplicate groups
        positions = defaultdict(list)
        for r, v in core:
            positions[round(v, 6)].append(r)
        block_h = r1 - r0
        dispersed = []
        dup_cells = 0
        for val, rs in positions.items():
            if len(rs) < 2:
                continue
            rs_sorted = sorted(rs)
            span = rs_sorted[-1] - rs_sorted[0]
            non_adjacent = any(b - a > 1 for a, b in zip(rs_sorted, rs_sorted[1:]))
            if span >= 0.5 * block_h and non_adjacent:
                dispersed.append((val, rs_sorted))
                dup_cells += len(rs_sorted)

        if len(dispersed) >= 10 and dup_cells >= 0.15 * m:
            dispersed.sort(key=lambda kv: -len(kv[1]))
            example_cells = []
            for _, rs in dispersed[:3]:
                for rr in rs[:8]:
                    example_cells.append((rr + 1, c + 1))
            col_name = header[c - c0] if c - c0 < len(header) else f"col{c}"
            core_arr = np.round(np.array([v for _, v in core]), 4)
            counts = Counter(core_arr.tolist())
            findings.append(dict(
                kind="within_col_dispersed_repeats",
                col=col_name, col_idx=c, n=m,
                n_repeat_groups=len(dispersed), dup_cells=dup_cells,
                frac_repeat=dup_cells / m,
                n_distinct=int(len(counts)), all_integer=False,
                value_sample=[float(v) for v, _ in counts.most_common(8)],
                example_cells=example_cells,
                severity="medium",
                rule=(f"col[{c}]: {len(dispersed)} distinct high-precision values "
                      f"each recur across dispersed rows ({dup_cells}/{m} cells)")))
    return findings


def detect_identical_after_rounding(sheet, r0, r1, c0, c1, header):
    """Detect pairs/groups of cells that differ at higher precision but match at lower (e.g.
       4.2735 vs 4.2812 — both round to 4.3). Kang Tiebang ED6h/6j signal."""
    findings = []
    cells = []
    for r in range(r0, r1):
        for c in range(c0, c1):
            v = sheet.cell(r, c)
            if is_num(v) and abs(v) > 1e-9:
                cells.append((r, c, float(v)))
    if len(cells) < 20:
        return findings
    # Bucket cells by 1-decimal rounded value
    from collections import defaultdict
    buckets = defaultdict(list)
    for r, c, v in cells:
        if abs(v) < 100:  # only meaningful for measurement-scale numbers
            buckets[round(v, 1)].append((r, c, v))
    # Find buckets where multiple DIFFERENT (>1e-4 apart) values map to the same rounded value
    suspicious = []
    for k, lst in buckets.items():
        if len(lst) >= 4:
            uniq = set(round(v, 4) for _, _, v in lst)
            if len(uniq) >= 3:
                suspicious.append((k, lst))
    suspicious.sort(key=lambda kv: -len(kv[1]))
    if suspicious:
        top = suspicious[:5]
        for k, lst in top:
            uniq = sorted(set(round(v, 4) for _, _, v in lst))
            findings.append(dict(kind="identical_after_rounding",
                                 rounded_to=float(k), n_cells=len(lst), n_unique=len(uniq),
                                 example_values=uniq[:6],
                                 example_cells=[(r + 1, c + 1) for r, c, _ in lst[:6]],
                                 severity="medium",
                                 rule=f"{len(lst)} cells share rounded value {k} but have {len(uniq)} distinct precise values"))
    return findings


def detect_grim_grimmer(sheet, r0, r1, c0, c1, header):
    """GRIM/GRIMMER: flag reported means (and SDs) impossible for integer-valued
    data at the stated n. Strictly gated — needs a header-located mean+n triple
    AND a count/score keyword in the MEAN column header signalling integer items —
    to stay false-positive-safe on continuous measurements where GRIM does not apply.
    GRIMMER runs only on a true SD column (SEM/SE columns are deliberately ignored,
    since GRIMMER is undefined for a standard error)."""
    findings = []

    def _find(rx, taken):
        for idx, h in enumerate(header):
            if idx not in taken and rx.search(str(h or "")):
                return idx
        return None

    taken = set()
    mean_i = _find(_GRIM_MEAN_RE, taken)
    if mean_i is not None:
        taken.add(mean_i)
    n_i = _find(_GRIM_N_RE, taken)
    if n_i is not None:
        taken.add(n_i)
    sd_i = _find(_GRIM_SD_RE, taken)
    if mean_i is None or n_i is None:
        return findings
    # Integer-data gate: the count/score keyword must be in the MEAN column header
    # itself, not anywhere in the row — otherwise a bookkeeping column such as
    # "number of replicates" would license GRIM on a continuous measurement.
    if not _GRIM_INT_RE.search(str(header[mean_i] or "")):
        return findings
    # Negative gate: a continuous ratio / percentage / index mean is not integer
    # data even when its header also contains a count word (e.g. "% positive cells").
    # NB: deliberately excludes "score"/"count" — GRIM's original domain is integer
    # composite/Likert scores, which must still be checked.
    if _GRIM_RATIO_RE.search(str(header[mean_i] or "")):
        return findings

    mean_c, n_c = c0 + mean_i, c0 + n_i
    sd_c = c0 + sd_i if sd_i is not None else None
    grim_fail, grimmer_fail = [], []
    checked = grimmer_checked = 0
    for r in range(r0, r1):
        mv = sheet.cell(r, mean_c)
        nv = sheet.cell(r, n_c)
        if not (is_num(mv) and is_num(nv)):
            continue
        n = int(round(float(nv)))
        if n < 2:
            continue
        mean = float(mv)
        d = _decimals_of(mean)
        if n >= 10 ** d:                 # power gate: no discriminating power
            continue
        checked += 1
        if not grim_consistent(mean, n, d):
            grim_fail.append((r, mean, n, d))
            continue                     # GRIM-failing rows are not re-reported
        if sd_c is not None:
            sv = sheet.cell(r, sd_c)
            if is_num(sv):
                sd = float(sv)
                ds = _decimals_of(sd)
                grimmer_checked += 1
                if not grimmer_consistent(mean, sd, n, d, ds):
                    grimmer_fail.append((r, mean, sd, n, ds))

    mean_name = str(header[mean_i] or f"col{mean_c}")
    n_name = str(header[n_i] or f"col{n_c}")
    sd_name = str(header[sd_i] or f"col{sd_c}") if sd_i is not None else None

    if grim_fail:
        f = dict(kind="grim_inconsistent", severity="high",
                 mean_col=mean_name, n_col=n_name, sd_col=sd_name,
                 col_a_idx=mean_c,
                 n=checked, n_rows_checked=checked, n_failed=len(grim_fail),
                 failed_rows=[dict(row=r + 1, mean=m, n=nn, decimals=dd,
                                   nearest_consistent=round(round(m * nn) / nn, dd))
                              for (r, m, nn, dd) in grim_fail[:8]],
                 example_cells=[[r + 1, mean_c + 1] for (r, *_rest) in grim_fail[:8]],
                 rule=(f"{len(grim_fail)}/{checked} rows report a mean impossible for "
                       f"integer data at the stated n (GRIM): col '{mean_name}'"))
        if sd_c is not None:
            f["col_b_idx"] = sd_c
        findings.append(f)
    if grimmer_fail:
        findings.append(dict(
            kind="grimmer_inconsistent", severity="high",
            mean_col=mean_name, n_col=n_name, sd_col=sd_name,
            col_a_idx=mean_c, col_b_idx=sd_c,
            n=grimmer_checked, n_rows_checked=grimmer_checked, n_failed=len(grimmer_fail),
            failed_rows=[dict(row=r + 1, mean=m, sd=s, n=nn, sd_decimals=ds)
                         for (r, m, s, nn, ds) in grimmer_fail[:8]],
            example_cells=[[r + 1, sd_c + 1] for (r, *_rest) in grimmer_fail[:8]],
            rule=(f"{len(grimmer_fail)}/{grimmer_checked} rows report an SD impossible for "
                  f"integer data at the stated mean & n (GRIMMER): col '{sd_name}'")))
    return findings


def detect_last_digit(values, label):
    digits = [int(d) for d in (last_significant_digit(v) for v in values) if d is not None and d != "0"]
    if len(digits) < 40:
        return None
    counts = Counter(digits)
    obs = np.array([counts.get(d, 0) for d in range(1, 10)], dtype=float)
    expected = np.full(9, obs.sum() / 9.0)
    chi2 = ((obs - expected) ** 2 / expected).sum()
    p = float(1 - stats.chi2.cdf(chi2, df=8))
    most_common = counts.most_common(3)
    return dict(label=label, n=int(obs.sum()), chi2=float(chi2), p=p,
                counts={str(d): int(counts.get(d, 0)) for d in range(0, 10)},
                top=[[str(d), c] for d, c in most_common])


def detect_repeated_decimals(values, label):
    endings = [trailing_decimal_digits(v, 2) for v in values]
    endings = [e for e in endings if e is not None]
    if len(endings) < 60:
        return None
    counts = Counter(endings)
    n = len(endings)
    flags = [(e, c) for e, c in counts.most_common(15) if c >= max(5, 5 * n / 100)]
    return dict(label=label, n=n, n_unique=len(counts), top=flags)


_TAIL_CLUSTER_MIN_N = int(os.environ.get("PAPERCONAN_TAIL_CLUSTER_MIN_N", "100"))
_TAIL_CLUSTER_SHARE = float(os.environ.get("PAPERCONAN_TAIL_CLUSTER_SHARE", "0.40"))


def detect_decimal_tail_clustering(values, label, top_k=6):
    """A few multi-digit fractional TAILS recurring far above chance across many
    INDEPENDENT high-precision values — a fingerprint of numbers drawn from a small set
    of fractional parts (copied/derived) rather than independently measured.

    Distinct from detect_last_digit (a single last digit), detect_repeated_decimals
    (2-digit endings, no concentration test) and within_col_value_duplication (repeated
    whole VALUES, not shared tails). Gated hard: only values with >=3 fractional digits
    (at read precision) count; needs >=_TAIL_CLUSTER_MIN_N of them; the top-`top_k`
    3-digit tails must cover >=_TAIL_CLUSTER_SHARE of them; AND the full fractional parts
    must be MOSTLY DISTINCT — otherwise a quantized / common-denominator column (values
    like k/7 or eighths) trivially shares tails and would false-positive. Large-magnitude
    values (>=1e7) are skipped: read-precision noise there reaches the captured digits."""
    tails, full, hp_vals = [], [], []
    for v in values:
        av = abs(float(v))
        if av >= 1e7:
            continue
        s = f"{av:.10f}".rstrip("0")
        if "." not in s:
            continue
        frac = s.split(".")[1]
        if len(frac) >= 3:
            tails.append(frac[-3:])
            full.append(frac)
            hp_vals.append(av)
    n = len(tails)
    if n < _TAIL_CLUSTER_MIN_N:
        return None
    # Quantized / common-denominator data (few distinct fractions) shares tails trivially.
    # The genuine fingerprint is MANY independent values that nonetheless collide on a few
    # tails, so require the full fractional parts to be mostly distinct.
    if len(set(full)) < max(50, n // 2):
        return None
    counts = Counter(tails)
    top = counts.most_common(top_k)
    top_sum = sum(c for _, c in top)
    share = top_sum / n
    if share < _TAIL_CLUSTER_SHARE:
        return None
    top_tails = [t for t, _ in top]
    # Averaging artifact: a reported MEAN of d replicates is (sum of d limited-precision
    # readings) / d, so its fractional tail is mechanically pinned to the residues of 1/d
    # (division by 3 -> .333/.667, by 6 -> .167/.333/.667/.833, ...). That concentration is
    # a benign consequence of averaging, not a copied-fraction fingerprint. If the values
    # carrying the dominant tails are almost all d-fold "terminating" for one small d — i.e.
    # value*d lands back on a short (<=4 dp) decimal — the cluster is an averaging artifact.
    # (Real JCI panels JCI195538 Fig1D/4A and JCI200564 Fig.2 false-positive exactly here.)
    carriers = [av for av, t in zip(hp_vals, tails) if t in set(top_tails)]
    if carriers:
        for d in range(2, 13):
            terminating = sum(1 for av in carriers
                              if abs(av * d - round(av * d, 4)) < 1e-6)
            if terminating >= 0.9 * len(carriers):
                return None
    # complementary pairs (t + t' = 1000) among the dominant tails — a stronger sub-signal
    comp = sum(1 for t in top_tails if int(t) < 500 and f"{1000 - int(t):03d}" in top_tails)
    return dict(label=label, n=n, n_unique=len(counts), n_distinct_fraction=len(set(full)),
                top=[[t, c] for t, c in top], top_share=round(share, 4),
                complementary_pairs=comp, severity="high",
                rule=(f"the {top_k} most common 3-digit fractional tails cover "
                      f"{top_sum}/{n} ({share:.0%}) of the high-precision values "
                      f"(uniform expectation ~{100 * top_k / 1000:.1f}%), which have "
                      f"{len(set(full))} distinct fractional parts"))


def benjamini_hochberg(pvals, alpha=0.05):
    """Benjamini-Hochberg step-up FDR. Returns (adjusted_pvals, significant_flags),
    both in the original order. Adjusted p (q-value) is the BH-corrected p; a sheet
    is significant when its q-value <= alpha. Controls false positives when dozens of
    per-sheet last-digit tests run at once."""
    m = len(pvals)
    if m == 0:
        return [], []
    order = sorted(range(m), key=lambda i: pvals[i])
    adj = [1.0] * m
    running_min = 1.0
    for rank in range(m, 0, -1):          # largest p (rank m) down to smallest (rank 1)
        i = order[rank - 1]
        running_min = min(running_min, pvals[i] * m / rank)
        adj[i] = min(running_min, 1.0)
    sig = [adj[i] <= alpha for i in range(m)]
    return adj, sig


def detect_equal_pairs(sheet, r0, r1, c0, c1, header):
    """Detect column pairs where many rows have identical values
    (e.g. tumor length == tumor width)."""
    findings = []
    A = sheet.block(r0, r1, c0, c1)
    for i in range(c1 - c0):
        for j in range(i + 1, c1 - c0):
            a, b = A[:, i], A[:, j]
            mask = ~np.isnan(a) & ~np.isnan(b)
            n = int(mask.sum())
            if n < 6:
                continue
            am, bm = a[mask], b[mask]
            # scale-relative tolerances, applied per row so one large metadata
            # coordinate does not make small measurement rows look equal.
            eq = int(_isclose_rowwise(am, bm, rtol=1e-6).sum())
            if eq >= max(6, n // 2) and eq / n >= 0.5 and not _allclose_rowwise(am, bm, rtol=1e-9):
                findings.append(dict(kind="many_equal_pairs", col_a=header[i], col_b=header[j],
                                     col_a_idx=c0 + i, col_b_idx=c0 + j, n=n, equal=eq,
                                     severity="medium" if eq < n else "high",
                                     col_a_sample=_sample(am), col_b_sample=_sample(bm),
                                     rule=f"col[{c0+i}] == col[{c0+j}] in {eq}/{n} rows"))
    return findings


# ---------- driver ----------

def _grid_from_rows(sheet, min_decimal_places=3, max_rows=200):
    """Build {(r, c): rounded_value} of decimal-bearing numeric cells from a Sheet.
    Only keeps non-integer values with >= min_decimal_places decimals in a sane range —
    these are the values whose bit-identical reuse across tables is suspicious."""
    grid = {}
    nm = sheet.numeric
    rmax = min(sheet.nrows, max_rows)
    for ri in range(rmax):
        for ci in range(sheet.ncols):
            fv = nm[ri, ci]
            if math.isnan(fv):
                continue
            if fv != int(fv) and 0.001 <= abs(fv) < 100000:
                s = repr(float(fv))
                if "." in s and "e" not in s.lower():
                    frac = s.split(".", 1)[1]
                    if len(frac) >= min_decimal_places:
                        grid[(ri, ci)] = round(fv, 9)
    return grid


import re as _re

# Matches a figure id inside a sheet name: an optional "extended/ED/ex" marker
# followed by a figure number, e.g. "Figure 5o", "exFig.6b-e", "ED_Fig8b", " exFig.6i".
_FIG_RE = _re.compile(r"(ext(?:ended)?|ed|ex)?\s*\.?\s*fig(?:ure)?\s*\.?\s*0*(\d+)", _re.I)
_CONTROL_BASELINE_LABEL_RE = _re.compile(
    r"\b(?:control|ctrl|baseline|vehicle|untreated|wt|wild[- ]?type|reference|mock|"
    r"naive|sham|pbs|dmso)\b|参照|对照|基线",
    _re.I,
)
_AXIS_CONTEXT_LABEL_RE = _re.compile(
    r"\b(?:time|day|dose|conc(?:entration)?|wavelength|m/z|mz|position|chr|"
    r"coordinate|coord|index|bin)\b|波长|时间|剂量",
    _re.I,
)


def figure_key(sheet_name):
    """Normalize a sheet name into a figure identity like 'main:5' or 'ext:6'.

    Returns None when no figure number can be parsed (e.g. 'Sheet1'). Two sheets
    with the SAME key are panels of the same display item — sharing data between
    them (a combined growth curve and its per-replicate breakdown) is expected and
    should not read as a cross-experiment duplication.
    """
    if not sheet_name:
        return None
    m = _FIG_RE.search(str(sheet_name))
    if not m:
        return None
    prefix = (m.group(1) or "").lower()
    namespace = "ext" if prefix else "main"
    return f"{namespace}:{m.group(2)}"


def _value_delta(ga, gb):
    """Characterize HOW two near-duplicate grids differ, so a clean re-plot can be
    told apart from a copy-then-tweak.

    - modified_cells: same (row,col) position, different value — only meaningful when
      the two tables share a layout; the copy-then-tweak fingerprint.
    - only_in_a / only_in_b: value-multiset members unique to each side (layout-robust).
    - pattern:
        value_tweaked : >=1 cell changed in place (most interesting — possible edit)
        perfect_dup   : identical value multisets, no in-place edits (clean re-plot)
        superset      : one side's values strictly contain the other's (e.g. an extra
                        replicate column — main shows n=5, extended shows n=6)
        value_divergent : both sides hold values the other lacks (partial overlap)
    """
    modified = sum(1 for k, v in ga.items() if k in gb and gb[k] != v)
    ca, cb = Counter(ga.values()), Counter(gb.values())
    shared = sum((ca & cb).values())
    only_a = sum(ca.values()) - shared
    only_b = sum(cb.values()) - shared
    # The value multiset is layout-robust, so decide on it FIRST: identical content is
    # a perfect_dup even if the two tables lay it out at different offsets (modified_cells
    # is then just a layout-shift artifact, meaningful only when layouts align).
    if only_a == 0 and only_b == 0:
        pattern = "perfect_dup"
    elif only_a == 0 or only_b == 0:
        pattern = "superset"
    elif modified > 0:
        pattern = "value_tweaked"
    else:
        pattern = "value_divergent"
    return dict(pattern=pattern, modified_cells=modified,
                shared_values=shared, only_in_a=only_a, only_in_b=only_b)


def value_tweak_subtype(delta: dict | None) -> str | None:
    """Sub-classify a ``value_tweaked`` cross-sheet overlap from an existing ``_value_delta``
    result, without changing detector output (reads fields only).

    - ``copy_then_edit``: a near-perfect copy with only a handful of cells retyped — the
      strongest manual-edit fingerprint (the page #8 pattern). Worth surfacing to judges.
    - ``block_edit``: a heavier rewrite of a shared block.
    - ``None``: not a ``value_tweaked`` pattern.

    Descriptive only — KEEP/DROP is unchanged; ``perfect_dup`` / ``mass`` / high-fraction
    overlaps stay KEEP-protected exactly as before.
    """
    if not delta or delta.get("pattern") != "value_tweaked":
        return None
    modified = delta.get("modified_cells") or 0
    shared = delta.get("shared_values") or 0
    denom = shared + modified
    if modified <= 3 or (denom and modified / denom <= 0.02):
        return "copy_then_edit"
    return "block_edit"


def _decimal_tail_signature(v, min_tail_digits=5, skip_decimal_digits=1):
    """Return a low-order decimal fingerprint for copy-then-edit detection.

    A common manual-edit fingerprint is that the leading integer/decimal digit is
    changed while the long fractional tail is left intact. For example,
    0.808902488 -> 0.908902488 preserves ``08902488`` after the first decimal
    digit. Short displayed decimals are ignored so ordinary one-decimal grids do
    not become tail matches.
    """
    try:
        fv = abs(float(v))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(fv):
        return None
    s = f"{fv:.9f}".rstrip("0").rstrip(".")
    if "." not in s:
        return None
    frac = s.split(".", 1)[1]
    if len(frac) < skip_decimal_digits + min_tail_digits:
        return None
    tail = frac[skip_decimal_digits:]
    # Padded/quantized tails such as 00000 or 99999 have little forensic value.
    if len(set(tail)) <= 1:
        return None
    return tail


def _detect_decimal_tail_reuse_for_pair(
    ga,
    gb,
    *,
    min_tail_digits=5,
    skip_decimal_digits=1,
    min_matches=8,
):
    """Find one aligned block where values differ but decimal tails are reused.

    This is layout-tolerant: if a table is pasted a few rows lower/upper, matching
    cells still share the same (row_delta, col_delta). Grouping by that offset
    distinguishes a copied block from isolated coincidental tail matches.
    """
    inv = {}
    for kb, vb in gb.items():
        sig = _decimal_tail_signature(
            vb,
            min_tail_digits=min_tail_digits,
            skip_decimal_digits=skip_decimal_digits,
        )
        if sig:
            inv.setdefault(sig, []).append((kb, vb))

    by_offset = {}
    for ka, va in ga.items():
        sig = _decimal_tail_signature(
            va,
            min_tail_digits=min_tail_digits,
            skip_decimal_digits=skip_decimal_digits,
        )
        if not sig:
            continue
        matches = inv.get(sig) or []
        # A very frequent tail is usually a quantization artifact; do not let it
        # create a combinatorial cloud of weak matches.
        if len(matches) > 20:
            continue
        for kb, vb in matches:
            if math.isclose(float(va), float(vb), rel_tol=1e-9, abs_tol=1e-12):
                continue
            off = (kb[0] - ka[0], kb[1] - ka[1])
            by_offset.setdefault(off, []).append((ka, kb, float(va), float(vb), sig))

    if not by_offset:
        return None
    off, pairs = max(by_offset.items(), key=lambda kv: len(kv[1]))
    if len(pairs) < min_matches:
        return None
    pairs = sorted(pairs, key=lambda p: (p[0][0], p[0][1], p[1][0], p[1][1]))
    return {
        "offset": off,
        "pairs": pairs,
        "tail_match_count": len(pairs),
        "min_tail_digits": min_tail_digits,
        "skip_decimal_digits": skip_decimal_digits,
    }


def _decimal_tail_constant_transform(pairs):
    """True if the matched value pairs share a constant additive offset (vb = va + k) or a constant
    ratio (vb = va * r). That is a benign linear/derived relationship between the two sheets (a shift,
    rescale, or baseline correction that incidentally preserves the fractional tail), NOT the
    irregular leading-digit edit pattern the detector targets."""
    vp = [(va, vb) for _ka, _kb, va, vb, _sig in pairs if va is not None and vb is not None]
    if len(vp) < 3:
        return False

    def _constant(vals):
        lo, hi = min(vals), max(vals)
        return (hi - lo) <= 1e-4 * max(abs(lo), abs(hi), 1e-9)

    if _constant([vb - va for va, vb in vp]):
        return True
    ratios = [vb / va for va, vb in vp if va not in (None, 0)]
    return len(ratios) >= 3 and _constant(ratios)


_DT_FIXED_DENOM_MAX_N = 400
_DT_FIXED_DENOM_TOL = 1e-6
_DT_FIXED_DENOM_FRAC = 0.85
_DT_AXIS_MIN_N = 6
_DT_FEWTAIL_MIN_PAIRS = 12
_DT_FEWTAIL_DOMINANCE = 0.80
_DT_PERCOL_MIN_GROUP = 3
_DT_LOGLABEL_RE = re.compile(
    r"\b("
    r"titer|titre|cfu|pfu|growth|log ?2|log ?10|log10|log2|"
    r"nt50|ic50|ec50|dilution|dilut|fold|"
    r"od600|od|absorbance|copy|copies|copy number|viral load|"
    r"qpcr|rt-qpcr|pcr|ct|cq|cycle threshold"
    r")\b",
    re.I,
)


def _dt_is_fractional(v):
    return v is not None and abs(float(v) - round(float(v))) > 1e-6


def _dt_fixed_denominator(pairs):
    """Return a benign reason when most decimal-tail values are k/N rates."""
    vals = [
        float(v)
        for _ka, _kb, va, vb, _s in pairs
        for v in (va, vb)
        if v is not None and _dt_is_fractional(v)
    ]
    if len(vals) < 6:
        return None
    need = max(6, math.ceil(_DT_FIXED_DENOM_FRAC * len(vals)))
    for n in range(2, _DT_FIXED_DENOM_MAX_N + 1):
        hit = sum(
            1
            for v in vals
            if abs(v * n - round(v * n)) < _DT_FIXED_DENOM_TOL * max(1, abs(v) * n)
        )
        if hit >= need:
            return f"fixed_denominator:1/{n}"
    return None


def _dt_progression(seq):
    """Conservative arithmetic/geometric progression test for axis-like values."""
    vals = [float(v) for v in seq if v is not None]
    if len(vals) < _DT_AXIS_MIN_N:
        return False
    if len({round(v, 9) for v in vals}) < _DT_AXIS_MIN_N:
        return False
    diffs = [vals[i + 1] - vals[i] for i in range(len(vals) - 1)]
    increasing = all(d > 1e-12 for d in diffs)
    decreasing = all(d < -1e-12 for d in diffs)
    if not (increasing or decreasing):
        return False

    base = min(abs(d) for d in diffs)
    if base and all(
        abs(d / base - round(d / base)) < 1e-4 * max(1, abs(d / base))
        for d in diffs
    ):
        return True

    if all(v != 0 for v in vals) and (all(v > 0 for v in vals) or all(v < 0 for v in vals)):
        ratios = [vals[i + 1] / vals[i] for i in range(len(vals) - 1)]
        return (max(ratios) - min(ratios)) < 1e-3 * max(abs(max(ratios)), abs(min(ratios)), 1e-9)
    return False


def _dt_axis(pairs):
    a = [va for _ka, _kb, va, _vb, _s in sorted(pairs, key=lambda p: (p[0][0], p[0][1]))]
    b = [vb for _ka, _kb, _va, vb, _s in sorted(pairs, key=lambda p: (p[1][0], p[1][1]))]
    return _dt_progression(a) and _dt_progression(b)


def _dt_few_tails(pairs):
    if len(pairs) < _DT_FEWTAIL_MIN_PAIRS:
        return False
    tails = [str(s) for _ka, _kb, _va, _vb, s in pairs if s is not None]
    if not tails:
        return False
    top = max(Counter(tails).values())
    return top >= _DT_FEWTAIL_DOMINANCE * len(tails)


def _dt_per_column_constant(pairs):
    """Return per-column constant offset/ratio reason, or None."""
    groups = {}
    for ka, _kb, va, vb, _s in pairs:
        if va is None or vb is None:
            continue
        groups.setdefault(ka[1], []).append((float(va), float(vb)))
    groups = {c: g for c, g in groups.items() if len(g) >= _DT_PERCOL_MIN_GROUP}
    if len(groups) < 2:
        return None

    def _const(xs):
        lo, hi = min(xs), max(xs)
        return (hi - lo) <= 1e-4 * max(abs(lo), abs(hi), 1e-9)

    offsets = []
    for c, g in sorted(groups.items()):
        diffs = [vb - va for va, vb in g]
        ratios = [vb / va for va, vb in g if va]
        if _const(diffs):
            offsets.append("c%d:%+.4g" % (c, sum(diffs) / len(diffs)))
        elif len(ratios) == len(g) and _const(ratios):
            offsets.append("c%d:x%.4g" % (c, sum(ratios) / len(ratios)))
        else:
            return None
    return "per_column_constant:[%s]" % ",".join(offsets)


def _dt_label_values(v):
    if not v:
        return []
    if isinstance(v, str):
        return [v]
    if isinstance(v, (list, tuple, set)):
        return [str(x) for x in v if x is not None]
    return [str(v)]


def _dt_label_blob(labels):
    parts = []
    for lc in labels or ():
        if not isinstance(lc, dict):
            continue
        for key in ("column_labels", "row_labels", "nearby_labels", "text"):
            parts.extend(_dt_label_values(lc.get(key)))
    return " ".join(parts)


def _dt_log_dilution_candidate(pairs, labels):
    """Return a note-only reason for likely log/dilution integer shifts."""
    diffs = [
        float(vb) - float(va)
        for _ka, _kb, va, vb, _s in pairs
        if va is not None and vb is not None
    ]
    if len(diffs) < 6:
        return None
    near_int = sum(1 for d in diffs if abs(d - round(d)) < 1e-6)
    if near_int < 0.8 * len(diffs):
        return None
    return (
        "log_or_dilution_integer_shift_candidate"
        if _DT_LOGLABEL_RE.search(_dt_label_blob(labels))
        else None
    )


def _decimal_tail_low_reason(pairs):
    if _decimal_tail_constant_transform(pairs):
        return "constant_transform"
    return _dt_fixed_denominator(pairs) or _dt_per_column_constant(pairs)


def _decimal_tail_note_reason(pairs, labels=None):
    if _dt_axis(pairs):
        return "axis_progression"
    if _dt_few_tails(pairs):
        return "constant_fraction_tail"
    return _dt_log_dilution_candidate(pairs, labels)


def _column_cells(grid, c):
    """Row-ordered [(row, value)] for column ``c`` of a decimal grid."""
    return sorted(((r, v) for (r, cc), v in grid.items() if cc == c), key=lambda t: t[0])


def _is_axis_progression(grid, c, min_n=4, rel_tol=1e-4, geo_tol=1e-3):
    """True when column ``c`` is a swept axis: its values lie on an arithmetic
    (constant step) or geometric (constant ratio) progression in row order.

    Catches dose ladders / serial dilutions (1:3 → geometric), time / frequency /
    voltage sweeps (linear → arithmetic) and integer-step index axes. Gaps from
    dropped integer rows are tolerated by fitting against the row index. ``geo_tol``
    is looser than ``rel_tol`` so a serial dilution stored at 3 significant figures
    (33.3 / 11.1 / 3.70 …) still reads as geometric.

    Blind spot worth noting: a *measurement* column that happens to be an exact
    arithmetic/geometric ramp is indistinguishable from an axis here. That is rare in
    real data, and paperconan's within-column arithmetic/geometric detectors flag such
    a column HIGH independently — so a copied exact-progression column is not silenced
    overall, only this one cross-sheet finding would be downgraded.
    """
    cells = _column_cells(grid, c)
    if len(cells) < min_n:
        return False
    rs = [r for r, _ in cells]
    vs = [v for _, v in cells]
    span = rs[-1] - rs[0]
    if span <= 0:
        return False
    # arithmetic: v linear in row index, non-flat
    step = (vs[-1] - vs[0]) / span
    if abs(step) > 1e-12:
        scale = max(abs(v) for v in vs) or 1.0
        if all(abs(v - (vs[0] + step * (r - rs[0]))) <= rel_tol * scale for r, v in cells):
            return True
    # geometric: same-sign nonzero values that are linear in log space
    if all(v != 0 for v in vs) and (all(v > 0 for v in vs) or all(v < 0 for v in vs)):
        logs = [math.log(abs(v)) for v in vs]
        lstep = (logs[-1] - logs[0]) / span
        if abs(lstep) > 1e-9:
            if all(abs(lg - (logs[0] + lstep * (r - rs[0]))) <= geo_tol for (r, _), lg in zip(cells, logs)):
                return True
    return False


def _axis_columns(grids, recur_min=3):
    """Classify, per (file, sheet), which columns are 'axis-like' so a cross-sheet
    overlap that lands only on them can be recognized as a shared-x-axis artifact.

    A column is axis-like if either:
      (A) its values form an arithmetic/geometric progression (a swept axis), or
      (B) its exact value-set recurs as a column across >= ``recur_min`` distinct
          (file, sheet) grids — i.e. the same axis was reused across many panels.
    """
    # (B) fingerprint columns by their value-set; count how many sheets carry each.
    fp_counts = Counter()
    col_fps = {}
    for key, grid in grids.items():
        cols = {c for (_, c) in grid}
        for c in cols:
            vals = frozenset(v for (r, cc), v in grid.items() if cc == c)
            if len(vals) >= 4:
                col_fps[(key, c)] = vals
                fp_counts[vals] += 1
    recurring = {fp for fp, n in fp_counts.items() if n >= recur_min}

    axis = {}
    for key, grid in grids.items():
        cols = {c for (_, c) in grid}
        axis[key] = {c for c in cols
                     if _is_axis_progression(grid, c) or col_fps.get((key, c)) in recurring}
    return axis


def _text_cell(sheet, r, c):
    if sheet is None or r < 0 or c < 0 or r >= sheet.nrows or c >= sheet.ncols:
        return ""
    v = sheet.cell(r, c)
    if isinstance(v, str):
        return v.strip()
    return ""


def _label_context_for_matches(sheet, shared, max_labels=40):
    if sheet is None:
        return {"column_labels": [], "row_labels": [], "nearby_labels": [], "text": ""}
    col_labels, row_labels, nearby = [], [], []
    for (r, c), _v in shared[:max_labels]:
        for rr in range(max(0, r - 3), r):
            label = _text_cell(sheet, rr, c)
            if label:
                col_labels.append(label)
        for cc in range(max(0, c - 3), c):
            label = _text_cell(sheet, r, cc)
            if label:
                row_labels.append(label)
        for rr in range(max(0, r - 2), min(sheet.nrows, r + 3)):
            for cc in range(max(0, c - 2), min(sheet.ncols, c + 3)):
                label = _text_cell(sheet, rr, cc)
                if label:
                    nearby.append(label)

    def uniq(vals):
        out, seen = [], set()
        for val in vals:
            key = val.lower()
            if key not in seen:
                seen.add(key)
                out.append(val)
        return out[:max_labels]

    ctx = {
        "column_labels": uniq(col_labels),
        "row_labels": uniq(row_labels),
        "nearby_labels": uniq(nearby),
    }
    ctx["text"] = " ".join(ctx["column_labels"] + ctx["row_labels"] + ctx["nearby_labels"])
    return ctx


def _shared_cross_sheet_context(ctx_a, ctx_b, pattern, fraction):
    text_a = (ctx_a or {}).get("text", "")
    text_b = (ctx_b or {}).get("text", "")
    both_control = bool(
        _CONTROL_BASELINE_LABEL_RE.search(text_a)
        and _CONTROL_BASELINE_LABEL_RE.search(text_b)
    )
    either_axis = bool(_AXIS_CONTEXT_LABEL_RE.search(text_a) or _AXIS_CONTEXT_LABEL_RE.search(text_b))
    non_perfect = pattern != "perfect_dup" and (fraction is None or fraction < 0.9)
    reason = None
    if non_perfect and both_control:
        reason = "matched cells are labelled as shared control/baseline/reference context"
    elif non_perfect and either_axis:
        reason = "matched cells are labelled as shared axis/coordinate context"
    return {
        "shared_control_or_baseline": bool(non_perfect and both_control),
        "shared_axis_or_coordinate": bool(non_perfect and either_axis),
        "context_reason": reason,
    }


def detect_collisions(grids, profile="review", sheets=None):
    """Find pairs of tables (sheets and/or flat files) with many bit-identical decimal
    values at the SAME (row, col). Catches copy-then-edit data inconsistencies,
    whether the copy lives in another sheet of the same workbook or in a separate file.

    `grids` maps (file, sheet) -> grid (from _grid_from_rows). Returns one dict per
    suspicious pair, with file_a/file_b set so same-file and cross-file pairs are
    distinguishable.

    Severity is context-aware on two axes:

    - SAME figure id (e.g. exFig.6i ↔ exFig.6k-n): the expected combined-vs-individual
      re-plot, downgraded to "low" with an explanatory `context`.
    - SHARED AXIS: when the bit-identical (row,col) cells concentrate (>=80%) on a
      column that is a swept axis / serial-dilution ladder / index reused across panels,
      AND the rest of the table diverges (pattern != perfect_dup), the overlap is just a
      shared x-axis (dose / time / frequency) — downgraded to "low" with `axis_overlap`.
      A full-table duplicate (perfect_dup) is NOT downgraded by this rule.

    Cross-figure overlaps that survive both checks keep their base severity.
    """
    findings = []
    keys = list(grids.keys())
    axis_cols = _axis_columns(grids)
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            (fa, sa), (fb, sb) = keys[i], keys[j]
            ga, gb = grids[keys[i]], grids[keys[j]]
            size_a, size_b = len(ga), len(gb)
            smaller = min(size_a, size_b)
            if smaller < 5:
                continue
            same_file = fa == fb
            # label_a / label_b disambiguate sheets when the pair spans two files
            la = sa if same_file else f"{fa}::{sa}"
            lb = sb if same_file else f"{fb}::{sb}"
            scope = "sheets" if same_file else "files"

            fig_a, fig_b = figure_key(sa), figure_key(sb)
            same_figure = bool(same_file and fig_a and fig_b and fig_a == fig_b)
            context = None
            if same_figure:
                context = (f"both sheets belong to the same display item ({fig_a}); "
                           f"a combined panel and its per-replicate breakdown share data "
                           f"by design, so this overlap is expected, not a cross-experiment reuse")

            same_pos = sum(1 for k, v in ga.items() if k in gb and gb[k] == v)
            vals_a, vals_b = set(ga.values()), set(gb.values())
            same_val = len(vals_a & vals_b)

            ctx_fields = dict(figure_a=fig_a, figure_b=fig_b, same_figure=same_figure,
                              delta=_value_delta(ga, gb))
            if context:
                ctx_fields["context"] = context

            if same_pos >= max(6, smaller * 0.15):
                shared = [(k, v) for k, v in ga.items() if k in gb and gb[k] == v]
                examples = shared[:5]
                label_context_a = _label_context_for_matches((sheets or {}).get(keys[i]), shared)
                label_context_b = _label_context_for_matches((sheets or {}).get(keys[j]), shared)
                fraction_of_smaller = same_pos / smaller
                shared_context = _shared_cross_sheet_context(
                    label_context_a,
                    label_context_b,
                    ctx_fields["delta"]["pattern"],
                    fraction_of_smaller,
                )
                # Shared-axis downgrade: if the bit-identical cells concentrate on a
                # column that is a swept/recurring axis AND the rest diverges, this is a
                # shared x-axis, not cross-experiment reuse. A perfect_dup spans every
                # column (incl. measurements), so it is excluded and stays high.
                pair_axis = axis_cols.get(keys[i], set()) | axis_cols.get(keys[j], set())
                on_axis = sum(1 for (_, c), _ in shared if c in pair_axis)
                non_axis_shared = len(shared) - on_axis
                # Downgrade only when the overlap is essentially confined to axis
                # columns: >=80% of shared cells on an axis AND no more than a couple of
                # stray matches off-axis (absolute backstop, so a wide axis can't drag a
                # real measurement overlap under the ratio). A perfect_dup spans every
                # column and is excluded above.
                axis_overlap = (
                    not same_figure
                    and ctx_fields["delta"]["pattern"] != "perfect_dup"
                    and on_axis >= 0.8 * len(shared)
                    and non_axis_shared <= 3
                )
                if axis_overlap:
                    ctx_fields["axis_overlap"] = True
                    axis_note = ("the bit-identical cells fall on a shared x-axis column "
                                 "(serial-dilution dose, time/frequency sweep, or an index "
                                 "reused across panels), while the measured values differ — "
                                 "a shared axis, not cross-experiment data reuse")
                    ctx_fields["context"] = axis_note
                    ctx_fields["likely_benign"] = axis_note
                if same_figure or axis_overlap:
                    sev = "low"
                else:
                    sev = "high"
                findings.append(dict(
                    kind="cross_sheet_position_identical",
                    file=fa if same_file else f"{fa} + {fb}",
                    file_a=fa, file_b=fb, same_file=same_file,
                    sheet_a=la, sheet_b=lb,
                    size_a=size_a, size_b=size_b,
                    same_position_count=same_pos,
                    fraction_of_smaller=fraction_of_smaller,
                    label_context_a=label_context_a,
                    label_context_b=label_context_b,
                    shared_context=shared_context,
                    examples=[dict(row=k[0] + 1, col=k[1] + 1, value=v) for k, v in examples],
                    severity=sev,
                    **ctx_fields,
                    rule=f"{la} and {lb} share {same_pos}/{smaller} ({same_pos/smaller*100:.0f}%) decimal values at SAME (row,col) across 2 {scope}",
                ))
            elif same_val >= max(8, smaller * 0.4):
                examples = sorted(list(vals_a & vals_b))[:5]
                findings.append(dict(
                    kind="cross_sheet_value_overlap",
                    file=fa if same_file else f"{fa} + {fb}",
                    file_a=fa, file_b=fb, same_file=same_file,
                    sheet_a=la, sheet_b=lb,
                    size_a=size_a, size_b=size_b,
                    shared_value_count=same_val,
                    fraction_of_smaller=same_val / smaller,
                    examples=examples,
                    severity="low" if same_figure else "medium",
                    **ctx_fields,
                    rule=f"{la} and {lb} share {same_val} bit-identical decimal values ({same_val/smaller*100:.0f}% of smaller) across 2 {scope}",
                ))

            tail_min_matches = max(8, min(20, math.ceil(smaller * 0.03)))
            tail_reuse = _detect_decimal_tail_reuse_for_pair(
                ga,
                gb,
                min_matches=tail_min_matches,
            )
            if tail_reuse:
                pairs = tail_reuse["pairs"]
                cells_a = [(ka, va) for ka, _kb, va, _vb, _sig in pairs]
                cells_b = [(kb, vb) for _ka, kb, _va, vb, _sig in pairs]
                label_context_a = _label_context_for_matches((sheets or {}).get(keys[i]), cells_a)
                label_context_b = _label_context_for_matches((sheets or {}).get(keys[j]), cells_b)
                fraction_of_smaller = tail_reuse["tail_match_count"] / smaller
                off_r, off_c = tail_reuse["offset"]
                low_reason = None if same_figure else _decimal_tail_low_reason(pairs)
                note_reason = None
                if same_figure:
                    sev = "low"
                elif low_reason:
                    # Strong benign decimal-tail structures: constant transform,
                    # fixed-denominator rates, or per-column shifts/ratios.
                    sev = "low"
                elif tail_reuse["tail_match_count"] >= 12 or fraction_of_smaller >= 0.10:
                    sev = "high"
                    note_reason = _decimal_tail_note_reason(pairs, (label_context_a, label_context_b))
                else:
                    sev = "medium"
                    note_reason = _decimal_tail_note_reason(pairs, (label_context_a, label_context_b))
                examples = [
                    {
                        "row_a": ka[0] + 1,
                        "col_a": ka[1] + 1,
                        "value_a": va,
                        "row_b": kb[0] + 1,
                        "col_b": kb[1] + 1,
                        "value_b": vb,
                        "decimal_tail": sig,
                    }
                    for ka, kb, va, vb, sig in pairs[:8]
                ]
                tail_fields = dict(ctx_fields)
                if same_figure and "context" not in tail_fields:
                    tail_fields["context"] = context
                tail_benign_reason = low_reason or note_reason
                if tail_benign_reason:
                    tail_fields["tail_benign_reason"] = tail_benign_reason
                findings.append(dict(
                    kind="cross_sheet_decimal_tail_reuse",
                    file=fa if same_file else f"{fa} + {fb}",
                    file_a=fa, file_b=fb, same_file=same_file,
                    sheet_a=la, sheet_b=lb,
                    size_a=size_a, size_b=size_b,
                    tail_match_count=tail_reuse["tail_match_count"],
                    fraction_of_smaller=fraction_of_smaller,
                    offset_rows=off_r,
                    offset_cols=off_c,
                    min_tail_digits=tail_reuse["min_tail_digits"],
                    skip_decimal_digits=tail_reuse["skip_decimal_digits"],
                    label_context_a=label_context_a,
                    label_context_b=label_context_b,
                    examples=examples,
                    severity=sev,
                    **tail_fields,
                    rule=(
                        f"{la} and {lb} share {tail_reuse['tail_match_count']}/{smaller} "
                        f"changed decimal cells with the same long fractional tail at "
                        f"offset ({off_r}, {off_c}) across 2 {scope}"
                    ),
                ))
    apply_profile_to_findings(findings, profile)
    return findings


def _column_axis_like(a):
    """True if a numeric column is an axis/index whose recurrence across panels is mundane: a
    (near-)constant column, a perfect ARITHMETIC progression (time/dose grid), or a perfect
    GEOMETRIC progression (serial-dilution axis) — the latter is legitimately shared across
    dose-response panels and must not read as a cross-experiment duplication."""
    if len(a) < 2:
        return True
    if len({round(float(v), 9) for v in a}) <= 1:
        return True                                   # constant
    scale = max(float(np.max(np.abs(a))), 1e-300)
    diffs = np.diff(a)
    if np.allclose(diffs, diffs[0], atol=1e-9 * scale, rtol=1e-9) and abs(diffs[0]) > 1e-9 * scale:
        return True                                   # arithmetic ladder
    if np.all(np.abs(a) > 1e-12):                     # geometric ladder (serial dilution)
        ratios = a[1:] / a[:-1]
        if np.allclose(ratios, ratios[0], atol=1e-9, rtol=1e-9) and abs(ratios[0] - 1) > 1e-9:
            return True
    return False


def detect_cross_sheet_column_duplicates(grid_sheets, profile="review", min_len=12):
    """B1 — full-column duplication ACROSS different (file, sheet) panels, including the
    integer / 1-decimal columns that `detect_collisions` misses (it grids only >=3-decimal
    values). Two panels that should be independent measurements carrying a byte-identical
    ordered column is a cross-experiment reuse fingerprint (e.g. a comet-assay 'No IR' column
    reproduced across two different figures). Same-figure panels are downgraded to low
    (a combined plot and its per-replicate breakdown legitimately share a column)."""
    if len(grid_sheets) < 2:
        return []                                     # a cross-panel duplicate needs >=2 panels
    # 1) collect candidate columns, deduped to the longest per (file, sheet, col_idx)
    best = {}
    for (fname, sname), sheet in grid_sheets.items():
        for (r0, r1, c0, c1) in find_numeric_blocks(sheet):
            header = header_for(sheet, r0, c0, c1)
            for c in range(c0, c1):
                a = col_array(sheet, r0, r1, c)
                a = a[~np.isnan(a)]
                if len(a) < min_len or _column_axis_like(a):
                    continue
                if len({round(float(v), 9) for v in a}) < max(6, len(a) // 2):
                    continue                          # low-cardinality column recurs benignly
                key = (fname, sname, c)
                if key not in best or len(a) > len(best[key][3]):
                    best[key] = (fname, sname, header[c - c0], a)

    # 2) bucket by exact rounded value-sequence; a bucket with >=2 distinct panels is a dup
    buckets = {}
    for (fname, sname, c), (_f, _s, label, a) in best.items():
        sig = tuple(round(float(v), 6) for v in a)
        buckets.setdefault(sig, []).append((fname, sname, c, label, a))

    findings = []
    for sig, group in buckets.items():
        panels = {(g[0], g[1]) for g in group}
        if len(panels) < 2:
            continue                                  # same-panel identical cols are identical_column's job
        a = group[0][4]
        n = len(a)
        all_int = all(abs(float(v) - round(float(v))) < 1e-9 for v in a)
        # all-integer sequences recur far more benignly (counts, indices) → require length + variety
        if all_int and (n < 25 or len({round(float(v), 9) for v in a}) < max(12, int(0.7 * n))):
            continue
        emitted = 0
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                fa, sa_name, _ca, la, _a = group[i]
                fb, sb_name, _cb, lb, _b = group[j]
                if (fa, sa_name) == (fb, sb_name):
                    continue                          # different columns, same sheet → identical_column
                fig_a, fig_b = figure_key(sa_name), figure_key(sb_name)
                same_figure = fig_a is not None and fig_a == fig_b
                same_file = fa == fb
                scope = "sheets" if same_file else "files"
                sev = "low" if (same_figure or all_int) else "high"
                findings.append(dict(
                    kind="cross_sheet_column_duplicate",
                    file=fa if same_file else f"{fa} + {fb}",
                    file_a=fa, file_b=fb, same_file=same_file,
                    sheet_a=sa_name, sheet_b=sb_name,
                    col_a=la, col_b=lb,
                    size_a=n, size_b=n,
                    same_position_count=n,
                    fraction_of_smaller=1.0,
                    figure_a=fig_a, figure_b=fig_b, same_figure=same_figure,
                    delta={"pattern": "column_duplicate"},
                    examples=[{"value": float(v)} for v in a[:5]],
                    severity=sev,
                    rule=(f"column '{la}' ({sa_name}) and column '{lb}' ({sb_name}) match to 6 decimal "
                          f"places over all {n} values across 2 {scope}"),
                ))
                emitted += 1
                if emitted >= 10:                     # cap per bucket
                    break
            if emitted >= 10:
                break
    apply_profile_to_findings(findings, profile)
    return findings


def _vector_is_patterned(vec):
    """Reject low-information tuples whose recurrence is mundane: near-constant, arithmetic or
    geometric ladders, and all-round-number (multiples of 10 / boundary) tuples."""
    if len({round(v, 6) for v in vec}) < 3:
        return True                                   # near-constant / too few distinct
    d = [vec[i + 1] - vec[i] for i in range(len(vec) - 1)]
    if all(abs(x - d[0]) < 1e-9 for x in d):
        return True                                   # arithmetic ladder
    nz = [v for v in vec if abs(v) > 1e-12]
    if len(nz) == len(vec):
        rat = [vec[i + 1] / vec[i] for i in range(len(vec) - 1)]
        if all(abs(r - rat[0]) < 1e-9 for r in rat):
            return True                               # geometric ladder
    if all(abs(v - round(v / 10.0) * 10.0) < 1e-9 for v in vec):
        return True                                   # all multiples of 10
    return False


# Per-row cell cap for the within-row pass: bounds the O(width * k) window build so one huge
# row (a wide genomics matrix) cannot balloon memory before the budget check fires.
_WR_MAX_ROW_CELLS = int(os.environ.get("PAPERCONAN_WR_MAX_ROW_CELLS", "20000"))


def detect_recurring_row_vectors(grid_sheets, profile="review",
                                 min_k=4, max_k=8, max_rows=300, max_findings=20):
    """B2 — a fixed ordered numeric tuple recurring as a contiguous row-slice across >=3 places
    spanning >=2 figure namespaces. Six independent mice cannot yield the identical six-value
    vector in several arms; a specific high-information tuple reappearing across unrelated figures
    is a copy fingerprint. Guarded hard (this is the most FP-prone pass): >=3 distinct values, no
    arithmetic/geometric/round-number ladders, >=3 occurrences in >=2 figure namespaces, and
    all-integer tuples need k>=5 with >=4 distinct values."""
    # The cross-FIGURE finding needs >=2 distinct figure namespaces; the within-ROW sibling
    # (a segment repeated inside one row) does not. Build the window index whenever EITHER is
    # possible — only the truly trivial single-sheet, unnamed corpus with no wide rows is
    # skipped (bounded by the budget below regardless).
    cross_figure_possible = (
        len({figure_key(s) for (_f, s) in grid_sheets if figure_key(s) is not None}) >= 2)
    occ = {}   # rounded tuple -> list of (file, sheet, figure_key, row, start_col)
    budget = 3_000_000   # bound worst-case work on genome-scale papers (linear, but many blocks)
    # The window index feeds ONLY the cross-figure path; skip building it entirely otherwise
    # (the within-row pass below scans rows directly).
    for (fname, sname), sheet in (grid_sheets.items() if cross_figure_possible else ()):
        fk = figure_key(sname)
        for (r0, r1, c0, c1) in find_numeric_blocks(sheet):
            for r in range(r0, min(r1, r0 + max_rows)):
                row = []
                for c in range(c0, c1):
                    v = sheet.cell(r, c)
                    row.append(float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None)
                for start in range(len(row)):
                    for k in range(min_k, max_k + 1):
                        window = row[start:start + k]
                        if len(window) < k or any(w is None for w in window):
                            continue
                        key = tuple(round(w, 6) for w in window)
                        occ.setdefault(key, []).append((fname, sname, fk, r, c0 + start))
                        budget -= 1
                if budget <= 0:
                    break
            if budget <= 0:
                break
        if budget <= 0:
            break

    cands = []
    for vec, places in (occ.items() if cross_figure_possible else ()):
        if len(places) < 3 or _vector_is_patterned(list(vec)):
            continue
        namespaces = {p[2] for p in places if p[2] is not None}
        if len(namespaces) < 2:
            continue                                  # recurrence within one figure is expected
        all_int = all(abs(v - round(v)) < 1e-9 for v in vec)
        if all_int and (len(vec) < 5 or len({round(v, 6) for v in vec}) < 4):
            continue
        # distinct (sheet,row) occurrences so the same cells aren't counted twice
        sites = {(p[0], p[1], p[3]) for p in places}
        if len(sites) < 3:
            continue
        # cells physically covered by all occurrences (file, sheet, row, col) — used to merge the
        # many overlapping windows that a single long recurring row-run produces. The file is part
        # of the key so two files that share a sheet name ('Sheet1') are not conflated.
        cells = {(p[0], p[1], p[3], p[4] + off) for p in places for off in range(len(vec))}
        cands.append((vec, places, namespaces, sites, cells))

    # Dedup by occurrence-cell overlap: a long recurring row-segment yields many overlapping
    # windows (k=4..8, shifted) at the SAME cells. Keep the strongest (most occurrences, then
    # longest) and drop any candidate whose covered cells overlap >=50% with a kept one, so one
    # physical recurring run reports as one finding.
    cands.sort(key=lambda x: (-len(x[3]), -len(x[0])))
    kept = []
    for c in cands:
        cells = c[4]
        if any(len(cells & k[4]) >= 0.5 * min(len(cells), len(k[4])) for k in kept):
            continue
        kept.append(c)

    findings = []
    for vec, places, namespaces, sites, _cells in kept:
        sheets_hit = sorted({p[1] for p in places})
        loc = "; ".join(sheets_hit[:6])
        files_hit = sorted({p[0] for p in places})
        findings.append(dict(
            kind="recurring_row_vector",
            file="; ".join(files_hit)[:120],
            file_a=files_hit[0], file_b=files_hit[-1], same_file=len(files_hit) == 1,
            sheet="; ".join(sheets_hit)[:120],
            sheet_a=sheets_hit[0], sheet_b=sheets_hit[-1],
            vector=[float(v) for v in vec],
            size_a=len(sites), size_b=len(sites),
            same_position_count=len(sites),
            fraction_of_smaller=1.0,
            n_occurrences=len(sites),
            n_figures=len(namespaces),
            same_figure=False,
            delta={"pattern": "recurring_row_vector"},
            pattern="recurring_row_vector",
            examples=[{"value": float(v)} for v in vec],
            severity="high" if (len(vec) >= 5 and len(sites) >= 3) else "medium",
            rule=(f"the {len(vec)}-value vector {list(vec)} recurs at {len(sites)} places across "
                  f"{len(namespaces)} figures ({loc})")))
        if len(findings) >= max_findings:
            break

    # Within-ROW member of the same family: the identical high-precision segment appearing at
    # >=2 NON-OVERLAPPING positions of ONE row (two cohorts of a row carrying the same tuple —
    # JCI196944 Fig S2H's CNO row). Scanned per-row DIRECTLY, not via the block index above: a
    # sparse sub-panel (S2H sits in columns the grid segmentation never blocks) would otherwise
    # be invisible. Same gates as the cross-figure path; the repeat corroborates itself, so one
    # figure namespace is enough.
    wr_budget = 2_000_000
    wr_cands = []
    for (fname, sname), sheet in grid_sheets.items():
        fk = figure_key(sname)
        for r in range(sheet.nrows):
            seq = []
            for c in range(sheet.ncols):
                v = sheet.cell(r, c)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    seq.append((c, float(v)))
                    if len(seq) >= _WR_MAX_ROW_CELLS:  # cap width so one huge row can't OOM
                        break
            if len(seq) < 2 * min_k:
                continue
            # per-row value frequency, keyed with the SAME quantization as the window (round-6)
            # so bucket and key agree at every magnitude: a value from a small quantized pool
            # (k/19 grid, dose plateau) recurs far more than the copies, unlike a copied segment.
            row_freq = Counter(round(v, 6) for _c, v in seq)
            wins = {}
            for start in range(len(seq)):
                for k in range(min_k, max_k + 1):
                    if start + k > len(seq):
                        break
                    wins.setdefault(tuple(round(seq[start + o][1], 6) for o in range(k)),
                                    []).append(start)
                    wr_budget -= 1
                if wr_budget <= 0:                     # gate INSIDE the loop, not per-row
                    break
            for vec, starts in wins.items():
                if len(starts) < 2 or _vector_is_patterned(list(vec)) or len(set(vec)) < 3:
                    continue
                all_int = all(abs(v - round(v)) < 1e-9 for v in vec)
                if all_int and (len(vec) < 5 or len(set(vec)) < 4):
                    continue
                chosen, last_end = [], -1
                for s in sorted(set(starts)):
                    if s >= last_end:                 # non-overlapping occurrences only
                        chosen.append(s)
                        last_end = s + len(vec)
                # quantized-pool signature is freq >> copies (not merely > copies): suppress only
                # when a value recurs beyond twice the copy count, else a genuine repeat whose
                # value happens to appear a couple extra times in a wide row is wrongly dropped.
                if len(chosen) >= 2 and max(row_freq[v] for v in vec) <= 2 * len(chosen):
                    cells = {(fname, sname, r, seq[s + o][0])
                             for s in chosen for o in range(len(vec))}
                    wr_cands.append((vec, fname, sname, fk, r, chosen, cells))
            if wr_budget <= 0:
                break
        if wr_budget <= 0:
            break
    if wr_budget <= 0:
        print("[paperconan] detect_recurring_row_vectors: within-row coverage bounded",
              file=sys.stderr)
    # One physical repeat yields many overlapping windows (k=4..8) — keep the strongest (most
    # copies, then longest) per row, dropping >=50%-cell-overlap duplicates.
    wr_cands.sort(key=lambda x: (-len(x[5]), -len(x[0])))
    wr_kept = []
    for c in wr_cands:
        if any(c[1:5] == kc[1:5] and len(c[6] & kc[6]) >= 0.5 * min(len(c[6]), len(kc[6]))
               for kc in wr_kept):
            continue
        wr_kept.append(c)
    for vec, fn, sn, fk, r, chosen, _cells in wr_kept:
        findings.append(dict(
            kind="within_row_repeated_segment",
            file=fn, file_a=fn, file_b=fn, same_file=True,
            sheet=sn, sheet_a=sn, sheet_b=sn, same_sheet=True,
            vector=[float(v) for v in vec],
            # same_position_count = matched CELLS (values x copies), consistent with the sibling
            # cross-sheet kinds, so evidence weighting isn't driven by the bare copy count.
            size_a=len(vec) * len(chosen), size_b=len(vec) * len(chosen),
            same_position_count=len(vec) * len(chosen),
            fraction_of_smaller=1.0, n_occurrences=len(chosen),
            figure_a=fk, figure_b=fk, same_figure=fk is not None,
            delta={"pattern": "within_row_repeat"}, pattern="within_row_repeat",
            examples=[{"value": float(v)} for v in vec],
            severity="high",
            rule=(f"the {len(vec)}-value segment {[round(float(v), 6) for v in vec]} repeats at "
                  f"{len(chosen)} non-overlapping positions within one row of {sn}")))
        if len(findings) >= max_findings:
            break

    apply_profile_to_findings(findings, profile)
    return findings


def detect_within_sheet_fraction_reuse(grid_sheets, profile="review", min_cells=10):
    """B3 — two numeric blocks in the SAME sheet whose positionally-corresponding cells reproduce
    each other's HIGH-PRECISION decimal fractions while their integer parts differ by whole numbers
    (e.g. two dose-response matrices where every cell shares the 5-decimal fraction but the value
    was shifted by an integer). detect_relations only compares columns within one block and
    detect_collisions only compares distinct sheets, so this matrix-to-matrix within-sheet reuse
    has no other detector. The precision + integer-shift + coverage requirements make chance
    coincidence negligible."""
    findings = []
    for (fname, sname), sheet in grid_sheets.items():
        grids = []
        for (r0, r1, c0, c1) in find_numeric_blocks(sheet):
            cells = {}
            for r in range(r0, r1):
                for c in range(c0, c1):
                    v = sheet.cell(r, c)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        cells[(r - r0, c - c0)] = float(v)
            if len(cells) >= min_cells:
                grids.append(((r0, r1, c0, c1), cells))
        best = None                                            # keep only the strongest pair per sheet
        for i in range(len(grids)):
            for j in range(i + 1, len(grids)):
                (ba, ca), (bb, cb) = grids[i], grids[j]
                common = [k for k in ca if k in cb]
                if len(common) < min_cells:
                    continue
                shared = int_diffs = hp = 0
                fracs, diffset = set(), set()
                for k in common:
                    x, y = ca[k], cb[k]
                    d = y - x
                    # Per-cell tolerance at THIS cell's magnitude, not the block-wide max. A single
                    # extreme value (a huge integer coordinate like distanceToTSS ~3.6e6, or a
                    # placeholder) must not inflate the tolerance so that every cell reads as an
                    # integer difference — that produced spurious whole-block fraction_reuse (M2-1).
                    tol = 1e-6 * max(abs(x), abs(y), 1.0)
                    if abs(d - round(d)) < tol:                 # integer difference => same fraction
                        shared += 1
                        if abs(round(d)) >= 1:
                            int_diffs += 1
                            diffset.add(round(d))
                        if _sig_frac_digits(x) >= 3:
                            hp += 1
                            fracs.add(round(x - round(x), 6))
                if (shared >= max(min_cells, int(round(0.8 * len(common))))
                        and hp >= max(6, int(round(0.5 * len(common))))
                        and int_diffs >= 3 and len(diffset) >= 2 and len(fracs) >= 5
                        and (best is None or shared > best[0])):
                    best = (shared, ba, bb, len(common))
        if best is not None:
            shared, ba, bb, ncommon = best
            findings.append(dict(
                kind="within_table_fraction_reuse",
                file=fname, file_a=fname, file_b=fname, same_file=True,
                sheet_a=sname, sheet_b=sname,
                size_a=ncommon, size_b=ncommon,
                same_position_count=shared,
                fraction_of_smaller=shared / ncommon,
                # both blocks live in ONE sheet, so there is no "two figures" to compare — leave
                # figure_a/b unset (None) rather than equal-but-not-same_figure (contradictory).
                figure_a=None, figure_b=None, same_figure=False,
                delta={"pattern": "fraction_reuse"},
                block_a=f"rows {ba[0]+1}-{ba[1]}, cols {ba[2]+1}-{ba[3]}",
                block_b=f"rows {bb[0]+1}-{bb[1]}, cols {bb[2]+1}-{bb[3]}",
                severity="high",
                rule=(f"two blocks in '{sname}' share identical decimal fractions on "
                      f"{shared}/{ncommon} positionally-corresponding cells but differ "
                      f"by whole numbers")))
    apply_profile_to_findings(findings, profile)
    return findings


def _longest_identical_run(a, b, c0, c1):
    """Longest contiguous column run where a[c] == b[c] (bit-identical to a tight
    relative tolerance). `a`, `b` are full row slices (may contain NaN); a NaN in
    either breaks the run. Returns (run_length, x_values_in_run) or None."""
    best_len, best_start = 0, 0
    cur_len, cur_start = 0, 0
    for idx in range(c1 - c0):
        av, bv = a[idx], b[idx]
        if math.isnan(av) or math.isnan(bv) or abs(av - bv) > 1e-9 * max(abs(av), abs(bv), 1e-300):
            cur_len = 0
            continue
        if cur_len == 0:
            cur_start = idx
        cur_len += 1
        if cur_len > best_len:
            best_len, best_start = cur_len, cur_start
    if best_len == 0:
        return None
    return best_len, a[best_start:best_start + best_len].astype(float)


def _row_bands(sheet):
    """Maximal runs of consecutive DATA rows (>= _ROW_REL_MIN_COLS finite cells),
    split by header/blank rows. A 'band' is one cohort/condition block laid out with
    conditions in rows. Yields (r_start, r_end) half-open row ranges."""
    is_data = []
    for r in range(sheet.nrows):
        finite = int(np.count_nonzero(~np.isnan(sheet.numeric[r, :])))
        is_data.append(finite >= _ROW_REL_MIN_COLS)
    bands, start = [], None
    for r, d in enumerate(is_data):
        if d and start is None:
            start = r
        elif not d and start is not None:
            bands.append((start, r))
            start = None
    if start is not None:
        bands.append((start, sheet.nrows))
    return bands


def _scaled_row_candidates(grid_sheets):
    """Collect high-information data-rows from row-oriented bands, tagged by band so
    same-band pairs (detect_row_relations' job) are excluded downstream."""
    cands = []
    for (fname, sname), sheet in grid_sheets.items():
        for bi, (r0, r1) in enumerate(_row_bands(sheet)):
            if (r1 - r0) < 2 or (r1 - r0) > _ROW_REL_MAX_ROWS:
                continue                                  # tall matrices are not this orientation
            for r in range(r0, r1):
                a = sheet.numeric[r, :]
                finite = a[~np.isnan(a)]
                if len(finite) < _ROW_REL_MIN_COLS or np.ptp(finite) <= 0 or len(np.unique(finite)) < 6:
                    continue
                if _vector_is_patterned(list(finite)):
                    continue                              # ladders / round-number rows recur benignly
                label = _row_label(sheet, r, 1)
                if _AXIS_CONTEXT_LABEL_RE.search(label):
                    continue
                cands.append(dict(file=fname, sheet=sname, band=(fname, sname, bi),
                                  rows=(r0, r1), row=r, label=label, a=a))
    return cands


def detect_scaled_row_reuse(grid_sheets, profile="review", max_candidates=1500,
                            max_findings=40):
    """Two DATA ROWS in DIFFERENT blocks (cross-block within a sheet) or different
    sheets that hold `row_B == row_A * k` over a long contiguous run of positionally-
    aligned columns — the scalar-multiple case (k != 1, `scaled_row_reuse`) and its
    k == 1 special case, a bit-identical data group reappearing under another cohort
    (`identical_row_reuse`).

    The Extended Data Fig. 5B pattern: the same condition measured under two
    treatments, where one cohort's row is an exact scalar multiple of the other's
    across ~200 cells. detect_row_relations only compares rows inside ONE block, so
    this cross-block reuse has no other detector. A whole power-of-ten ratio is a
    unit/percentage restatement (benign); an arbitrary constant is unexplained. An
    identical row across cohorts/figures may be a disclosed shared control — confirm
    against the legend.
    """
    cands = _scaled_row_candidates(grid_sheets)
    truncated = len(cands) > max_candidates
    if truncated:
        cands = cands[:max_candidates]
    findings = []
    budget = 4_000_000
    for i in range(len(cands)):
        A = cands[i]
        for j in range(i + 1, len(cands)):
            B = cands[j]
            if A["band"] == B["band"]:
                continue                                  # same block → detect_row_relations
            a, b = A["a"], B["a"]
            m = min(len(a), len(b))
            budget -= m
            if budget <= 0:
                break
            # Prefer the identical run (k==1) when it is at least as long as the best
            # scaling run — an exact duplicate is a cleaner statement than a ratio.
            ident = _longest_identical_run(a[:m], b[:m], 0, m)
            ratio = _longest_constant_ratio_run(a[:m], b[:m], 0, m)
            ident_ok = ident is not None and ident[0] >= _ROW_REL_MIN_COLS and len(np.unique(ident[1])) >= 6
            ratio_ok = ratio is not None and ratio[1] >= _ROW_REL_MIN_COLS and len(np.unique(ratio[2])) >= 6
            if ident_ok and (not ratio_ok or ident[0] >= ratio[1]):
                kind, k, run_len, x_run = "identical_row_reuse", 1.0, ident[0], ident[1]
            elif ratio_ok:
                kind, k, run_len, x_run = "scaled_row_reuse", ratio[0], ratio[1], ratio[2]
            else:
                continue
            fa, fb = A["file"], B["file"]
            sa_name, sb_name = A["sheet"], B["sheet"]
            fig_a, fig_b = figure_key(sa_name), figure_key(sb_name)
            same_file = fa == fb
            same_sheet = same_file and sa_name == sb_name
            scope = "blocks" if same_sheet else ("sheets" if same_file else "files")
            rel = (f"== row '{B['label']}'" if kind == "identical_row_reuse"
                   else f"= row '{B['label']}' ({sb_name}) * {k:.6g}")
            findings.append(dict(
                kind=kind,
                file=fa if same_file else f"{fa} + {fb}",
                file_a=fa, file_b=fb, same_file=same_file, same_sheet=same_sheet,
                sheet_a=sa_name, sheet_b=sb_name,
                row_a=A["label"], row_b=B["label"],
                size_a=run_len, size_b=run_len,
                same_position_count=run_len,
                fraction_of_smaller=1.0,
                ratio=k, run_length=run_len,
                figure_a=fig_a, figure_b=fig_b,
                same_figure=(fig_a is not None and fig_a == fig_b),
                delta={"pattern": "identical_row" if kind == "identical_row_reuse" else "scaled_row"},
                block_a=f"rows {A['rows'][0] + 1}-{A['rows'][1]}",
                block_b=f"rows {B['rows'][0] + 1}-{B['rows'][1]}",
                examples=[{"row": A["label"], "col": None, "value": float(v)}
                          for v in x_run[:5]],
                severity="high",
                rule=(f"row '{A['label']}' ({sa_name}) {rel} over a run of {run_len} "
                      f"positionally-aligned columns across 2 {scope}")))
            if len(findings) >= max_findings:
                break
        if budget <= 0 or len(findings) >= max_findings:
            break
    if truncated or budget <= 0:
        # Never silently cap coverage — say what was bounded (stderr only; scan.json stays
        # deterministic). Real condition-layout papers stay far under these limits.
        print(f"[paperconan] detect_scaled_row_reuse: coverage bounded "
              f"(candidates={len(cands)}{'+truncated' if truncated else ''}, "
              f"budget_exhausted={budget <= 0})", file=sys.stderr)
    apply_profile_to_findings(findings, profile)
    return findings


def _sigfigs_and_frac(v):
    """(significant-figure count, fractional-digit count) of |v| at 10-decimal read
    precision. 169.8665 -> (7, 4); 0.95705 -> (5, 5); a 5-digit INTEGER 10234 -> (5, 0)."""
    av = abs(float(v))
    if not math.isfinite(av) or av == 0.0:
        return 0, 0
    s = f"{av:.10f}".rstrip("0")
    ip, _, fr = s.partition(".")
    sig = len(fr.lstrip("0")) if ip == "0" else len(ip) + len(fr)
    return sig, len(fr)


def _is_short_hp(v):
    """A value is 'high-precision' for short-run matching only if it carries real
    FRACTIONAL precision — >=3 fractional digits AND >=5 significant figures. Requiring a
    fractional part is what keeps collision-prone INTEGER data (read counts, IDs, genomic
    coordinates) — where a 3-cell match is easy chance — out of the detector entirely."""
    if math.isnan(v):
        return False
    sig, frac = _sigfigs_and_frac(v)
    return frac >= 3 and sig >= _SHORT_ROW_MIN_SIGFIGS


def _near_power_of_ten(k):
    """True if k is a whole power of ten to the SAME relative tolerance a short ratio run is
    accepted at. `_is_round_power_of_ten` only matches to 1e-9, but a run holds to
    _SHORT_ROW_RTOL, so a unit conversion between two panels stored at different decimal
    precision yields a mean ratio ~1e-5 off an exact power of ten — still a benign
    restatement, not two independent measurements. Checked both k and 1/k directions."""
    ak = abs(float(k))
    if ak <= 1e-300 or not math.isfinite(ak):
        return False
    e = round(math.log10(ak))
    return abs(ak - 10.0 ** e) <= _SHORT_ROW_RTOL * (10.0 ** e)


def _longest_hp_identical_run(a, b):
    """Longest contiguous run where a[c] == b[c] (tight tol) AND both cells are
    high-precision (>=5 sig figs). A low-precision or NaN column breaks the run, so a
    coincidental match on small integers can never extend one. Returns (run_len, x_run)."""
    best_len, best_start = 0, 0
    cur_len, cur_start = 0, 0
    for i in range(len(a)):
        av, bv = a[i], b[i]
        if (not _is_short_hp(av) or not _is_short_hp(bv)
                or abs(av - bv) > 1e-9 * max(abs(av), abs(bv), 1e-300)):
            cur_len = 0
            continue
        if cur_len == 0:
            cur_start = i
        cur_len += 1
        if cur_len > best_len:
            best_len, best_start = cur_len, cur_start
    if best_len == 0:
        return None
    return best_len, a[best_start:best_start + best_len].astype(float)


def _longest_hp_ratio_run(a, b):
    """Longest contiguous run where b[c] == k * a[c] for a fixed k != 1, every cell
    high-precision, k held to _SHORT_ROW_RTOL. Returns (k, run_len, x_run) or None."""
    best_len, best_start, best_k = 0, 0, None
    cur_len, cur_k, cur_start = 0, None, 0
    for i in range(len(a)):
        av, bv = a[i], b[i]
        if not _is_short_hp(av) or not _is_short_hp(bv) or abs(av) <= 1e-12:
            cur_len, cur_k = 0, None
            continue
        r = bv / av
        if cur_k is None or abs(r - cur_k) > _SHORT_ROW_RTOL * max(abs(cur_k), 1e-300):
            cur_k, cur_len, cur_start = r, 1, i
        else:
            cur_len += 1
        if cur_len > best_len and abs(cur_k - 1.0) > _SHORT_ROW_RTOL:
            best_len, best_start, best_k = cur_len, cur_start, cur_k
    if best_len == 0 or best_k is None:
        return None
    x_run = a[best_start:best_start + best_len].astype(float)
    k = float(np.mean(b[best_start:best_start + best_len].astype(float) / x_run))
    if abs(k - 1.0) <= _SHORT_ROW_RTOL or abs(k) <= 1e-9:
        return None
    return k, best_len, x_run


def _longest_hp_offset_run(a, b):
    """Longest contiguous run where b[c] == a[c] + c for a fixed NON-zero constant c, every
    cell high-precision, c held to a tight absolute tolerance (a genuine copy+shift is exact).
    The row twin of `constant_offset`. Returns (c, run_len, x_run) or None."""
    best_len, best_start, best_c = 0, 0, None
    cur_len, cur_c, cur_start, cur_scale = 0, None, 0, 1.0
    for i in range(len(a)):
        av, bv = a[i], b[i]
        if not _is_short_hp(av) or not _is_short_hp(bv):
            cur_len, cur_c = 0, None
            continue
        d = bv - av
        # membership tolerance scales with the CELL magnitude (no fixed floor): for near-zero
        # rows a fixed 1e-4 floor is huge relative to the values, so unrelated small rows read
        # as a constant difference; a genuine copy+shift matches to read precision (~1e-9).
        tol = _SHORT_ROW_RTOL * max(abs(av), abs(bv), 1e-300)
        if cur_c is None or abs(d - cur_c) > tol:
            cur_c, cur_len, cur_start = d, 1, i
            cur_scale = max(abs(av), abs(bv), 1e-300)    # anchor magnitude, fixed for the run
        else:
            cur_len += 1
        # non-triviality is anchored to the run START magnitude, not the current cell — a
        # per-cell threshold flips as the run crosses magnitudes and truncates genuine runs
        # when a large cell lands at the tail.
        if cur_len > best_len and abs(cur_c) > _SHORT_ROW_RTOL * cur_scale:
            best_len, best_start, best_c = cur_len, cur_start, cur_c
    if best_len == 0 or best_c is None:
        return None
    x_run = a[best_start:best_start + best_len].astype(float)
    c = float(np.mean(b[best_start:best_start + best_len].astype(float) - x_run))
    if abs(c) <= 1e-9:
        return None
    return c, best_len, x_run


def _short_row_candidates(grid_sheets):
    """Every data row carrying >=_SHORT_ROW_MIN_COLS high-precision values with >=3
    DISTINCT such values — including ISOLATED single rows that `_scaled_row_candidates`
    drops (its bands need >=2 rows of >=12 finite cells). Grouped by sheet downstream."""
    cands = []
    for (fname, sname), sheet in grid_sheets.items():
        rows = []
        for r in range(sheet.nrows):
            a = sheet.numeric[r, :]
            hp = [v for v in a if _is_short_hp(v)]
            if len(hp) < _SHORT_ROW_MIN_COLS or len(set(hp)) < 3:
                continue
            if _vector_is_patterned(hp):
                continue
            label = _row_label(sheet, r, 1)
            if _AXIS_CONTEXT_LABEL_RE.search(label):
                continue
            rows.append(dict(file=fname, sheet=sname, row=r, label=label, a=a))
            if len(rows) >= _SHORT_ROW_MAX_ROWS_PER_SHEET:
                break
        cands.extend(rows)
    return cands


def detect_short_row_reuse(grid_sheets, profile="review", max_findings=60):
    """SHORT high-precision identical or constant-ratio runs (3..11 columns) between two
    data rows of one sheet — the JCI "Supporting Data Values" fingerprint that the >=12
    column `detect_scaled_row_reuse` and `detect_row_relations` cannot see: a control
    block copied verbatim across two sub-panels, a whole condition row shared by two
    different genes, or one panel's row = another's * a constant (e.g. Group B = 0.8409 *
    Group A). Every run cell must be >=5 significant figures so a short run is not chance.
    A whole power-of-ten ratio is a unit restatement (benign) and is skipped. Signal, not
    verdict — confirm against the legend/Methods before drawing any conclusion."""
    by_sheet = {}
    for c in _short_row_candidates(grid_sheets):
        by_sheet.setdefault((c["file"], c["sheet"]), []).append(c)
    findings = []
    budget = 4_000_000
    for (fname, sname), rows in by_sheet.items():
        fig = figure_key(sname)
        # Sheet-wide frequency of each high-precision value, BUCKETED to 5 significant
        # figures: a value shared by many rows is a quantized grid (k/19) or a fitted-curve
        # plateau (a saturated asymptote repeats across many consecutive rows, wobbling in
        # the last digit) — not a distinctive duplicate. The coarse bucket keeps a plateau's
        # 24.670713/24.670714 wobble in ONE bin so it is counted as common, not many rares.
        def _freq_key(v):
            return float(f"{float(v):.5g}")
        freq = Counter(_freq_key(v) for R in rows for v in R["a"] if _is_short_hp(v))

        def _rare(run):
            return all(freq.get(_freq_key(v), 0) <= _SHORT_ROW_MAX_VALUE_FREQ for v in run)

        # A smooth fitted curve (dose-response, binding) sampled along an axis makes every
        # consecutive row an approximate scalar multiple of the next — a benign ~1.01 step,
        # not a copied panel. Those rows sit in ONE contiguous data band; a genuine cross-
        # panel scaling (Group B = 0.84 * Group A) is separated by a header/blank row, i.e.
        # a DIFFERENT band. So a SCALED (ratio) pair with no non-data row between them is a
        # curve step and is dropped. Identical pairs are kept (a curve never repeats a row
        # verbatim; plateaus are already removed by the frequency gate).
        cand_idx = {R["row"] for R in rows}

        def _same_band(ra, rb):
            lo, hi = (ra, rb) if ra < rb else (rb, ra)
            return all(k in cand_idx for k in range(lo + 1, hi))

        for i in range(len(rows)):
            A = rows[i]
            for j in range(i + 1, len(rows)):
                B = rows[j]
                a, b = A["a"], B["a"]
                m = min(len(a), len(b))
                budget -= m
                if budget <= 0:
                    break
                ident = _longest_hp_identical_run(a[:m], b[:m])
                ratio = _longest_hp_ratio_run(a[:m], b[:m])
                offset = _longest_hp_offset_run(a[:m], b[:m])
                ident_ok = (ident is not None
                            and _SHORT_ROW_MIN_COLS <= ident[0] < _ROW_REL_MIN_COLS
                            and len(np.unique(ident[1])) >= 3
                            and _rare(ident[1]))
                ratio_ok = (ratio is not None
                            and _SHORT_ROW_MIN_COLS <= ratio[1] < _ROW_REL_MIN_COLS
                            and len(np.unique(ratio[2])) >= 3
                            and _rare(ratio[2])
                            and not _same_band(A["row"], B["row"])
                            and not _near_power_of_ten(ratio[0]))
                # Constant additive offset (B = A + c) — the row twin of constant_offset. Like
                # the scaled case, an adjacent same-band offset is a smooth-curve step (a linear
                # stretch of a curve), so it is suppressed.
                offset_ok = (offset is not None
                             and _SHORT_ROW_MIN_COLS <= offset[1] < _ROW_REL_MIN_COLS
                             and len(np.unique(offset[2])) >= 3
                             and _rare(offset[2])
                             and not _same_band(A["row"], B["row"]))
                # Prefer identical; among the two one-parameter relations pick the longer run.
                if ident_ok and ident[0] >= max(ratio[1] if ratio_ok else 0,
                                                offset[1] if offset_ok else 0):
                    kind, k, run_len, x_run = "identical_row_reuse", 1.0, ident[0], ident[1]
                elif offset_ok and (not ratio_ok or offset[1] >= ratio[1]):
                    kind, k, run_len, x_run = "offset_row_reuse", offset[0], offset[1], offset[2]
                elif ratio_ok:
                    kind, k, run_len, x_run = "scaled_row_reuse", ratio[0], ratio[1], ratio[2]
                else:
                    continue
                rel = ("== row '{}'".format(B["label"]) if kind == "identical_row_reuse"
                       else "= row '{}' + {:.6g}".format(B["label"], k)
                       if kind == "offset_row_reuse"
                       else "= row '{}' * {:.6g}".format(B["label"], k))
                findings.append(dict(
                    kind=kind, short_run=True,
                    file=fname, file_a=fname, file_b=fname,
                    same_file=True, same_sheet=True,
                    sheet_a=sname, sheet_b=sname,
                    row_a=A["label"], row_b=B["label"],
                    size_a=run_len, size_b=run_len, same_position_count=run_len,
                    fraction_of_smaller=1.0, run_length=run_len,
                    ratio=(k if kind == "scaled_row_reuse" else None),
                    offset=(k if kind == "offset_row_reuse" else None),
                    figure_a=fig, figure_b=fig, same_figure=fig is not None,
                    delta={"pattern": {"identical_row_reuse": "identical_row",
                                       "offset_row_reuse": "offset_row",
                                       "scaled_row_reuse": "scaled_row"}[kind]},
                    block_a=f"row {A['row'] + 1}", block_b=f"row {B['row'] + 1}",
                    examples=[{"row": A["label"], "col": None, "value": float(v)}
                              for v in x_run[:5]],
                    severity="high",
                    rule=(f"row '{A['label']}' {rel} over a short run of {run_len} "
                          f"high-precision columns in {sname}")))
                if len(findings) >= max_findings:
                    break
            if budget <= 0 or len(findings) >= max_findings:
                break
        if budget <= 0 or len(findings) >= max_findings:
            break
    if budget <= 0:
        print("[paperconan] detect_short_row_reuse: coverage bounded (budget exhausted)",
              file=sys.stderr)
    apply_profile_to_findings(findings, profile)
    return findings


# Within-row shared-fraction: a fractional tail this long, shared by two cells whose integer
# parts differ, is ~1e-6 by chance per pair — low enough that a single pair is worth a look
# and a multi-pair segment (a copied row-slice with integers rewritten) is near-certain.
_WITHIN_ROW_FRAC_MIN_DIGITS = int(os.environ.get("PAPERCONAN_WITHIN_ROW_FRAC_MIN_DIGITS", "6"))
# Row-PAIR shared fraction: a RUN of aligned columns corroborates the match, so each cell's
# tail can be shorter (>=4 digits) than the single-cell within-row bar (>=6) — a >=3 column
# run of distinct shared tails is ~(1e-4)^3 by chance.
_ROW_PAIR_FRAC_MIN_DIGITS = int(os.environ.get("PAPERCONAN_ROW_PAIR_FRAC_MIN_DIGITS", "4"))
_ROW_PAIR_MIN_RUN = int(os.environ.get("PAPERCONAN_ROW_PAIR_MIN_RUN", "3"))
# Per-sheet candidate-row cap bounds the O(rows^2) pair loop so one huge sheet cannot
# starve the rest (the budget is also reset per sheet, so sheet order never matters).
_ROW_PAIR_MAX_ROWS_PER_SHEET = int(os.environ.get("PAPERCONAN_ROW_PAIR_MAX_ROWS", "400"))


# --- shared-fraction primitives -------------------------------------------------------
# "Same high-precision decimal tail, different integer part" (copy-then-integer-shift) is
# one fingerprint that shows up in several orientations: two COLUMNS
# (integer_diff_shared_fraction / round_shift_shared_fraction), two BLOCKS
# (within_table_fraction_reuse), one ROW's cells (within_row_shared_fraction), and a ROW
# PAIR (detect_row_pair_shared_fraction). These two helpers are the shared substrate — a
# magnitude-safe tail extractor and the small-denominator/quantization gate — so the tail
# semantics and the false-positive gate are defined once.

def _reliable_frac_tail(av):
    """Fractional-digit string of |av| at a precision the magnitude can actually carry.
    float64 holds ~15 significant figures, so formatting the WHOLE value at a fixed 10
    decimals when the integer part is large prints representation NOISE as real tail digits
    (5000000.137 -> '1370000001'). Cap decimals at 15 - integer-digits so a short real tail
    stays short. Returns '' for an integer / no fractional part / non-finite input."""
    av = abs(float(av))
    if not math.isfinite(av):
        return ""
    prec = min(10, max(0, 15 - len(str(int(av)))))
    s = f"{av:.{prec}f}".rstrip("0")
    return s.split(".")[1] if "." in s else ""


def _shared_frac_is_small_denominator(frac, max_q=128):
    """True if the shared fractional VALUE is a simple fraction p/q for some small q. Any two
    values x + p/q and y + p/q (same residue, different integer) trivially share that tail —
    a small-denominator artifact (triplicate means .333/.667, k/13 .923076, k/19, dyadic
    1/128 …), NOT a copied fraction. A genuine copy-then-shift tail is an arbitrary
    high-entropy decimal with no small denominator. Guards the same trap as the tail-cluster
    and short-row detectors, per shared tail. `frac` may be the digit-string after the decimal
    point or the fractional value itself (any float; its |·| mod 1 is used)."""
    if isinstance(frac, str):
        f = float("0." + frac) if frac else 0.0
    else:
        f = abs(float(frac)) % 1.0
    if f == 0.0:
        return False
    for q in range(2, max_q + 1):
        if abs(f * q - round(f * q)) < 2e-6:
            return True
    return False


def detect_within_row_shared_fraction(grid_sheets, profile="review", max_findings=60):
    """Two cells of ONE row that share a long high-precision fractional tail while their
    integer parts differ (e.g. 20.316768 and 102.316768) — a copy-then-shift: a value or a
    whole row-slice reused with the integer part rewritten but the decimals left intact.
    `integer_diff_shared_fraction` / `round_shift_shared_fraction` only compare two COLUMNS
    and `within_table_fraction_reuse` only compares two BLOCKS, so a segment copied across
    the columns of a single row has no other detector. A shared >=6-digit tail across
    different integers is ~1e-6 by chance, so requiring that tail length is the FP control.
    Signal, not verdict — confirm against the legend/Methods before drawing a conclusion."""
    findings = []
    budget = 8_000_000
    for (fname, sname), sheet in grid_sheets.items():
        fig = figure_key(sname)
        for r in range(sheet.nrows):
            row = sheet.numeric[r, :]
            budget -= sheet.ncols
            if budget <= 0:
                break
            by_frac = {}
            for c in range(sheet.ncols):
                v = row[c]
                if math.isnan(v):
                    continue
                av = abs(float(v))
                if av >= 1e7:                         # read-precision noise reaches the tail
                    continue
                frac = _reliable_frac_tail(av)
                if len(frac) < _WITHIN_ROW_FRAC_MIN_DIGITS:
                    continue
                by_frac.setdefault(frac, []).append((c, float(v), int(av)))
            groups = [(frac, cells) for frac, cells in by_frac.items()
                      if len(cells) >= 2 and len({ip for _, _, ip in cells}) >= 2
                      and not _shared_frac_is_small_denominator(frac)]
            if not groups:
                continue
            groups.sort(key=lambda g: g[1][0][0])     # deterministic: by first column
            label = _row_label(sheet, r, 1)
            examples = []
            for frac, cells in groups[:5]:
                vs = [v for _, v, _ in cells[:3]]
                examples.append({"row": label, "col": None, "tail": frac,
                                 "values": [float(x) for x in vs]})
            sample = " / ".join(f"{v:.10g}" for _, v, _ in groups[0][1][:2])
            findings.append(dict(
                kind="within_row_shared_fraction",
                file=fname, file_a=fname, file_b=fname,
                same_file=True, same_sheet=True,
                sheet_a=sname, sheet_b=sname,
                row=label, row_a=label, row_b=label,
                n_groups=len(groups),
                # NOTE: size_a / same_position_count = number of shared-tail FAMILIES in the
                # row (what _distill_cross_sheet reads as `n`), not a count of shared cells.
                size_a=len(groups), same_position_count=len(groups),
                fraction_of_smaller=1.0,
                figure_a=fig, figure_b=fig, same_figure=fig is not None,
                delta={"pattern": "shared_fraction"},
                block_a=f"row {r + 1}", block_b=f"row {r + 1}",
                examples=examples,
                severity="high",
                rule=(f"row '{label}': {len(groups)} value pair(s) in the same row share a "
                      f">={_WITHIN_ROW_FRAC_MIN_DIGITS}-digit fractional tail but differ in "
                      f"the integer part (e.g. {sample}) — a copy-then-integer-shift pattern "
                      f"in {sname}")))
            if len(findings) >= max_findings:
                break
        if budget <= 0 or len(findings) >= max_findings:
            break
    if budget <= 0:
        print("[paperconan] detect_within_row_shared_fraction: coverage bounded "
              "(budget exhausted)", file=sys.stderr)
    apply_profile_to_findings(findings, profile)
    return findings


def detect_row_pair_shared_fraction(grid_sheets, profile="review", max_findings=60):
    """Two ROWS that share the same high-precision decimal fraction at a contiguous run of
    aligned columns while their integer parts differ — the row-oriented twin of
    `integer_diff_shared_fraction` (which only compares two COLUMNS). A concentration row
    reused at another concentration with the integer parts rewritten (JCI201090 C4-2: 20 nM
    and 100 nM share .27037/.85351/.86076 over 3 columns, integers 95/85, 90/88, 91/87) has
    no other detector. The run corroborates each cell match, so a >=3 column run of >=3
    DISTINCT non-small-denominator shared tails is near-zero chance. Signal, not verdict."""
    by_sheet = {}
    for (fname, sname), sheet in grid_sheets.items():
        rows = []
        for r in range(sheet.nrows):
            a = sheet.numeric[r, :]
            cells, hp = [], 0
            for c in range(sheet.ncols):
                v = a[c]
                av = abs(float(v)) if not math.isnan(v) else float("nan")
                if math.isnan(av) or av >= 1e7:
                    cells.append(None)
                    continue
                tail = _reliable_frac_tail(av)
                if len(tail) < _ROW_PAIR_FRAC_MIN_DIGITS:
                    cells.append(None)
                    continue
                cells.append((tail, int(av), float(v)))
                hp += 1
            if hp >= _ROW_PAIR_MIN_RUN:
                rows.append((r, _row_label(sheet, r, 1), cells))
                if len(rows) >= _ROW_PAIR_MAX_ROWS_PER_SHEET:
                    print(f"[paperconan] detect_row_pair_shared_fraction: {sname} candidate "
                          f"rows capped at {_ROW_PAIR_MAX_ROWS_PER_SHEET} — later rows not "
                          "compared", file=sys.stderr)
                    break
        if len(rows) >= 2:
            by_sheet[(fname, sname)] = rows
    findings = []
    budget = 40_000_000                       # global compute bound across all sheets
    for (fname, sname), rows in by_sheet.items():
        fig = figure_key(sname)
        for i in range(len(rows)):
            ri, la, ca = rows[i]
            for j in range(i + 1, len(rows)):
                rj, lb, cb = rows[j]
                m = min(len(ca), len(cb))
                budget -= m
                if budget <= 0:
                    break
                # ALL maximal contiguous runs of columns sharing a tail with different
                # integers (evaluating only the longest would let a benign long run mask a
                # genuine shorter one).
                runs, cur = [], []
                for c in range(m):
                    A, B = ca[c], cb[c]
                    if A is not None and B is not None and A[0] == B[0] and A[1] != B[1]:
                        cur.append((c, A[0], A[2], B[2]))
                    else:
                        if len(cur) >= _ROW_PAIR_MIN_RUN:
                            runs.append(cur)
                        cur = []
                if len(cur) >= _ROW_PAIR_MIN_RUN:
                    runs.append(cur)
                # keep the longest run that PASSES the gate: >=3 distinct non-small-
                # denominator shared tails AND >=2 distinct integer differences (the latter
                # matches the column twin integer_diff_shared_fraction so a pure constant
                # offset B = A + k stays with constant_offset, not a copy-shift).
                best = None
                for run in sorted(runs, key=len, reverse=True):
                    distinct_nonsd = {t for _, t, _, _ in run
                                      if not _shared_frac_is_small_denominator(t)}
                    int_diffs = {int(round(vb - va)) for _, _, va, vb in run}
                    if len(distinct_nonsd) >= 3 and len(int_diffs) >= 2:
                        best = run
                        break
                if best is None:
                    continue
                sample = ", ".join(f"{va:.10g}/{vb:.10g}" for _, _, va, vb in best[:3])
                findings.append(dict(
                    kind="shared_fraction_row_pair",
                    file=fname, file_a=fname, file_b=fname,
                    same_file=True, same_sheet=True,
                    sheet_a=sname, sheet_b=sname,
                    row_a=la, row_b=lb,
                    run_length=len(best),
                    size_a=len(best), same_position_count=len(best),
                    fraction_of_smaller=1.0,
                    figure_a=fig, figure_b=fig, same_figure=fig is not None,
                    delta={"pattern": "shared_fraction"},
                    block_a=f"row {ri + 1}", block_b=f"row {rj + 1}",
                    examples=[{"row": f"{la} / {lb}", "col": None, "tail": t,
                               "values": [float(va), float(vb)]}
                              for _, t, va, vb in best[:5]],
                    severity="high",
                    rule=(f"rows '{la}' and '{lb}' share the same decimal fraction at "
                          f"{len(best)} aligned columns while the integer part differs "
                          f"(e.g. {sample}) — a copy-then-integer-shift across two rows "
                          f"in {sname}")))
                if len(findings) >= max_findings:
                    break
            if budget <= 0 or len(findings) >= max_findings:
                break
        if budget <= 0 or len(findings) >= max_findings:
            break
    if budget <= 0:
        print("[paperconan] detect_row_pair_shared_fraction: coverage bounded "
              "(global budget exhausted)", file=sys.stderr)
    apply_profile_to_findings(findings, profile)
    return findings


def _load_provenance(in_dir, paper):
    """Resolve scan provenance: an explicit `paper` override wins; otherwise read a
    paperconan_source.json sidecar left by `fetch`; otherwise None."""
    if paper:
        return paper
    sidecar = os.path.join(in_dir, "paperconan_source.json")
    if os.path.isfile(sidecar):
        try:
            with open(sidecar, encoding="utf-8") as fh:
                return json.load(fh)
        except (OSError, ValueError):
            return None
    return None


# Per-file memory guard: workbooks above this size expand to many GB of Python objects
# when fully materialized, so they are skipped (recorded as oversized) before loading.
# Coarse byte backstop (generous — the precise guard is the cell-count cap below).
_MAX_FILE_MB = float(os.environ.get("PAPERCONAN_MAX_FILE_MB", "200"))
_MAX_FILE_BYTES = int(_MAX_FILE_MB * 1024 * 1024)
# Precise memory guard: the columnar substrate stores numeric cells in a dense float64
# array (~8 bytes/cell) instead of ~100-200 bytes/cell as Python objects, so a given cell
# budget now bounds far less RAM. Skip a sheet whose cell count exceeds this, checked from
# the sheet dimensions BEFORE materializing. Default 10M cells ≈ an 80MB numeric array.
_MAX_CELLS = int(os.environ.get("PAPERCONAN_MAX_CELLS", "10000000"))
# Wide blocks (dense correlation matrices) blow up the O(col²) relation/equal-pair detectors in
# both compute time and output size (scan.json / report.html). Skip just those two detectors when
# a block is wider than this; the cheap column-wise detectors still run. 0 disables the skip.
_MAX_BLOCK_COLS = int(os.environ.get("PAPERCONAN_MAX_BLOCK_COLS", "120"))
# Output cap: each finding embeds a table-snippet as evidence, so a paper with thousands of
# findings balloons scan.json to many GB. Stop collecting blocks once this many have findings.
_MAX_REPORT_BLOCKS = int(os.environ.get("PAPERCONAN_MAX_REPORT_BLOCKS", "2000"))
# Per-finding evidence cap: each finding embeds a sub-rectangle of its block as evidence.
# On a dense matrix a single block can be hundreds of rows × cols, and that copy is duplicated
# across thousands of findings — ballooning the scan dict / scan.json to many GB and OOMing the
# worker. Bound each evidence snippet to a contiguous window of this many rows × cols (always
# including the highlighted cells). Small blocks are emitted whole and stay byte-identical.
_MAX_EV_ROWS = int(os.environ.get("PAPERCONAN_MAX_EVIDENCE_ROWS", "50"))
_MAX_EV_COLS = int(os.environ.get("PAPERCONAN_MAX_EVIDENCE_COLS", "30"))
# Per-block finding cap: the pairwise detectors are O(col²), so a single dense, highly
# correlated block (a correlation matrix, an expression panel with many proportional columns)
# can emit thousands of findings. Each carries its own embedded evidence snippet, so the count —
# not just the per-snippet size — is what balloons scan.json / report.html past 1 GB (GH #15).
# Keep at most this many findings per block, retaining the highest-severity ones, and record how
# many were dropped in the block's `findings_omitted` field (never a silent truncation). 0 disables.
_MAX_FINDINGS_PER_BLOCK = int(os.environ.get("PAPERCONAN_MAX_FINDINGS_PER_BLOCK", "150"))
# Global backstop across all blocks: MAX_REPORT_BLOCKS × MAX_FINDINGS_PER_BLOCK could still be
# large on a pathological corpus, so stop retaining findings once this many have been kept.
_MAX_TOTAL_FINDINGS = int(os.environ.get("PAPERCONAN_MAX_TOTAL_FINDINGS", "5000"))

# Severity rank for deterministic, highest-first truncation when a block is over budget.
_SEVERITY_RANK = {"high": 0, "medium": 1, "low": 2}


def _cap_block_findings(groups, cap):
    """Trim a block's findings to at most `cap`, keeping the highest-severity ones.

    `groups` maps a BLOCK_FINDING_GROUPS key to its list of findings; each list is
    trimmed IN PLACE. Selection is by severity (high > medium > low); ties keep the
    original detector/emission order, so output stays deterministic. Returns the number
    of findings dropped. `cap is None` means unlimited (no trimming); `cap == 0` drops all."""
    if cap is None:
        return 0
    cap = max(0, cap)
    total = sum(len(v) for v in groups.values())
    if total <= cap:
        return 0
    flat = [(name, idx, f)
            for name, lst in groups.items()
            for idx, f in enumerate(lst)]
    # Stable sort by severity keeps original order within a severity band.
    flat.sort(key=lambda t: _SEVERITY_RANK.get((t[2].get("severity") or "low").lower(), 3))
    keep = {(name, idx) for name, idx, _ in flat[:cap]}
    omitted = 0
    for name, lst in groups.items():
        kept = [f for i, f in enumerate(lst) if (name, i) in keep]
        omitted += len(lst) - len(kept)
        lst[:] = kept
    return omitted


def _optional_image_error(prefix, exc):
    detail = " ".join((str(exc) or exc.__class__.__name__).split())
    if len(detail) > 500:
        detail = detail[:497] + "..."
    return {"error": f"{prefix}: {detail}"}


def _elapsed_ms(start):
    """Milliseconds since ``start``, or ``None`` when timing is not recorded.

    ``start`` is ``None`` unless runtime metadata is requested, which keeps the
    wall clock out of the default (byte-reproducible) scan output.
    """
    if start is None:
        return None
    return round((time.perf_counter() - start) * 1000, 3)


def scan_dir(in_dir, out_dir, *, write_md=False, write_html=True, paper=None,
             profile="review", write_json=True, evidence=True, images=False,
             image_diagnostics=False, runtime_metadata=False):
    profile = normalize_profile(profile)
    # The HTML report renders the evidence snippets, so it requires them.
    if write_html:
        evidence = True
    table_files = sorted({
        p for pattern in (
            "*.xlsx", "*.xls", "*.xlsm", "*.xlsb",
            "*.csv", "*.tsv", "*.pdf", "*.docx",
        )
        for p in glob.glob(os.path.join(in_dir, pattern))
    })
    from .fetch._files import is_image
    local_images = sorted(
        (
            path for path in glob.glob(os.path.join(in_dir, "*"))
            if os.path.isfile(path) and is_image(os.path.basename(path))
        ),
        key=lambda path: (os.path.basename(path).casefold(), os.path.basename(path)),
    )
    if not table_files and not (images and local_images):
        supported = ".xlsx / .xls / .xlsm / .xlsb / .csv / .tsv / .pdf / .docx"
        if images:
            supported += " / .png / .jpg / .jpeg / .tif / .tiff / .webp"
        raise PaperconanInputError(
            f"no {supported} files in {in_dir}"
        )

    report_blocks = []
    findings_omitted_total = 0   # findings dropped by the per-block / global finding caps
    findings_kept_total = 0      # findings retained across all blocks (for the global backstop)
    per_sheet_numbers = {}
    grids = {}  # (file, sheet) -> decimal grid, for the unified collision pass
    grid_sheets = {}  # (file, sheet) -> Sheet, for local cross-sheet label context
    scan_errors = []
    scan_stats = {"files": [], "sheets": []}
    scan_start = time.perf_counter() if runtime_metadata else None

    for f in table_files:
        file_start = time.perf_counter() if runtime_metadata else None
        file_stat = {"file": os.path.basename(f), "path": f}
        # Memory guard: a large workbook expands to many GB of Python objects when fully
        # loaded, so cap file size BEFORE loading. Oversized files are recorded (never
        # silently treated as clean) and skipped. Raise PAPERCONAN_MAX_FILE_MB on big-RAM hosts.
        try:
            fsize = os.path.getsize(f)
        except OSError:
            fsize = 0
        if fsize > _MAX_FILE_BYTES:
            msg = (f"oversized: {fsize / 1048576:.1f}MB exceeds {_MAX_FILE_MB:.0f}MB cap "
                   f"(set PAPERCONAN_MAX_FILE_MB to raise) — skipped to bound memory")
            print(f"  skipping {os.path.basename(f)}: {msg}", file=sys.stderr)
            scan_errors.append({"file": os.path.basename(f), "error": msg})
            file_stat["error"] = msg
            file_stat["oversized"] = True
            file_stat["elapsed_ms"] = _elapsed_ms(file_start)
            scan_stats["files"].append(file_stat)
            continue
        try:
            sheets = load_table(f)
        except Exception as e:
            print(f"  failed to read {os.path.basename(f)}: {e}", file=sys.stderr)
            scan_errors.append({"file": os.path.basename(f), "error": str(e)})
            file_stat["error"] = str(e)
            file_stat["elapsed_ms"] = _elapsed_ms(file_start)
            scan_stats["files"].append(file_stat)
            continue
        file_stat["n_sheets"] = len(sheets)
        file_stat["elapsed_ms"] = _elapsed_ms(file_start)
        scan_stats["files"].append(file_stat)
        for sn, rows in sheets.items():
            sheet_start = time.perf_counter() if runtime_metadata else None
            if rows is None:        # oversized sheet (>_MAX_CELLS): recorded, never audited
                msg = (f"oversized sheet exceeds {_MAX_CELLS} cells "
                       f"(set PAPERCONAN_MAX_CELLS to raise) — skipped to bound memory")
                scan_errors.append({"file": os.path.basename(f), "sheet": sn, "error": msg})
                scan_stats["sheets"].append({
                    "file": os.path.basename(f), "sheet": sn, "oversized": True,
                    "elapsed_ms": _elapsed_ms(sheet_start)})
                continue
            sheet = rows if isinstance(rows, Sheet) else Sheet.from_rows(rows)
            grids[(os.path.basename(f), sn)] = _grid_from_rows(sheet)
            grid_sheets[(os.path.basename(f), sn)] = sheet
            sheet_nums = sheet.numeric_values()
            per_sheet_numbers[(os.path.basename(f), sn)] = sheet_nums
            blocks = find_numeric_blocks(sheet)
            max_cols = sheet.ncols
            scan_stats["sheets"].append({
                "file": os.path.basename(f),
                "sheet": sn,
                "n_rows": sheet.nrows,
                "n_cols": max_cols,
                "numeric_cells": len(sheet_nums),
                "n_blocks": len(blocks),
                "elapsed_ms": _elapsed_ms(sheet_start),
            })
            for (r0, r1, c0, c1) in blocks:
                if len(report_blocks) >= _MAX_REPORT_BLOCKS:   # output budget reached; stop collecting
                    break
                if _MAX_TOTAL_FINDINGS > 0 and findings_kept_total >= _MAX_TOTAL_FINDINGS:
                    break   # global finding budget spent; stop collecting (subsequent sheets short-circuit here too)
                header = header_for(sheet, r0, c0, c1)
                # On very wide blocks (dense correlation matrices) the O(col²) relation and
                # equal-pair detectors explode in compute + output, so skip just those two; the
                # column-wise detectors below still run. (_MAX_BLOCK_COLS=0 disables the skip.)
                wide = _MAX_BLOCK_COLS and (c1 - c0) > _MAX_BLOCK_COLS
                rel = [] if wide else detect_relations(sheet, r0, r1, c0, c1, header)
                ap = detect_arithmetic_progression(sheet, r0, r1, c0, c1, header)
                eq = [] if wide else detect_equal_pairs(sheet, r0, r1, c0, c1, header)
                rp = [] if wide else detect_row_pair_digit_coupling(sheet, r0, r1, c0, c1, header)
                # Runs UNCONDITIONALLY (not gated by `wide`): row-oriented condition/measurement
                # layouts are exactly the wide blocks the column detectors skip, and this is
                # where a "row B = row A * k" relationship lives. Self-gates on rows/cols.
                rr = detect_row_relations(sheet, r0, r1, c0, c1, header)
                wc = detect_within_column_patterns(sheet, r0, r1, c0, c1, header)
                wc = wc + detect_dispersed_repeats(sheet, r0, r1, c0, c1, header)
                iar = detect_identical_after_rounding(sheet, r0, r1, c0, c1, header)
                gg = detect_grim_grimmer(sheet, r0, r1, c0, c1, header)
                if rel or ap or eq or rp or rr or wc or iar or gg:
                    sheet_context = " ".join([os.path.basename(f), sn, *[str(h) for h in header]])
                    # Bound this block's finding count BEFORE attaching evidence, so trimmed
                    # findings never pay the (large) embedded-snippet cost. The per-block cap is
                    # further clamped by the remaining global budget; keep highest severity first.
                    # NOTE: this precedes the count-based demotion passes below
                    # (_demote_reused_progressions / _demote_dense_relations / _demote_within_col_flood),
                    # which judge floods by cross-block counts. Trimming can lower those counts, so a
                    # benign finding may keep an elevated severity that demotion would have lowered —
                    # an over-report only (never drops a genuine high; keeping highest-severity first
                    # protects exactly the findings demotion targets). Capping after demotion would
                    # require attaching evidence to every finding first, re-introducing the GH#15 OOM.
                    groups = {"relations": rel, "progressions": ap, "equal_pairs": eq,
                              "row_pairs": rp, "row_relations": rr, "within_col": wc,
                              "identical_after_rounding": iar, "grim": gg}
                    # Effective cap = the tighter of the per-block limit and the remaining global
                    # budget; None means unlimited (both caps disabled). A spent global budget
                    # yields 0, which drops the whole block (recorded via `omitted`).
                    per_block = _MAX_FINDINGS_PER_BLOCK if _MAX_FINDINGS_PER_BLOCK > 0 else None
                    if _MAX_TOTAL_FINDINGS > 0:
                        remaining = max(0, _MAX_TOTAL_FINDINGS - findings_kept_total)
                        block_cap = remaining if per_block is None else min(per_block, remaining)
                    else:
                        block_cap = per_block
                    omitted = _cap_block_findings(groups, block_cap)
                    findings_omitted_total += omitted
                    findings_kept_total += sum(len(v) for v in groups.values())
                    for group in groups.values():
                        if evidence:
                            _attach_evidence(group, sheet, r0, r1, c0, c1, header)
                        _attach_benign(group)
                        apply_profile_to_findings(group, profile,
                                                  sheet_context=sheet_context)
                    report_blocks.append(dict(file=os.path.basename(f), sheet=sn,
                                              block=dict(rows=f"{r0+1}-{r1}", cols=f"{c0+1}-{c1}", header=header),
                                              relations=groups["relations"], progressions=groups["progressions"],
                                              equal_pairs=groups["equal_pairs"],
                                              row_pairs=groups["row_pairs"],
                                              row_relations=groups["row_relations"],
                                              within_col=groups["within_col"],
                                              identical_after_rounding=groups["identical_after_rounding"],
                                              grim=groups["grim"],
                                              findings_omitted=omitted))

    # Down-weight dense/correlated sheets: judged by per-sheet relation totals, so a
    # wide matrix's expected identical/linear columns don't flood high-severity output.
    _demote_dense_sheets(report_blocks)
    _demote_reused_progressions(report_blocks)   # reused perfect progression = re-plotted axis

    # Unified collision pass: every (file, sheet) grid against every other —
    # covers both intra-workbook sheet pairs and cross-file duplicates.
    cross_sheet_findings = detect_collisions(grids, profile=profile, sheets=grid_sheets)
    # B1: full-column duplication across panels, incl. the integer / 1-decimal columns the
    # >=3-decimal collision grids miss.
    cross_sheet_findings += detect_cross_sheet_column_duplicates(grid_sheets, profile=profile)
    # B3: matrix-to-matrix decimal-fraction reuse between two blocks of the same sheet.
    cross_sheet_findings += detect_within_sheet_fraction_reuse(grid_sheets, profile=profile)
    # B2: a fixed high-information row-vector recurring across >=2 figures.
    cross_sheet_findings += detect_recurring_row_vectors(grid_sheets, profile=profile)
    # B6: a condition ROW that is an exact scalar multiple of a row in another block /
    # sheet (cross-block sibling of detect_row_relations; the Extended Data Fig. 5B case).
    cross_sheet_findings += detect_scaled_row_reuse(grid_sheets, profile=profile)
    # B6b: the SHORT high-precision variant — a 3-11 column identical/scaled run between two
    # rows (incl. isolated single-row panels) that the >=12 column detectors above miss. No
    # dedup against the long-run detector is needed: it only emits runs >=12 columns and this
    # one only runs <12, so the two never report the same relation on the same pair.
    cross_sheet_findings += detect_short_row_reuse(grid_sheets, profile=profile)
    # B6c: copy-then-integer-shift WITHIN a row — two cells sharing a long fractional tail
    # with different integer parts (the column- and block-pair shared-fraction detectors
    # never look across the columns of a single row).
    cross_sheet_findings += detect_within_row_shared_fraction(grid_sheets, profile=profile)
    # B6d: the row-PAIR twin of integer_diff_shared_fraction — two rows sharing a decimal
    # fraction at aligned columns with different integer parts (copy-then-shift across rows).
    cross_sheet_findings += detect_row_pair_shared_fraction(grid_sheets, profile=profile)
    _attach_benign(cross_sheet_findings)

    digit_reports, decimal_reports, tail_cluster_reports = [], [], []
    for key, nums in per_sheet_numbers.items():
        d = detect_last_digit(nums, label=f"{key[0]}::{key[1]}")
        if d:
            digit_reports.append(d)
        dec = detect_repeated_decimals(nums, label=f"{key[0]}::{key[1]}")
        if dec:
            decimal_reports.append(dec)
        tc = detect_decimal_tail_clustering(nums, label=f"{key[0]}::{key[1]}")
        if tc:
            tail_cluster_reports.append(tc)

    # Multiple-testing control: dozens of per-sheet χ² tests run at once, so a raw
    # p-threshold over-reports. Attach a BH-adjusted q-value + significance flag.
    if digit_reports:
        adj, sig = benjamini_hochberg([d["p"] for d in digit_reports], alpha=0.05)
        for d, a, s in zip(digit_reports, adj, sig):
            d["p_adj"] = a
            d["fdr_significant"] = bool(s)

    image_assets = []
    image_findings = []
    if images:
        from .image import ImageDependencyError
        from .image._assets import prepare_image_assets
        from .image._budget import ImageArtifactBudget
        try:
            image_budget = ImageArtifactBudget.from_environment()
        except ValueError as exc:
            scan_errors.append({"error": str(exc)})
        else:
            inventory_ready = False
            try:
                from .image._dependencies import preflight_image_dependencies
                preflight_image_dependencies(
                    render_pdf=False,
                    diagnostics=False,
                )
                image_assets, image_errors = prepare_image_assets(
                    in_dir,
                    out_dir,
                    artifact_budget=image_budget,
                )
                scan_errors.extend(image_errors)
                inventory_ready = True
            except ImageDependencyError as exc:
                scan_errors.append(_optional_image_error(
                    "optional image inventory unavailable",
                    exc,
                ))
            except Exception as exc:
                scan_errors.append(_optional_image_error(
                    "optional image inventory unavailable",
                    exc,
                ))
            if image_diagnostics and inventory_ready:
                try:
                    preflight_image_dependencies(
                        render_pdf=False,
                        diagnostics=True,
                    )
                    from .image._diagnostics import diagnose_image_assets
                    image_findings, diagnostic_errors = diagnose_image_assets(
                        image_assets,
                        out_dir,
                        artifact_budget=image_budget,
                    )
                    scan_errors.extend(diagnostic_errors)
                except ImageDependencyError as exc:
                    scan_errors.append(_optional_image_error(
                        "optional image diagnostics unavailable",
                        exc,
                    ))
                except Exception as exc:
                    scan_errors.append(_optional_image_error(
                        "optional image diagnostics unavailable",
                        exc,
                    ))

    out = dict(tool="paperconan",
               tool_version=_version(),
               scanned_at=(
                   datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")
                   if runtime_metadata else None),
               profile=profile,
               input_dir=in_dir,
               paper=_load_provenance(in_dir, paper),
               n_files=len(table_files),
               n_image_source_files=len(local_images),
               n_image_assets=len(image_assets),
               n_blocks_with_findings=len(report_blocks),
               findings_omitted=findings_omitted_total,
               scan_errors=scan_errors,
               scan_stats={**scan_stats,
                           "elapsed_ms": _elapsed_ms(scan_start)},
               relations_blocks=report_blocks,
               digit_distribution=digit_reports,
               decimal_endings=decimal_reports,
               decimal_tail_clusters=tail_cluster_reports,
               cross_sheet_findings=cross_sheet_findings,
               image_assets=image_assets,
               image_findings=image_findings)
    os.makedirs(out_dir, exist_ok=True)
    if write_json:
        with open(os.path.join(out_dir, "scan.json"), "w") as fh:
            json.dump(out, fh, indent=2, default=str)
    if write_md:
        write_markdown_report(out, os.path.join(out_dir, "REPORT.md"))
    if write_html:
        from ._html import write_html_report
        write_html_report(out, os.path.join(out_dir, "report.html"))
    return out


def write_markdown_report(out, path):
    lines = ["# Paper data audit report\n",
             f"- Input: `{out['input_dir']}`",
             f"- Files scanned: {out['n_files']}",
             f"- Blocks with findings: {out['n_blocks_with_findings']}\n"]

    high = []
    medium = []
    def push(b, r):
        sev = r.get("severity", "low")
        row = dict(file=b["file"], sheet=b["sheet"], block_rows=b["block"]["rows"],
                   kind=r["kind"], rule=r.get("rule", ""), n=r.get("n", r.get("n_cells", "?")))
        (high if sev == "high" else medium).append(row)

    for b in out["relations_blocks"]:
        for r in b["relations"]:
            push(b, r)
        for r in b["equal_pairs"]:
            push(b, r)
        for r in b.get("row_pairs", []):
            push(b, r)
        for r in b["progressions"]:
            push(b, r)
        for r in b.get("within_col", []):
            push(b, r)
        for r in b.get("identical_after_rounding", []):
            push(b, r)
        for r in b.get("grim", []):
            push(b, r)

    csf = out.get("cross_sheet_findings", [])
    if csf:
        lines.append(f"## ⚠️ Cross-sheet bit-identical collisions ({len(csf)})\n")
        for cf in csf:
            sev = cf.get("severity", "?")
            lines.append(f"- **[{cf['kind']}]** ({sev}) `{cf['file']}` — {cf['rule']}")
            for ex in cf.get("examples", [])[:3]:
                if isinstance(ex, dict):
                    lines.append(f"    example: row {ex['row']}, col {ex['col']}, value {ex['value']}")
                else:
                    lines.append(f"    shared value: {ex}")
        lines.append("")

    lines.append(f"## High-severity findings ({len(high)})\n")
    for r in high[:40]:
        lines.append(f"- **[{r['kind']}]** `{r['file']}::{r['sheet']}` rows {r['block_rows']}, n={r['n']}  \n  → `{r['rule']}`")
    if len(high) > 40:
        lines.append(f"- … and {len(high) - 40} more (see scan.json)")
    lines.append("")

    lines.append(f"## Medium-severity findings ({len(medium)})\n")
    for r in medium[:30]:
        lines.append(f"- [{r['kind']}] `{r['file']}::{r['sheet']}` rows {r['block_rows']}, n={r['n']} → `{r['rule']}`")
    if len(medium) > 30:
        lines.append(f"- … and {len(medium) - 30} more (see scan.json)")
    lines.append("")

    # last-digit chi-square (BH-FDR-significant, falling back to raw p for old scans)
    def _digit_sig(d):
        return d["fdr_significant"] if "fdr_significant" in d else d["p"] < 1e-6
    sig_digits = sorted([d for d in out["digit_distribution"] if _digit_sig(d)],
                        key=lambda d: d.get("p_adj", d["p"]))
    lines.append(f"## Last-digit χ² anomalies ({len(sig_digits)} sheets, BH-FDR q ≤ 0.05)\n")
    for d in sig_digits[:20]:
        top = ", ".join([f"{k}×{v}" for k, v in d["top"]])
        qv = f" q={d['p_adj']:.1e}" if "p_adj" in d else ""
        lines.append(f"- `{d['label']}` n={d['n']} χ²={d['chi2']:.1f} p={d['p']:.1e}{qv} top: {top}")
    lines.append("")

    # decimal endings
    sig_dec = [d for d in out["decimal_endings"] if d["top"]]
    lines.append(f"## Over-represented two-decimal endings ({len(sig_dec)} sheets)\n")
    for d in sig_dec[:20]:
        top = ", ".join([f".{e}×{c}" for e, c in d["top"][:5]])
        lines.append(f"- `{d['label']}` n={d['n']}, unique={d['n_unique']}, top: {top}")
    lines.append("")

    with open(path, "w") as fh:
        fh.write("\n".join(lines))


def main():
    if len(sys.argv) > 1 and sys.argv[1] == "fetch":
        from .fetch._cli import fetch_main
        sys.exit(fetch_main(sys.argv[2:]))
    if len(sys.argv) > 1 and sys.argv[1] == "report":
        import json
        from ._adjudicated_html import write_adjudicated_report

        rp = argparse.ArgumentParser(
            prog="paperconan report",
            description="Render an adjudicated HTML report from scan.json and verdict.json",
        )
        rp.add_argument("scan_json", help="Path to paperconan scan.json")
        rp.add_argument("--verdict", required=True, help="Path to verdict JSON")
        rp.add_argument("--out", required=True, help="Output HTML path")
        rargs = rp.parse_args(sys.argv[2:])
        try:
            with open(rargs.scan_json, encoding="utf-8") as fh:
                scan = json.load(fh)
            with open(rargs.verdict, encoding="utf-8") as fh:
                verdict = json.load(fh)
            write_adjudicated_report(
                scan,
                verdict,
                rargs.out,
                artifact_dir=os.path.dirname(os.path.abspath(rargs.scan_json)),
            )
        except ValueError as exc:
            sys.exit(str(exc))
        print(f"wrote {rargs.out}")
        return
    ap = argparse.ArgumentParser(
        description=(
            "Scan a paper's source data (xlsx/csv/tsv, or tables inside "
            "pdf/docx) for statistical signals and data inconsistencies"
        )
    )
    ap.add_argument("in_dir", help="Directory with the paper's source data (*.xlsx/*.csv/*.tsv, or *.pdf/*.docx supplements)")
    ap.add_argument("--out", default=None, help="Output directory (default: <in_dir>/audit)")
    ap.add_argument("--md", action="store_true",
                    help="Also write REPORT.md (default: only scan.json + report.html)")
    ap.add_argument("--no-html", action="store_true",
                    help="Skip the HTML report (only scan.json, plus REPORT.md if --md)")
    ap.add_argument("--doi", default=None,
                    help="Record this paper DOI as scan.json provenance "
                         "(overrides any paperconan_source.json sidecar)")
    ap.add_argument("--title", default=None, help="Record this paper title as provenance")
    ap.add_argument("--profile", choices=("review", "forensic", "triage"),
                    default="review",
                    help="False-positive handling profile: review (default), forensic, or triage")
    ap.add_argument(
        "--images",
        action="store_true",
        help="inventory local/fetched images and render PDF pages into scan.json image_assets",
    )
    ap.add_argument(
        "--image-diagnostics",
        action="store_true",
        help="also run optional non-gating deterministic image similarity helpers",
    )
    ap.add_argument(
        "--runtime-metadata",
        action="store_true",
        help="record wall-clock scan timestamp and elapsed times "
             "(omitted by default so scan.json stays byte-reproducible)",
    )
    ap.add_argument("--version", action="version", version=f"paperconan {_version()}")
    args = ap.parse_args()
    if args.image_diagnostics and not args.images:
        ap.error("--image-diagnostics requires --images")
    out_dir = args.out or os.path.join(args.in_dir, "audit")
    write_html = not args.no_html
    paper = None
    if args.doi or args.title:
        paper = {"doi": args.doi, "title": args.title}
    from .image import ImageDependencyError
    try:
        res = scan_dir(args.in_dir, out_dir, write_md=args.md, write_html=write_html,
                       paper=paper, profile=args.profile, images=args.images,
                       image_diagnostics=args.image_diagnostics,
                       runtime_metadata=args.runtime_metadata)
    except PaperconanInputError as e:
        sys.exit(str(e))
    except ImageDependencyError as e:
        sys.exit(str(e))
    outputs = [f"{out_dir}/scan.json"]
    if write_html:
        outputs.append(f"{out_dir}/report.html")
    if args.md:
        outputs.append(f"{out_dir}/REPORT.md")
    print("wrote " + ", ".join(outputs))
    print(f"  files: {res['n_files']}, blocks with findings: {res['n_blocks_with_findings']}")
    print(f"  digit anomaly sheets: {len(res['digit_distribution'])}, decimal anomaly sheets: {len(res['decimal_endings'])}")
    if write_html:
        print(f"\n  → open {out_dir}/report.html in a browser to review findings")


if __name__ == "__main__":
    main()
