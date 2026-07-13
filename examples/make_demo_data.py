"""Generate the synthetic demo dataset for examples/demo_paper/.

This is a *synthetic* paper's source data — every planted pattern below was
planted on purpose so the generated report shows what paperconan can find. No
real paper or person is involved.

Run:
    python examples/make_demo_data.py examples/demo_paper
    paperconan examples/demo_paper --out examples/demo_paper/audit

Each planted pattern is annotated with the detector it is meant to trip. After
generating, the actual findings are whatever paperconan reports — the README
walks through them.
"""
from __future__ import annotations

import os
import sys

import openpyxl


def _round_messy(x: float) -> float:
    """A 'measurement-looking' number with 4 decimals."""
    return round(x, 4)


def _build_tumor(path: str) -> None:
    """ED_Fig2: in-vivo tumor growth. Planted patterns:

    - day            : 0, 3, 6, ... 21  -> arithmetic_progression (a LEGIT time axis;
                        included on purpose to show a benign false-positive)
    - ctrl_volume    : messy increasing measurements
    - treat_volume   : == ctrl_volume + 120.0 exactly  -> constant_offset (high)
    - ctrl_replicate : == ctrl_volume exactly           -> identical_column (high)
    - tumor_length /
      tumor_width    : identical in 7/8 rows            -> many_equal_pairs (medium)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "growth_curve"
    ws.append(["day", "ctrl_volume", "treat_volume", "ctrl_replicate",
               "tumor_length", "tumor_width"])

    ctrl = [98.2156, 142.8923, 201.5347, 268.7412, 339.2588,
            418.6071, 502.4419, 591.8364]
    length = [4.2156, 5.8923, 6.5347, 7.1412, 7.9588, 8.6071, 9.4419, 10.1364]
    for i in range(8):
        day = i * 3
        c = ctrl[i]
        treat = _round_messy(c + 120.0)        # constant_offset
        replicate = c                          # identical_column
        ln = length[i]
        wd = ln if i != 5 else _round_messy(ln + 0.21)  # equal in 7/8 rows
        ws.append([day, c, treat, replicate, ln, wd])

    wb.save(path)


def _build_qpcr(path: str) -> None:
    """ED_Fig4: qPCR relative expression across two 'independent' donors.

    - donor_A / donor_B : most decimal values identical at the SAME (row, col),
                          with two cells tweaked -> cross_sheet_position_identical (high)
    - rel_expr (donor_A): the value 1.0837 repeats in 8/12 rows
                          -> within_col_value_duplication (high)
                          (those repeats also share last-2 decimals '.37'
                           -> within_col_decimal_repetition (high))
    - ct_value (donor_A): every value's last significant digit is 0 or 5
                          -> rounded_to_half_or_int (medium)
    """
    wb = openpyxl.Workbook()

    # --- donor_A ---
    a = wb.active
    a.title = "donor_A"
    a.append(["sample", "rel_expr", "ct_value", "efficiency", "input_ng"])

    rel_expr = [1.0837, 2.4519, 1.0837, 3.1142, 1.0837, 0.8734,
                1.0837, 2.9913, 1.0837, 1.0837, 1.0837, 1.0837]   # 8x 1.0837
    ct_value = [20.0, 22.5, 25.0, 27.5, 30.0, 32.5,
                35.0, 20.5, 25.5, 30.5, 35.5, 22.5]   # last sig digit always 0 or 5
    efficiency = [0.9812, 0.9743, 0.9881, 0.9659, 0.9927, 0.9788,
                  0.9834, 0.9716, 0.9902, 0.9771, 0.9845, 0.9693]
    input_ng = [10.2371, 10.4188, 9.8642, 10.1095, 9.9523, 10.3367,
                10.0814, 9.7459, 10.2956, 9.9138, 10.1772, 10.0285]
    for i in range(12):
        a.append([f"A{i+1}", rel_expr[i], ct_value[i], efficiency[i], input_ng[i]])

    # --- donor_B : copy donor_A's efficiency + input_ng at SAME positions,
    #               tweak two cells (the classic "copy then nudge" fingerprint) ---
    b = wb.create_sheet("donor_B")
    b.append(["sample", "rel_expr", "ct_value", "efficiency", "input_ng"])
    rel_expr_b = [1.5523, 2.1190, 1.7834, 2.6651, 1.4427, 0.9913,
                  2.0148, 1.8876, 1.6692, 2.3318, 1.9905, 1.7741]
    ct_value_b = [23.0, 25.5, 22.5, 24.0, 23.5, 21.5,
                  25.0, 22.0, 24.5, 23.0, 21.0, 25.5]
    for i in range(12):
        eff = efficiency[i]
        ing = input_ng[i]
        if i == 4:
            eff = _round_messy(eff + 0.0007)   # one tweaked cell
        if i == 9:
            ing = _round_messy(ing - 0.0033)   # one tweaked cell
        b.append([f"B{i+1}", rel_expr_b[i], ct_value_b[i], eff, ing])

    wb.save(path)


def build(out_dir: str) -> list[str]:
    os.makedirs(out_dir, exist_ok=True)
    paths = []
    p1 = os.path.join(out_dir, "ED_Fig2_tumor_volume.xlsx")
    p2 = os.path.join(out_dir, "ED_Fig4_qPCR.xlsx")
    _build_tumor(p1)
    _build_qpcr(p2)
    paths.extend([p1, p2])
    return paths


def main(argv: list[str]) -> int:
    out_dir = argv[1] if len(argv) > 1 else "examples/demo_paper"
    paths = build(out_dir)
    for p in paths:
        print(f"wrote {p}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
