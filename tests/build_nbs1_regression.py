"""Build a synthetic, deterministic regression corpus reproducing every anomaly CLASS that a
heavily-duplicated Nature paper's Source Data exhibited (the real .xls are gitignored binaries,
so this reconstructs the patterns in code). Mixes one legacy .xls with .xlsx to also exercise the
calamine read path. Used by tests/test_detection_recall_e2e.py.

Embedded (each maps to a detector):
  - identical_column (HIGH)            : two byte-identical measurement columns in a block
  - cross_sheet_column_duplicate (B1)  : a 1-decimal column repeated across two DIFFERENT figures
  - integer_diff_shared_fraction (B5)  : two cols share high-precision fractions, differ by integers
  - partial_constant_offset (B4)       : a long consecutive run where B = A - 0.3
  - within_table_fraction_reuse (B3)   : two matrix blocks share decimal fractions, differ by ints
  - recurring_row_vector (B2)          : a fixed 6-vector recurring across >=2 figures
  - arithmetic_progression axis (LOW)  : a leftmost 'week' 0.5-step axis column -> demoted to low
"""
from __future__ import annotations

import os
import sys

import openpyxl


VEC = [220.0, 188.0, 122.0, 166.0, 128.0, 166.0]     # recurs across figures (B2)
DUP_COL = [3.0, 3.2, 2.5, 2.8, 2.9, 2.2, 5.0, 5.2, 4.5, 4.8, 4.9, 4.2, 6.1, 6.3]  # 1-dp (B1), len 14


def _weeks(n):
    return [round(0.5 * (i + 1), 4) for i in range(n)]


def _matrix(base_seed):
    return [[round(100.0 - r * 7 - c * 5 + (r * c % 9) * 0.10007 + 0.00051 * (r + c) + base_seed, 5)
             for c in range(7)] for r in range(7)]


def _add(ws, rows):
    for r in rows:
        ws.append(list(r))


def _irregular(i, j, seed):
    """A non-monotone, non-linear value so incidental progression/linear detectors stay quiet."""
    return round(50 + 17 * ((i * 7 + j * 13 + seed) % 11) + 0.5 * ((i * j + seed) % 7) + j, 3)


def _vec_block(seed):
    """A >=3-row numeric block whose FIRST row is VEC; the rest are sheet-distinct so only VEC
    recurs across sheets. VEC starts at column 0 so it sits inside the block."""
    return ([list(VEC)]
            + [[round(seed * 1.3 + 0.31 * j + 1.7 * k, 4) for j in range(len(VEC))] for k in range(1, 4)])


def build(out_dir: str) -> list[str]:
    os.makedirs(out_dir, exist_ok=True)
    paths = []

    # ---- MOESM_A.xlsx ----------------------------------------------------
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Figure 1e"                               # main:1
    # leftmost week axis (0.5 step -> LOW) + tumor columns; a later block carries VEC (B2 site 1)
    _add(ws, [["week", "g1", "g2", "g3", "g4", "g5", "g6"]])
    weeks = _weeks(10)
    for i, wk in enumerate(weeks):
        _add(ws, [[wk] + [_irregular(i, j, 1) for j in range(6)]])
    _add(ws, [[], []])
    _add(ws, _vec_block(11))                             # B2 occurrence (site 1)

    ws2 = wb.create_sheet("Figure 3i")                   # identical_column (HIGH)
    _add(ws2, [["t", "WT_a", "WT_b", "KO_a", "KO_b"]])
    meas = [1.03, 1.09, 1.12, 1.15, 1.19, 1.32, 1.50, 1.45, 1.47, 1.55, 1.52, 1.61, 1.66]
    other = [1.10, 1.14, 1.20, 1.05, 1.31, 1.40, 1.41, 1.52, 1.49, 1.60, 1.58, 1.63, 1.70]
    for i in range(len(meas)):
        _add(ws2, [[i * 50, meas[i], other[i], meas[i], other[i] + 0.02]])  # col1==col3 identical

    ws3 = wb.create_sheet("Figure 3c")                   # B5: shared high-precision fraction + int diff
    _add(ws3, [["rep", "WT", "K388R"]])
    wt = [167.9312, 178.7615, 169.8687, 155.2044, 190.4471, 143.6698, 172.5123, 161.8890]
    for i, v in enumerate(wt):
        _add(ws3, [[i, v, round(v - (58 + (i * 7 % 13)), 4)]])  # K388R = WT - varying integer

    p = os.path.join(out_dir, "MOESM_A.xlsx"); wb.save(p); paths.append(p)

    # ---- MOESM_B.xlsx ----------------------------------------------------
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Figure 4b"                               # main:4 ; B2 site 2
    _add(ws, [["week", "a", "b", "c", "d", "e", "f"]])
    for i, wk in enumerate(_weeks(8)):
        _add(ws, [[wk] + [_irregular(i, j, 2) for j in range(6)]])
    _add(ws, [[], []])
    _add(ws, _vec_block(44))

    ws2 = wb.create_sheet("Figure 4a")                   # B3: two matrix blocks, fraction reuse
    _add(ws2, [["PDO #4"]])
    m4 = _matrix(0.0)
    _add(ws2, m4)
    _add(ws2, [[], [], []])
    _add(ws2, [["PDO #5"]])
    shifts = [[(r * 3 + c * 2) % 17 - 8 for c in range(7)] for r in range(7)]
    m5 = [[round(m4[r][c] + shifts[r][c], 5) for c in range(7)] for r in range(7)]
    _add(ws2, m5)

    ws3 = wb.create_sheet("Figure 1l")                   # B4: partial constant offset run (-0.3)
    _add(ws3, [["n", "Control", "Lactate"]])
    ctrl = [round(1.0 + 0.11 * i + (i * i % 7) * 0.037, 4) for i in range(50)]
    lact = list(ctrl)
    for i in range(30):                                  # 30 consecutive rows: Lactate = Control - 0.3
        lact[i] = round(ctrl[i] - 0.3, 4)
    for i in range(30, 50):                              # rest diverges
        lact[i] = round(ctrl[i] + 0.4 + 0.05 * i, 4)
    for i in range(50):
        _add(ws3, [[i, ctrl[i], lact[i]]])

    p = os.path.join(out_dir, "MOESM_B.xlsx"); wb.save(p); paths.append(p)

    # ---- MOESM_C.xlsx : B1 cross-figure column duplicate (Fig 3b col == ED Fig 9d col) ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Figure 3b"                               # main:3
    _add(ws, [["cell", "No_IR", "IR_05", "IR_8"]])
    for i in range(len(DUP_COL)):
        _add(ws, [[i, DUP_COL[i], round(DUP_COL[i] * 1.7 + 3, 3), round(DUP_COL[i] + 9.4, 3)]])
    ws2 = wb.create_sheet("Extended Data Fig. 9d")       # ext:9  -> different figure namespace
    _add(ws2, [["cell", "siNC", "siLDHA", "other"]])
    for i in range(len(DUP_COL)):
        _add(ws2, [[i, round(DUP_COL[i] * 0.9 + 1.1, 3), DUP_COL[i], round(2 + i * 0.3, 3)]])  # siLDHA == Fig3b No_IR
    # B2 site 3: VEC in a third figure namespace (ext:2)
    ws3 = wb.create_sheet("Extended Data Fig. 2a")       # ext:2
    _add(ws3, [["week", "x", "y", "z", "u", "v", "w"]])
    for i, wk in enumerate(_weeks(8)):
        _add(ws3, [[wk] + [_irregular(i, j, 3) for j in range(6)]])
    _add(ws3, [[], []])
    _add(ws3, _vec_block(77))
    p = os.path.join(out_dir, "MOESM_C.xls"); wb.save(p.replace(".xls", ".xlsx"))
    # rewrite MOESM_C as a legacy .xls via xlwt so the corpus exercises the calamine read path
    try:
        import xlwt
        xwb = xlwt.Workbook()
        for name in wb.sheetnames:
            src = wb[name]
            xs = xwb.add_sheet(name[:31])
            for r, row in enumerate(src.iter_rows(values_only=True)):
                for c, val in enumerate(row):
                    if val is not None:
                        xs.write(r, c, val)
        xwb.save(p)
        os.remove(p.replace(".xls", ".xlsx"))
        paths.append(p)
    except Exception:
        paths.append(p.replace(".xls", ".xlsx"))       # xlwt absent: fall back to .xlsx
    return paths


def main(argv):
    if len(argv) < 2:
        print("usage: python tests/build_nbs1_regression.py <out_dir>", file=sys.stderr)
        return 2
    for p in build(argv[1]):
        print("wrote", p)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
