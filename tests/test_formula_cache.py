"""OOXML formula-cache inspection + its coverage wiring.

A formula cell keeps both the formula (``<f>``) and a cached computed value
(``<v>``). calamine reads the cached value, so a formula cell with no cached
value is silently under-read; the scan records that as a coverage limitation.
"""

import json
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
import pytest

from paperconan._audit import scan_dir
from paperconan._formula_cache import (
    OoxmlFormulaInspectionLimit,
    inspect_ooxml_formula_cache,
)


def _write_xlsx(path, rows, *, sheet_title="Stats"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


_M = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _set_cached_value(path, cell_ref, value):
    """Inject a cached ``<v>`` node into a formula cell (in place).

    openpyxl writes formulas without a computed value; Excel would store one.
    This reproduces the Excel-saved case so we can assert a *cached* formula is
    not flagged.
    """
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        data = {n: zf.read(n) for n in names}
    ET.register_namespace("", _M)
    root = ET.fromstring(data["xl/worksheets/sheet1.xml"])
    for cell in root.iter("{%s}c" % _M):
        if cell.attrib.get("r") == cell_ref:
            # openpyxl emits an empty <v></v> for formulas; populate it (or add
            # one) so the cell carries a real cached value like an Excel save.
            v = cell.find("{%s}v" % _M)
            if v is None:
                v = ET.SubElement(cell, "{%s}v" % _M)
            v.text = str(value)
    data["xl/worksheets/sheet1.xml"] = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:
            zf.writestr(n, data[n])
    return path


def _formula_only_xlsx(tmp_path, name="book.xlsx", *, cell="A3", formula="=A1+A2"):
    path = tmp_path / name
    _write_xlsx(path, [[1.5], [2.5], [None]])
    # openpyxl writes the formula but no cached value, which is exactly the
    # under-read case; set it explicitly then confirm no <v> is emitted.
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws[cell] = formula
    wb.save(path)
    return path


def test_no_formula_sheet_is_not_reported(tmp_path):
    # A plain numeric sheet (no formulas) has nothing to report.
    path = _write_xlsx(tmp_path / "clean.xlsx", [[1.0, 2.0], [3.0, 4.0]])
    assert inspect_ooxml_formula_cache(str(path)) == {}


def test_cached_formula_is_not_reported(tmp_path):
    # A formula that DOES carry a cached value is fully readable → not flagged.
    path = _formula_only_xlsx(tmp_path)
    _set_cached_value(path, "A3", 4.0)
    assert inspect_ooxml_formula_cache(str(path)) == {}


def test_formula_without_cached_value_is_reported(tmp_path):
    path = _formula_only_xlsx(tmp_path)
    gaps = inspect_ooxml_formula_cache(str(path))
    assert gaps == {"Stats": {"count": 1, "cells": ["A3"]}}


def test_non_ooxml_path_returns_empty(tmp_path):
    csv = tmp_path / "data.csv"
    csv.write_text("a,b\n1,2\n")
    assert inspect_ooxml_formula_cache(str(csv)) == {}


def test_examples_are_bounded(tmp_path):
    path = tmp_path / "many.xlsx"
    _write_xlsx(path, [[1], [2]])
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for i in range(3, 9):
        ws[f"A{i}"] = "=A1+A2"
    wb.save(path)
    gaps = inspect_ooxml_formula_cache(str(path), max_examples=3)
    assert gaps["Stats"]["count"] == 6
    assert len(gaps["Stats"]["cells"]) == 3


def test_examples_can_be_disabled(tmp_path):
    path = _formula_only_xlsx(tmp_path)
    gaps = inspect_ooxml_formula_cache(str(path), max_examples=0)
    assert gaps == {"Stats": {"count": 1, "cells": []}}


def test_accepted_sheets_restricts_inspection(tmp_path):
    path = _formula_only_xlsx(tmp_path)
    # Restricting to a sheet that is not the formula sheet reports nothing.
    assert inspect_ooxml_formula_cache(str(path), accepted_sheets={"Other"}) == {}


def test_sheet_limit_is_enforced(tmp_path, monkeypatch):
    monkeypatch.setenv("PAPERCONAN_OOXML_FORMULA_SHEET_LIMIT", "1")
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "One"
    wb.create_sheet("Two")
    wb.save(path)
    with pytest.raises(OoxmlFormulaInspectionLimit) as exc:
        inspect_ooxml_formula_cache(str(path))
    assert exc.value.reason == "formula_metadata_sheet_limit"


# --- end-to-end through scan_dir / coverage -------------------------------


def _scan(tmp_path):
    out = tmp_path / "audit"
    res = scan_dir(str(tmp_path), str(out), write_html=False, write_json=True)
    disk = json.loads((out / "scan.json").read_text())
    return res, disk


def test_formula_gap_marks_scan_partial(tmp_path):
    _formula_only_xlsx(tmp_path, name="book.xlsx")
    res, disk = _scan(tmp_path)
    assert res["scan_status"] == "partial"
    cov = res["coverage"]
    assert cov["files_succeeded"] == 1
    assert cov["files_failed"] == 0
    assert cov["sheets_succeeded"] == 1
    assert cov["sheets_skipped"] == 0
    assert cov["limitations"] == [{
        "scope": "sheet",
        "reason": "formula_cache_missing",
        "file": "book.xlsx",
        "sheet": "Stats",
        "count": 1,
        "cells": ["A3"],
    }]
    assert disk["coverage"] == cov


def test_clean_workbook_stays_complete(tmp_path):
    _write_xlsx(tmp_path / "clean.xlsx", [[1.0, 2.0, 3.0], [4.0, 5.0, 6.0], [7.0, 8.0, 9.0]])
    res, _ = _scan(tmp_path)
    assert res["scan_status"] == "complete"
    assert res["coverage"]["limitations"] == []


def test_inspection_can_be_disabled(tmp_path, monkeypatch):
    monkeypatch.setenv("PAPERCONAN_OOXML_FORMULA_INSPECT", "0")
    import importlib
    import paperconan._audit as audit
    importlib.reload(audit)
    try:
        _formula_only_xlsx(tmp_path, name="book.xlsx")
        out = tmp_path / "audit"
        res = audit.scan_dir(str(tmp_path), str(out), write_html=False, write_json=True)
        assert all(
            l["reason"] != "formula_cache_missing"
            for l in res["coverage"]["limitations"]
        )
    finally:
        monkeypatch.delenv("PAPERCONAN_OOXML_FORMULA_INSPECT", raising=False)
        importlib.reload(audit)


def test_malformed_package_degrades_to_file_limitation(tmp_path, monkeypatch):
    # A .xlsx that loads as a table via one reader but is not a valid zip for the
    # XML pass must degrade to a limitation, never crash the scan.
    import paperconan._audit as audit
    monkeypatch.setattr(
        audit, "inspect_ooxml_formula_cache",
        lambda *a, **k: (_ for _ in ()).throw(zipfile.BadZipFile("boom")),
    )
    cov = audit.ScanCoverage(files_discovered=1)
    audit._record_formula_cache_gaps(cov, "x.xlsx", {"Stats"})
    assert cov.to_dict()["limitations"] == [
        {"scope": "file", "reason": "formula_cache_unreadable", "file": "x.xlsx"}
    ]
